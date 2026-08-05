import io
import os
import re
import json
import base64
from datetime import datetime
import pandas as pd
from nicegui import ui, app
import theme
from database import get_db_connection, get_service_db_connection

THEME = theme.colors

# ═══════════════════════════════════════════════════════════════
# CONSTANTES E FUNÇÕES AUXILIARES GLOBAIS DE IMPRESSÃO (PLACAS JADE)
# ═══════════════════════════════════════════════════════════════

# Insígnias oficiais por Posto/Graduação
RANK_INSIGNIAS = {
    'AE':   {'stars': '★★★★', 'title': 'ALMIRANTE DE ESQUADRA',   'color': '#FFD700'},
    'VA':   {'stars': '★★★',  'title': 'VICE-ALMIRANTE',           'color': '#FFD700'},
    'CA':   {'stars': '★★',   'title': 'CONTRA-ALMIRANTE',         'color': '#FFD700'},
    'CMG':  {'stars': '★',    'title': 'CAPITÃO DE MAR E GUERRA',  'color': '#C0C0C0'},
    'CF':   {'stars': '⚓',   'title': 'CAPITÃO DE FRAGATA',       'color': '#C0C0C0'},
    'CC':   {'stars': '⚓',   'title': 'CAPITÃO DE CORVETA',       'color': '#C0C0C0'},
    'CT':   {'stars': '⚓',   'title': 'CAPITÃO-TENENTE',          'color': '#B0B0B0'},
    '1TEN': {'stars': '▬',    'title': '1º TENENTE',               'color': '#B0B0B0'},
    '2TEN': {'stars': '▬',    'title': '2º TENENTE',               'color': '#B0B0B0'},
    'SO':   {'stars': '◆',    'title': 'SUBOFICIAL',               'color': '#CD7F32'},
    '1SG':  {'stars': '▲▲▲',  'title': '1º SARGENTO',              'color': '#CD7F32'},
    '2SG':  {'stars': '▲▲',   'title': '2º SARGENTO',              'color': '#CD7F32'},
    '3SG':  {'stars': '▲',    'title': '3º SARGENTO',              'color': '#CD7F32'},
    'CB':   {'stars': '∨∨',   'title': 'CABO',                     'color': '#808080'},
    'SD':   {'stars': '∨',    'title': 'SOLDADO',                  'color': '#808080'},
    'MN':   {'stars': '∨',    'title': 'MARINHEIRO',               'color': '#808080'},
    'Dr.':  {'stars': '⚖️',   'title': 'AUTORIDADE CIVIL',         'color': '#4A90D9'},
    'Min.': {'stars': '🏛️',   'title': 'MINISTRO DE ESTADO',       'color': '#9B59B6'},
    'Dep.': {'stars': '🏛️',   'title': 'DEPUTADO',                 'color': '#27AE60'},
    'Sen.': {'stars': '🏛️',   'title': 'SENADOR',                  'color': '#2980B9'},
    'Gen.': {'stars': '★★★★', 'title': 'GENERAL DE EXÉRCITO',      'color': '#FFD700'},
    'Cel.': {'stars': '★',    'title': 'CORONEL',                  'color': '#C0C0C0'},
    'TC':   {'stars': '★',    'title': 'TENENTE-CORONEL',          'color': '#C0C0C0'},
    'Maj':  {'stars': '★',    'title': 'MAJOR',                    'color': '#C0C0C0'},
}

# Brasão padrão CGCFN em SVG compacto (fallback offline quando sem internet)
_BRASAO_FALLBACK_SVG = (
    "data:image/svg+xml;base64,"
    + base64.b64encode(
        b'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 100 100" width="100" height="100">'
        b'<text y="80" font-size="80" text-anchor="middle" x="50">\xe2\x9a\x93</text></svg>'
    ).decode()
)

def gen_qr_base64(data_str: str) -> str:
    """Gera QR Code em Base64 offline (sem internet). Fallback para URL pública se falhar."""
    try:
        import qrcode
        qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=5, border=1)
        qr.add_data(str(data_str))
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        return f"data:image/png;base64,{base64.b64encode(buf.getvalue()).decode('utf-8')}"
    except Exception:
        return f"https://api.qrserver.com/v1/create-qr-code/?size=120x120&data={data_str}&color=000000&bgcolor=ffffff"


def url_to_base64(url: str) -> str:
    """Converte URL de imagem para data:image/...;base64. Fallback ao brasão SVG se falhar."""
    if not url or url.startswith('data:'):
        return url or _BRASAO_FALLBACK_SVG
    if os.path.exists(url):
        try:
            with open(url, 'rb') as f:
                raw = f.read()
            ext = url.rsplit('.', 1)[-1].lower()
            mime = {'png': 'png', 'jpg': 'jpeg', 'jpeg': 'jpeg', 'svg': 'svg+xml', 'gif': 'gif'}.get(ext, 'png')
            return f"data:image/{mime};base64,{base64.b64encode(raw).decode()}"
        except Exception:
            return _BRASAO_FALLBACK_SVG
    try:
        import urllib.request
        with urllib.request.urlopen(url, timeout=3) as resp:
            raw = resp.read()
        ct = resp.headers.get_content_type() or 'image/png'
        return f"data:{ct};base64,{base64.b64encode(raw).decode()}"
    except Exception:
        return _BRASAO_FALLBACK_SVG


def clean_authority_name(raw_name: str) -> str:
    """Remove prefixos de acompanhante e índices numéricos do nome da autoridade."""
    if not raw_name:
        return ""
    name = str(raw_name).strip()
    name = re.sub(r'^(ACOMP\.|ACOMPANHANTE)\s*', '', name, flags=re.IGNORECASE)
    name = re.sub(r'\s*\(\d+(/\d+)?\)$', '', name).strip()
    return name.upper()


def get_rank_logo_asset(posto_str: str):
    """Retorna URL/path da insígnia para o posto, buscando no Supabase Storage ou assets locais."""
    if not posto_str:
        return None
    p = str(posto_str).upper().strip()

    sigla = None
    if any(k in p for k in ['ESQUADRA', 'SQUADRA', 'AE', 'ALMIRANTE DE ESQUADRA']):
        sigla = 'AE'
    elif any(k in p for k in ['VICE', 'VADM', 'V-ADM', 'VA', 'VICE-ALMIRANTE']):
        sigla = 'VA'
    elif any(k in p for k in ['CONTRA', 'CALTE', 'C-ADM', 'CA', 'CONTRA-ALMIRANTE']):
        sigla = 'CA'
    elif any(k in p for k in ['MAR E GUERRA', 'CMG']):
        sigla = 'CMG'
    elif any(k in p for k in ['FRAGATA', 'CF']):
        sigla = 'CF'
    elif any(k in p for k in ['CORVETA', 'CC']):
        sigla = 'CC'
    elif any(k in p for k in ['TENENTE', 'CT']):
        sigla = 'CT'
    else:
        sigla = p.split()[0] if p else None

    if not sigla:
        return None

    sigla_clean = re.sub(r'\W+', '', sigla).upper()

    # 1. Procura no bucket 'logos' do Supabase
    try:
        from database import list_supabase_storage_files
        bucket_files = list_supabase_storage_files("logos")
        for f in bucket_files:
            fname = f.get('name', '')
            fname_no_ext = os.path.splitext(fname)[0].upper()
            if fname_no_ext == sigla_clean:
                return f.get('url')
    except Exception as b_err:
        print(f"[RANK LOGO BUCKET ERR] {b_err}")

    # 2. Procura localmente em assets/insignias/
    for p_name in [f"{sigla_clean.lower()}.png", f"{sigla_clean}.png", f"{sigla_clean.lower()}.jpg"]:
        local_p = os.path.join('assets', 'insignias', p_name)
        if os.path.exists(local_p):
            return local_p

    return None


def parse_almirantado_stars(posto_str: str) -> dict:
    """Retorna dict com estrelas, título, cor e asset PNG para o posto/graduação."""
    if not posto_str:
        return {'eh_almirante': False, 'stars': '', 'title': '', 'color': '#000000', 'png_asset': None}
    p = str(posto_str).upper().strip()
    png_path = get_rank_logo_asset(p)

    if any(k in p for k in ['ESQUADRA', 'SQUADRA', 'AE', 'ALMIRANTE DE ESQUADRA']):
        return {'eh_almirante': True, 'stars': '★ ★ ★ ★', 'title': 'ALMIRANTE DE ESQUADRA', 'color': '#000000', 'png_asset': png_path}
    elif any(k in p for k in ['VICE', 'VADM', 'V-ADM', 'VA', 'VICE-ALMIRANTE']):
        return {'eh_almirante': True, 'stars': '★ ★ ★', 'title': 'VICE-ALMIRANTE', 'color': '#000000', 'png_asset': png_path}
    elif any(k in p for k in ['CONTRA', 'CALTE', 'C-ADM', 'CA', 'CONTRA-ALMIRANTE']):
        return {'eh_almirante': True, 'stars': '★ ★', 'title': 'CONTRA-ALMIRANTE', 'color': '#000000', 'png_asset': png_path}

    return {'eh_almirante': False, 'stars': '', 'title': p, 'color': '#000000', 'png_asset': png_path}



# Estado local do módulo
class ModuleState:
    def __init__(self):
        self.selected_event_id = None
        self.edit_mode = "alocacao"  # "alocacao" ou "layout"
        self.search_query = ""
        self.filter_category = "Todos"
        self.filter_only_unallocated = False
        self.selected_sector = "Todos"  # "Todos" ou nome específico do Setor
        self.zoom_level = "normal"  # "compact", "normal", "large"

state = ModuleState()

def get_row_label(index):
    """Retorna 'A', 'B', ... 'Z', 'AA', etc."""
    label = ""
    while index >= 0:
        label = chr(index % 26 + 65) + label
        index = index // 26 - 1
    return label

def sync_companions(main_guest_id, main_guest_name, max_acomp, event_id, category):
    db = get_service_db_connection() or get_db_connection()
    if not db:
        return
    try:
        # 1. Buscar acompanhantes existentes
        res = db.table('jade_convidados').select('*').eq('convidado_principal_id', main_guest_id).order('id', desc=False).execute()
        existing = res.data if res.data else []
        existing_count = len(existing)
        
        # 2. Se precisamos de mais
        if existing_count < max_acomp:
            needed = max_acomp - existing_count
            new_comps = []
            for i in range(existing_count + 1, max_acomp + 1):
                new_comps.append({
                    'evento_id': event_id,
                    'nome': f"ACOMP. {main_guest_name} ({i}/{max_acomp})",
                    'categoria': category,
                    'convidado_principal_id': main_guest_id,
                    'max_acompanhantes': 0,
                    'cargo_funcao': f"Acompanhante de {main_guest_name}"
                })
            if new_comps:
                db.table('jade_convidados').insert(new_comps).execute()
            
        # 3. Se temos demais (deleta excedentes dando prioridade aos NÃO ALOCADOS)
        elif existing_count > max_acomp:
            needed_delete = existing_count - max_acomp
            # Separa acompanhantes sem assento e com assento
            unallocated = [d for d in existing if not d.get('assento_id')]
            allocated = [d for d in existing if d.get('assento_id')]
            
            # Ordena não alocados decrescente por ID e alocados decrescente por ID
            unallocated_sorted = sorted(unallocated, key=lambda x: x['id'], reverse=True)
            allocated_sorted = sorted(allocated, key=lambda x: x['id'], reverse=True)
            
            # Prioriza deletar os não alocados primeiro
            candidates = unallocated_sorted + allocated_sorted
            to_delete = candidates[:needed_delete]
            delete_ids = [d['id'] for d in to_delete]
            
            if delete_ids:
                # 3.1 Garante que se algum alocado for deletado, a cadeira é liberada no mapa
                for d in to_delete:
                    if d.get('assento_id'):
                        try:
                            db.table('jade_convidados').update({'assento_id': None}).eq('id', d['id']).execute()
                        except Exception:
                            pass
                # 3.2 Deleta os registros excedentes
                db.table('jade_convidados').delete().in_('id', delete_ids).execute()

        # 4. Atualiza obrigatoriamente nomes, categorias e índices de todos os acompanhantes remanescentes
        res_remaining = db.table('jade_convidados').select('*').eq('convidado_principal_id', main_guest_id).order('id', desc=False).execute()
        remaining = res_remaining.data if res_remaining.data else []
        for idx, comp in enumerate(remaining, 1):
            db.table('jade_convidados').update({
                'nome': f"ACOMP. {main_guest_name} ({idx}/{max_acomp})",
                'categoria': category,
                'cargo_funcao': f"Acompanhante de {main_guest_name}"
            }).eq('id', comp['id']).execute()
    except Exception as e:
        print(f"[SYNC COMPANIONS ERR] {e}")


def render_page():
    ui.label('🪑 PROJETAR ASSENTOS (PLACAS JADE)').classes('text-2xl font-bold text-white cyber-title gt-xs q-mb-md q-ml-md')
    
    user_data = app.storage.user.get('user_data', {})
    user_name = user_data.get('nome_guerra', 'Operador')
    
    @ui.refreshable
    def render_content():
        try:
            from database import sync_rsvp_with_jade
            sync_rsvp_with_jade()
        except Exception as sync_err:
            print(f"[RSVP JADE SYNC ERR] {sync_err}")

        db = get_service_db_connection() or get_db_connection()
        if not db:
            with ui.column().classes('w-full items-center justify-center q-py-xl gap-2 text-grey-4'):
                ui.icon('cloud_off', size='4rem')
                ui.label('Banco de dados não disponível. Verifique a conexão.').classes('text-md font-bold')
            return

        # 1. Carregar lista de eventos
        eventos = []
        try:
            res_ev = db.table('jade_eventos').select('*').order('data_evento', desc=True).execute()
            eventos = res_ev.data if res_ev.data else []
        except Exception as e:
            print(f"[JADE EVENTS FETCH ERR] {e}")


        if not state.selected_event_id and eventos:
            state.selected_event_id = eventos[0]['id']

        current_event = next((e for e in eventos if e['id'] == state.selected_event_id), None)

        convidados = []
        layout = {}
        if current_event:
            try:
                res_conv = db.table('jade_convidados').select('*').eq('evento_id', current_event['id']).order('id', desc=False).execute()
                convidados = res_conv.data if res_conv.data else []
            except Exception as e:
                print(f"[JADE GUESTS FETCH ERR] {e}")

            try:
                layout = json.loads(current_event['layout_json']) if current_event.get('layout_json') else {}
            except Exception as e:
                print(f"[LAYOUT PARSE ERR] {e}")

        # --- CABEÇALHO DE CONTROLE DE EVENTOS ---
        with ui.card().classes('w-full q-pa-md no-shadow rounded-xl q-mb-md').style(
            f'background: {THEME["bg_panel"]}; border: 1px solid {THEME["border"]};'
        ):

            # ── Linha 1: Seletor + Gestão do Evento ──
            with ui.row().classes('w-full items-center gap-3 q-mb-sm wrap'):
                ui.label('Solenidade Ativa:').classes('text-xs text-grey-4 font-bold')
                if eventos:
                    event_options = {e['id']: f"{e['nome']} ({e['data_evento']})" for e in eventos}
                    ui.select(
                        options=event_options,
                        value=state.selected_event_id,
                        on_change=lambda e: select_event(e.value)
                    ).props('dark outlined dense').style('min-width: 300px;')
                else:
                    ui.label('Nenhum evento cadastrado.').classes('text-sm text-amber font-bold')

                ui.space()
                # Controles do Evento
                ui.button('＋ Novo Evento', icon='add', on_click=open_create_event_dialog).props('unelevated color=primary text-color=black dense').classes('q-px-sm text-xs')
                if current_event:
                    ui.button('✏️ Editar', icon='edit', on_click=lambda: open_edit_event_dialog(current_event, layout)).props('unelevated color=grey-7 text-color=white dense').classes('q-px-sm text-xs')
                    ui.button('🗑️', icon='delete', on_click=lambda: confirm_delete_event(current_event)).props('unelevated color=negative text-color=white dense').classes('q-px-xs text-xs').tooltip('Excluir Evento')

            if current_event:
                ui.separator().classes('q-my-xs').style('border-color: rgba(255,255,255,0.06);')

                # ── Linha 2: Grupo Placas + Grupo Dados + Grupo Campo ──
                with ui.row().classes('w-full items-center gap-2 wrap'):
                    # GRUPO 1: Impressão de Placas (alta prioridade)
                    ui.badge('PLACAS').props('color=cyan text-color=black').classes('text-[9px] font-bold q-mr-xs')
                    ui.button('⚡ Placa Express', icon='bolt', on_click=lambda: open_express_plate_dialog(current_event, convidados, layout)).props('unelevated color=deep-orange text-color=white dense bold').classes('q-px-sm text-xs').tooltip('Buscar ou criar placa avulsa de emergência (1 clique para impressora)')
                    ui.button('🖨️ Imprimir Placas', icon='print', on_click=lambda: open_print_cards_dialog(current_event, convidados, layout)).props('unelevated color=cyan text-color=black dense bold').classes('q-px-sm text-xs').tooltip('Central de impressão: modelos, brasões, fontes, lote completo')
                    ui.button('✅ Confirmar', icon='how_to_reg', on_click=lambda: open_mass_confirmation_dialog(current_event, convidados)).props('unelevated color=green text-color=white dense bold').classes('q-px-sm text-xs').tooltip('Confirmar Presenças em Massa')

                    ui.separator().props('vertical').classes('q-mx-xs').style('height: 24px; border-color: rgba(255,255,255,0.12);')

                    # GRUPO 2: Dados e Cadastro
                    ui.badge('DADOS').props('color=indigo text-color=white').classes('text-[9px] font-bold q-mr-xs')
                    ui.button('📥 Importar Excel', icon='file_upload', on_click=lambda: open_smart_excel_import_dialog(current_event)).props('unelevated color=deep-purple text-color=white dense bold').classes('q-px-sm text-xs')
                    ui.button('🏛️ Mestre', icon='account_balance', on_click=lambda: open_master_authorities_dialog(current_event)).props('unelevated color=indigo text-color=white dense bold').classes('q-px-sm text-xs').tooltip('Cadastro Mestre de Autoridades')
                    ui.button('📊 Planilhão', icon='table_chart', on_click=lambda: open_event_spreadsheet_dialog(current_event, convidados)).props('unelevated color=teal text-color=white dense bold').classes('q-px-sm text-xs').tooltip('Planilhão completo do evento com exportação CSV')

                    ui.separator().props('vertical').classes('q-mx-xs').style('height: 24px; border-color: rgba(255,255,255,0.12);')


                    # GRUPO 3: Operações de Campo
                    ui.badge('CAMPO').props('color=amber text-color=black').classes('text-[9px] font-bold q-mr-xs')
                    ui.button('🔍 Scanner', icon='qr_code_scanner', on_click=lambda: open_tactical_scanner_dialog(current_event, convidados)).props('unelevated color=amber text-color=black dense bold').classes('q-px-sm text-xs').tooltip('Scanner & Conferência Tática por QR Code')
                    ui.button('📋 Checklist', icon='checklist', on_click=lambda: open_production_checklist_dialog(current_event, convidados)).props('unelevated color=amber-9 text-color=black dense bold').classes('q-px-sm text-xs').tooltip('Checklist de Produção de Placas')
                    ui.button('🎖️ Precedência', icon='verified', on_click=lambda: open_seniority_checklist_dialog(current_event, convidados)).props('unelevated color=blue-grey-7 text-color=white dense bold').classes('q-px-sm text-xs').tooltip('Ordenação por Precedência / Antiguidade')
                    ui.button('📄 Montagem', icon='assignment', on_click=lambda: open_field_assembly_report_dialog(current_event, convidados)).props('unelevated color=light-blue-9 text-color=white dense bold').classes('q-px-sm text-xs').tooltip('Relatório de Montagem de Campo')



        # ═══════════════════════════════════════════════════════════════
        # FASE 1: PAINEL DE FILA DE PRODUÇÃO DE PLACAS JADE
        # ═══════════════════════════════════════════════════════════════
        if current_event and convidados:
            # Contadores de status de placa
            count_pending = sum(1 for c in convidados if c.get('status_placa') == 'pendente')
            count_producing = sum(1 for c in convidados if c.get('status_placa') == 'em_producao')
            count_printed = sum(1 for c in convidados if c.get('status_placa') == 'impressa')
            count_reprint = sum(1 for c in convidados if c.get('status_placa') == 'reimpressao')
            count_delivered = sum(1 for c in convidados if c.get('status_placa') == 'entregue')
            count_not_needed = sum(1 for c in convidados if c.get('status_placa', 'nao_necessaria') == 'nao_necessaria')
            
            # Contadores de confirmação
            count_confirmed = sum(1 for c in convidados if c.get('status_confirmacao') == 'confirmado')
            count_refused = sum(1 for c in convidados if c.get('status_confirmacao') == 'recusado')
            count_probable = sum(1 for c in convidados if c.get('status_confirmacao') == 'provavel')
            count_conf_pending = len(convidados) - count_confirmed - count_refused - count_probable
            
            total_plates_active = count_pending + count_producing + count_reprint

            if total_plates_active > 0 or count_printed > 0 or count_confirmed > 0:
                with ui.card().classes('w-full q-pa-sm no-shadow rounded-xl q-mb-md').style(
                    f'background: linear-gradient(135deg, rgba(0,20,40,0.9) 0%, rgba(0,40,60,0.8) 100%); border: 1px solid rgba(0,229,255,0.3);'
                ):
                    with ui.row().classes('w-full items-center justify-between wrap gap-2'):
                        with ui.row().classes('items-center gap-1'):
                            ui.icon('print', color='cyan').classes('text-lg')
                            ui.label('FILA DE PRODUÇÃO JADE').classes('text-xs font-bold text-cyan tracking-widest')
                        
                        with ui.row().classes('items-center gap-2 wrap'):
                            # Badges de status com cores
                            if count_pending > 0:
                                with ui.badge(f'🟡 {count_pending} Pendentes').props('color=amber text-color=black').classes('text-xs cursor-pointer'):
                                    pass
                            if count_producing > 0:
                                with ui.badge(f'🔵 {count_producing} Em Produção').props('color=blue text-color=white').classes('text-xs'):
                                    pass
                            if count_printed > 0:
                                with ui.badge(f'🟢 {count_printed} Impressas').props('color=green text-color=white').classes('text-xs'):
                                    pass
                            if count_reprint > 0:
                                with ui.badge(f'🔴 {count_reprint} Reimpressão').props('color=red text-color=white').classes('text-xs'):
                                    pass
                            if count_delivered > 0:
                                with ui.badge(f'✅ {count_delivered} Entregues').props('color=teal text-color=white').classes('text-xs'):
                                    pass
                        
                        with ui.row().classes('items-center gap-1'):
                            ui.label(f'👥 {count_confirmed} conf. | {count_conf_pending} pend. | {count_refused} rec.').classes('text-[10px] text-grey-4')
                            
                            # Botões de ação rápida
                            async def mark_pending_as_producing():
                                _db = get_service_db_connection() or get_db_connection()
                                if _db:
                                    for c in convidados:
                                        if c.get('status_placa') == 'pendente':
                                            try:
                                                _db.table('jade_convidados').update({'status_placa': 'em_producao'}).eq('id', c['id']).execute()
                                            except Exception:
                                                pass
                                    ui.notify(f'🔵 {count_pending} placas movidas para "Em Produção"', color='info')
                                    render_content.refresh()
                            
                            async def mark_producing_as_printed():
                                _db = get_service_db_connection() or get_db_connection()
                                if _db:
                                    for c in convidados:
                                        if c.get('status_placa') == 'em_producao':
                                            try:
                                                _db.table('jade_convidados').update({'status_placa': 'impressa'}).eq('id', c['id']).execute()
                                            except Exception:
                                                pass
                                    ui.notify(f'🟢 {count_producing} placas marcadas como "Impressas"', color='success')
                                    render_content.refresh()
                            
                            if count_pending > 0:
                                ui.button('▶ Iniciar Produção', on_click=mark_pending_as_producing).props('unelevated color=blue-8 text-color=white dense').classes('text-[10px] q-px-xs')
                            if count_producing > 0:
                                ui.button('✅ Marcar Impressas', on_click=mark_producing_as_printed).props('unelevated color=green-8 text-color=white dense').classes('text-[10px] q-px-xs')

        if not current_event:
            with ui.column().classes('w-full items-center justify-center q-py-xl gap-4'):
                ui.icon('event_seat', size='5rem', color='cyan')
                ui.label('Por favor, crie um evento para iniciar o mapeamento de assentos.').classes('text-md text-white font-bold')
                ui.button('Criar Primeiro Evento', icon='add', on_click=open_create_event_dialog).props('unelevated color=primary text-color=black')
            return
            
        rows_count = layout.get('rows', 5)
        cols_count = layout.get('cols', 8)
        blocked_seats = layout.get('blocked_seats', [])

        categories = sorted(list(set(c['categoria'] for c in convidados if c.get('categoria'))))
        category_options = ["Todos"] + categories

        allocated_map = {c['assento_id']: c for c in convidados if c.get('assento_id')}

        # Parse dos setores / blocos de assentos
        sectors = layout.get('sectors', [])
        
        # Se existirem setores, garante que o cols_count abranja até a última coluna cadastrada nos setores
        if sectors:
            max_sector_col = max(s.get('end_col', 1) for s in sectors)
            if max_sector_col > cols_count:
                cols_count = max_sector_col
        else:
            if cols_count >= 4:
                sectors = [
                    {'name': 'SETOR ALPHA (ESQUERDA)', 'start_col': 1, 'end_col': max(1, cols_count // 3)},
                    {'name': 'SETOR NOBRE (CENTRO)', 'start_col': max(1, cols_count // 3) + 1, 'end_col': max(1, (cols_count * 2) // 3)},
                    {'name': 'SETOR BRAVO (DIREITA)', 'start_col': max(1, (cols_count * 2) // 3) + 1, 'end_col': cols_count}
                ]

        # Determinar faixa de colunas ativas conforme filtro de setor
        active_start_col = 1
        active_end_col = cols_count

        if getattr(state, 'selected_sector', 'Todos') != "Todos" and sectors:
            target_sec = next((s for s in sectors if s['name'] == state.selected_sector), None)
            if target_sec:
                active_start_col = target_sec['start_col']
                active_end_col = target_sec['end_col']

        display_cols = list(range(active_start_col, active_end_col + 1))
        display_cols_count = len(display_cols)

        # --- CABEÇALHO DO MAPA DE ASSENTOS COM FILTRO DE SETOR ---
        with ui.card().classes('w-full q-pa-md no-shadow rounded-xl q-mb-md').style(
            f'background: {THEME["bg_panel"]}; border: 1px solid {THEME["border"]};'
        ):
            with ui.row().classes('w-full items-center justify-between q-mb-sm wrap-mobile gap-2'):
                with ui.column().classes('gap-0'):
                    ui.label('🗺️ MAPA DE ASSENTOS DA SOLENIDADE').classes('text-md font-bold text-cyan cyber-title')
                    ui.label(f"Grid: {rows_count} fileiras × {cols_count} colunas • Exibindo: {getattr(state, 'selected_sector', 'Todos').upper()} ({display_cols_count} colunas)").classes('text-[11px] text-grey-4')
                
                    with ui.row().classes('items-center gap-2 wrap'):
                        # SELETOR DE SETOR ATIVO NO MESMO EVENTO
                        if sectors:
                            sector_options = {'Todos': '🌐 Todos os Setores'}
                            for s in sectors:
                                sector_options[s['name']] = f"📍 {s['name']}"
                                
                            ui.select(
                                options=sector_options,
                                value=getattr(state, 'selected_sector', 'Todos'),
                                on_change=lambda e: [setattr(state, 'selected_sector', e.value), render_content.refresh()]
                            ).props('dark outlined dense').style('min-width: 200px;').classes('text-xs')

                        # SELETOR DE ZOOM / TAMANHO DOS ASSENTOS (COMPACTO / NORMAL / AMPLO)
                        ui.select(
                            options={'compact': '🔍 Zoom: Compacto', 'normal': '🔍 Zoom: Normal', 'large': '🔍 Zoom: Amplo'},
                            value=getattr(state, 'zoom_level', 'normal'),
                            on_change=lambda e: [setattr(state, 'zoom_level', e.value), render_content.refresh()]
                        ).props('dark outlined dense').style('width: 145px;').classes('text-xs')

                        # Seletor de Modo de Edição
                        with ui.row().classes('items-center bg-black/30 rounded-lg q-pa-xs border border-white/10'):
                            ui.button('Alocação Rápida', icon='event_seat', on_click=lambda: toggle_mode("alocacao")).props(f'dense unelevated {"color=primary text-color=black" if state.edit_mode == "alocacao" else "flat text-color=grey"}').classes('text-xs q-px-sm')
                            ui.button('Editor de Layout / Corredores', icon='edit_road', on_click=lambda: toggle_mode("layout")).props(f'dense unelevated {"color=primary text-color=black" if state.edit_mode == "layout" else "flat text-color=grey"}').classes('text-xs q-px-sm')

            # Dimensões dinâmicas conforme Zoom
            zoom = getattr(state, 'zoom_level', 'normal')
            if zoom == 'compact':
                seat_w, seat_h, font_name, font_sub = '52px', '38px', '8px', '6px'
            elif zoom == 'large':
                seat_w, seat_h, font_name, font_sub = '90px', '58px', '11px', '8px'
            else:
                seat_w, seat_h, font_name, font_sub = '70px', '48px', '9px', '7px'

            # Renderizador de Grid de Assentos por Setores com Rolagem Dupla Completa
            with ui.column().classes('w-full items-start justify-start q-py-md scroll-container').style('overflow-x: auto; overflow-y: auto; max-height: 580px;'):
                ref_top = layout.get('ref_top', 'PALCO PRINCIPAL')
                if ref_top:
                    with ui.row().classes('w-full justify-center q-mb-sm'):
                        ui.label(f"▲ {ref_top.upper()} ▲").classes('text-[10px] font-black tracking-widest text-cyan px-4 py-1 rounded-full border border-cyan-500/20 bg-cyan-500/5')

                # Cálculo de largura mínima necessária para caber TODAS as colunas sem cortar
                num_px = int(seat_w.replace('px',''))
                min_grid_w = max(600, (display_cols_count + 1) * (num_px + 8) + 60)

                # Cores vibrantes e bem visíveis para diferenciar nitidamente os setores do auditório
                sector_colors = [
                    {'bg': 'rgba(0, 229, 255, 0.18)', 'border': 'rgba(0, 229, 255, 0.45)', 'text': '#00e5ff'},   # Setor 1: Cyan Forte
                    {'bg': 'rgba(255, 183, 77, 0.20)', 'border': 'rgba(255, 183, 77, 0.50)', 'text': '#ffb74d'},  # Setor 2: Amber/Dourado Vívido
                    {'bg': 'rgba(171, 71, 188, 0.20)', 'border': 'rgba(171, 71, 188, 0.50)', 'text': '#ab47bc'},  # Setor 3: Roxo Marcante
                    {'bg': 'rgba(102, 187, 106, 0.20)', 'border': 'rgba(102, 187, 106, 0.50)', 'text': '#66bb6a'}, # Setor 4: Verde Intenso
                    {'bg': 'rgba(239, 83, 80, 0.20)', 'border': 'rgba(239, 83, 80, 0.50)', 'text': '#ef5350'}     # Setor 5: Vermelho Vibrante
                ]

                def get_seat_sector_info(col_num, row_num):
                    if not sectors:
                        return {'bg': 'rgba(0, 230, 118, 0.12)', 'border': 'rgba(0, 230, 118, 0.35)', 'text': '#00e676', 'active': True}
                    for idx, sec in enumerate(sectors):
                        if sec['start_col'] <= col_num <= sec['end_col']:
                            sec_rows = sec.get('rows_count', rows_count)
                            is_active = row_num < sec_rows
                            style_info = sector_colors[idx % len(sector_colors)].copy()
                            style_info['active'] = is_active
                            return style_info
                    return {'bg': 'rgba(255, 255, 255, 0.05)', 'border': 'rgba(255, 255, 255, 0.15)', 'text': '#9e9e9e', 'active': True}

                with ui.grid(columns=display_cols_count + 1).classes('gap-2 items-center').style(f'min-width: {min_grid_w}px;'):
                    ui.label('').classes('text-center font-bold text-grey-5').style('width: 40px;')
                    
                    for col in display_cols:
                        ui.label(str(col)).classes('text-center font-bold text-grey-5').style(f'width: {seat_w}; font-size: 11px;')
                        
                    for r in range(rows_count):
                        row_label = get_row_label(r)
                        ui.label(row_label).classes('text-center font-bold text-grey-5 text-md').style('width: 40px;')
                        
                        for col in display_cols:
                            seat_id = f"{row_label}-{col}"
                            is_blocked = seat_id in blocked_seats
                            guest = allocated_map.get(seat_id)
                            sec_info = get_seat_sector_info(col, r)
                            
                            # Se a fileira estiver fora da quantidade de fileiras deste setor específico
                            if not sec_info['active']:
                                ui.label('').style(f'width: {seat_w}; height: {seat_h};')
                                continue

                            if is_blocked:
                                if state.edit_mode == "layout":
                                    with ui.column().classes('items-center justify-center cursor-pointer transition-all hover:scale-105').style(
                                        f'width: {seat_w}; height: {seat_h}; border: 1px dashed rgba(255,255,255,0.15); border-radius: 4px; background: rgba(255,255,255,0.02); gap: 0;'
                                    ).on('click', lambda s=seat_id: toggle_seat_block(s, current_event, layout)):
                                        ui.label(seat_id).classes('text-[8px] text-grey-5 font-mono')
                                        ui.label('CORREDOR').classes('text-[7px] text-grey-6 font-bold')
                                else:
                                    ui.label('').style(f'width: {seat_w}; height: {seat_h};')
                            else:
                                if guest:
                                    display_name = f"{guest.get('posto_graduacao') or ''} {guest['nome']}".strip()
                                    if len(display_name) > 12:
                                        display_name = display_name[:10] + '..'
                                        
                                    is_vip = guest.get('categoria') == 'VIP'
                                    is_acomp = bool(guest.get('convidado_principal_id'))
                                    
                                    border_c = THEME['primary'] if is_vip else ('#ffb74d' if is_acomp else THEME['accent'])
                                    bg_c = 'rgba(0, 229, 255, 0.25)' if is_vip else ('rgba(255, 183, 77, 0.22)' if is_acomp else 'rgba(0, 162, 255, 0.25)')
                                    text_c = THEME['primary'] if is_vip else ('#ffb74d' if is_acomp else THEME['accent'])
                                    
                                    with ui.column().classes('items-center justify-between q-pa-xs cursor-pointer transition-all hover:scale-105 border').style(
                                        f'width: {seat_w}; height: {seat_h}; border-radius: 4px; border-color: {border_c} !important; background: {bg_c}; gap: 0;'
                                    ).on('click', lambda s=seat_id, g=guest: open_seat_actions_dialog(s, g, convidados, current_event['id'])):
                                        ui.label(seat_id).classes('text-[8px] text-grey-4 font-mono leading-none')
                                        ui.label(display_name).classes(f'font-bold text-center leading-none text-white overflow-hidden w-full').style(f'font-size: {font_name};')
                                        
                                        category_label = 'ACOMP' if is_acomp else str(guest.get('categoria', 'Geral')).upper()
                                        if len(category_label) > 10:
                                            category_label = category_label[:8] + '..'
                                        ui.label(category_label).classes(f'text-center leading-none').style(f'color: {text_c}; font-weight: bold; font-size: {font_sub};')
                                else:
                                    if state.edit_mode == "layout":
                                        with ui.column().classes('items-center justify-center cursor-pointer transition-all hover:scale-105 border').style(
                                            f'width: {seat_w}; height: {seat_h}; border-radius: 4px; border-color: rgba(255,255,255,0.15) !important; background: #1b2535; gap: 0;'
                                        ).on('click', lambda s=seat_id: toggle_seat_block(s, current_event, layout)):
                                            ui.label(seat_id).classes('text-[8px] text-grey-4 font-mono')
                                            ui.label('BLOQUEAR').classes('text-[7px] text-grey-5 font-bold')
                                    else:
                                        with ui.column().classes('items-center justify-between q-pa-xs cursor-pointer transition-all hover:scale-105 border').style(
                                            f'width: {seat_w}; height: {seat_h}; border-radius: 4px; border-color: {sec_info["border"]} !important; background: {sec_info["bg"]}; gap: 0;'
                                        ).on('click', lambda s=seat_id: open_allocate_seat_dialog(s, convidados, current_event['id'])):
                                            ui.label(seat_id).classes('text-[8px] text-grey-4 font-mono leading-none')
                                            ui.label('LIVRE').classes('font-bold text-center leading-none').style(f'color: {sec_info["text"]}; font-size: {font_name};')
                                            ui.label('(vazio)').classes('text-grey-4 text-center leading-none').style(f'font-size: {font_sub};')

                ref_bottom = layout.get('ref_bottom', 'ENTRADA / FACHADA')
                if ref_bottom:
                    with ui.row().classes('w-full justify-center q-mt-sm q-mb-sm'):
                        ui.label(f"▼ {ref_bottom.upper()} ▼").classes('text-[10px] font-black tracking-widest text-cyan px-4 py-1 rounded-full border border-cyan-500/20 bg-cyan-500/5')

            # Faixa Informativa de Legenda e Dimensões do Auditório na Base
            with ui.row().classes('w-full justify-between items-center wrap-mobile gap-2 q-mt-sm bg-black/40 q-pa-xs px-3 rounded-lg border border-cyan-500/20'):
                with ui.row().classes('items-center gap-3'):
                    ui.label(f"🏛️ Dimensão Total: {rows_count} Fileiras × {cols_count} Colunas").classes('text-xs text-cyan font-bold')
                    ui.separator().props('vertical').classes('q-my-none').style('height: 16px; border-color: rgba(255,255,255,0.1);')
                    ui.label('💡 Clique nos lugares vagos para alocar convidados. A estrutura de setores é gerida no menu "Editar Evento".').classes('text-[11px] text-grey-4 italic')

        # =========================================================================
        # SEÇÃO 2 (ABAIXO DO GRID): LISTA DE CONVIDADOS HIERÁRQUICA E ACOMPANHANTES
        # =========================================================================
        with ui.card().classes('w-full q-pa-md no-shadow rounded-xl').style(
            f'background: {THEME["bg_panel"]}; border: 1px solid {THEME["border"]};'
        ):
            with ui.row().classes('w-full justify-between items-center wrap-mobile gap-2 q-mb-md'):
                ui.label('👥 PAINEL DE CONVIDADOS E ACOMPANHANTES').classes('text-md font-bold text-cyan cyber-title')
                
                with ui.row().classes('items-center gap-2'):
                    ui.button('Modelo Excel', icon='download', on_click=download_template).props('unelevated color=cyan dense outline').classes('text-xs')

                    # ui.upload com aparência de botão
                    ui.upload(
                        on_upload=lambda e: handle_import_list(e, current_event['id']),
                        multiple=False,
                        auto_upload=True
                    ).props('accept=.xlsx,.xls,.xlsm,.csv,.tsv,.txt,.ods flat color=primary text-color=black dense label="Importar Lista" icon=upload').classes('text-xs')

                    ui.button('➕ Adicionar Convidado Principal', icon='person_add', on_click=lambda: open_edit_guest_dialog(None, current_event['id'])).props('unelevated color=primary text-color=black dense').classes('text-xs')

            # Estado de colapsar tudo / expandir tudo se desejado
            if not hasattr(state, 'filter_seat_status'):
                state.filter_seat_status = "Todos"

            # Filtros de Convidados
            with ui.row().classes('w-full gap-2 items-center q-mb-md wrap-mobile'):
                ui.input(
                    placeholder='Buscar autoridade, convidado ou acompanhante...',
                    on_change=lambda e: update_search(e.value)
                ).props('dark outlined dense clearable').classes('col')
                
                ui.select(
                    options=category_options,
                    value=state.filter_category,
                    on_change=lambda e: update_filter_category(e.value),
                    label='Categoria'
                ).props('dark outlined dense').style('width: 140px;')
                
                ui.select(
                    options={'Todos': 'Status: Todos', 'pendentes': '⏳ Com Pendências', 'completos': '✅ Assentos Completos'},
                    value=getattr(state, 'filter_seat_status', 'Todos'),
                    on_change=lambda e: update_filter_seat_status(e.value)
                ).props('dark outlined dense').style('width: 170px;')

            # Filtra autoridades principais (sem convidado_principal_id)
            principais = [c for c in convidados if not c.get('convidado_principal_id')]
            
            # Filtro por busca de texto
            if state.search_query:
                q = state.search_query.lower()
                filtered_principais = []
                for p in principais:
                    acomp_p = [c for c in convidados if c.get('convidado_principal_id') == p['id']]
                    match_p = q in p['nome'].lower() or (p.get('cargo_funcao') and q in p['cargo_funcao'].lower()) or (p.get('posto_graduacao') and q in p['posto_graduacao'].lower())
                    match_ac = any(q in a['nome'].lower() for a in acomp_p)
                    if match_p or match_ac:
                        filtered_principais.append(p)
                principais = filtered_principais

            if state.filter_category != "Todos":
                principais = [p for p in principais if p.get('categoria') == state.filter_category]

            # Filtro por status de alocação (Pendente x Completo)
            status_filter = getattr(state, 'filter_seat_status', 'Todos')
            if status_filter == 'pendentes':
                # Tem pendência se a própria autoridade ou algum acompanhante não tem assento
                principais = [
                    p for p in principais 
                    if not p.get('assento_id') or any(not c.get('assento_id') for c in convidados if c.get('convidado_principal_id') == p['id'])
                ]
            elif status_filter == 'completos':
                # 100% alocado se a autoridade E todos os seus acompanhantes têm assento
                principais = [
                    p for p in principais 
                    if p.get('assento_id') and all(c.get('assento_id') for c in convidados if c.get('convidado_principal_id') == p['id'])
                ]

            if principais:
                with ui.column().classes('w-full gap-3 scroll-container q-pr-xs').style('max-height: 680px; overflow-y: auto;'):
                    with ui.grid(columns='1 md:grid-cols-2 lg:grid-cols-3').classes('w-full gap-4'):
                        for p in principais:
                            acomp_list = [c for c in convidados if c.get('convidado_principal_id') == p['id']]
                            
                            is_p_allocated = bool(p.get('assento_id'))
                            all_acomp_allocated = len(acomp_list) > 0 and all(bool(ac.get('assento_id')) for ac in acomp_list)
                            is_group_complete = is_p_allocated and (len(acomp_list) == 0 or all_acomp_allocated)

                            card_border = 'rgba(0, 230, 118, 0.4)' if is_group_complete else ('rgba(0, 229, 255, 0.4)' if is_p_allocated else 'rgba(255, 255, 255, 0.08)')
                            header_icon = '✅' if is_group_complete else ('⏳' if is_p_allocated else '⚠️')
                            
                            nome_p = f"{p.get('posto_graduacao') or ''} {p['nome']}".strip()
                            cargo_p = p.get('cargo_funcao') or p.get('categoria') or 'Autoridade'
                            
                            with ui.expansion().classes('w-full rounded-xl border no-shadow').style(
                                f'background: rgba(19, 26, 38, 0.95); border-color: {card_border} !important;'
                            ) as exp:
                                with exp.add_slot('header'):
                                    with ui.row().classes('w-full justify-between items-center no-wrap gap-2'):
                                        with ui.column().classes('gap-0 col'):
                                            ui.label(f"{header_icon} {nome_p}").classes('text-sm font-bold text-white cyber-title')
                                            ui.label(f"[{p.get('categoria', 'Geral')}] {cargo_p}").classes('text-xs text-grey-4')

                                        with ui.row().classes('items-center gap-1'):
                                            if is_p_allocated:
                                                ui.badge(f"Assento {p['assento_id']}").props('color=cyan text-color=black bold').classes('text-[10px]')
                                            else:
                                                ui.badge('Sem Assento').props('color=grey-8').classes('text-[10px]')

                                # --- CONTEÚDO DO CARD COLAPSÁVEL ---
                                with ui.column().classes('w-full q-pa-sm gap-2'):
                                    # Ações da Autoridade
                                    with ui.row().classes('w-full justify-between items-center bg-black/30 q-pa-xs rounded-lg'):
                                        ui.label('Ações da Autoridade:').classes('text-[11px] text-grey-4 font-bold')
                                        with ui.row().classes('items-center gap-1'):
                                            if is_p_allocated:
                                                ui.button('Desalocar', icon='cancel', on_click=lambda p=p: remove_guest_allocation(p)).props('unelevated color=danger dense flat').classes('text-xs')
                                            ui.button('Editar', icon='edit', on_click=lambda p=p: open_edit_guest_dialog(p)).props('unelevated color=primary dense flat').classes('text-xs')
                                            ui.button('Excluir', icon='delete', on_click=lambda p=p: confirm_delete_guest(p)).props('unelevated color=danger dense flat').classes('text-xs')

                                    # Controle Quantitativo (+ / -)
                                    max_ac = p.get('max_acompanhantes', 0)
                                    with ui.row().classes('w-full justify-between items-center q-py-xs bg-black/20 px-2 rounded-lg'):
                                        ui.label(f"Acompanhantes Vagas: {max_ac}").classes('text-xs text-amber font-bold')
                                        
                                        with ui.row().classes('items-center gap-1'):
                                            if max_ac > 0:
                                                def dec_acomp(p_ref=p):
                                                    new_ac = max(0, p_ref.get('max_acompanhantes', 0) - 1)
                                                    reg = {
                                                        'nome': p_ref['nome'],
                                                        'posto_graduacao': p_ref.get('posto_graduacao'),
                                                        'cargo_funcao': p_ref.get('cargo_funcao'),
                                                        'categoria': p_ref.get('categoria', 'Geral'),
                                                        'max_acompanhantes': new_ac
                                                    }
                                                    save_guest(p_ref['id'], reg, current_event['id'])
                                                
                                                ui.button('-', on_click=dec_acomp).props('unelevated color=amber text-color=black dense round').style('width: 22px; height: 22px; font-weight: bold;')

                                            def inc_acomp(p_ref=p):
                                                new_ac = p_ref.get('max_acompanhantes', 0) + 1
                                                reg = {
                                                    'nome': p_ref['nome'],
                                                    'posto_graduacao': p_ref.get('posto_graduacao'),
                                                    'cargo_funcao': p_ref.get('cargo_funcao'),
                                                    'categoria': p_ref.get('categoria', 'Geral'),
                                                    'max_acompanhantes': new_ac
                                                }
                                                save_guest(p_ref['id'], reg, current_event['id'])

                                            ui.button('+', on_click=inc_acomp).props('unelevated color=amber text-color=black dense round').style('width: 22px; height: 22px; font-weight: bold;')

                                    # Lista de Acompanhantes
                                    if acomp_list:
                                        with ui.column().classes('w-full gap-1 q-mt-xs pl-2 border-l-2 border-amber-500/40'):
                                            for ac in acomp_list:
                                                is_ac_allocated = bool(ac.get('assento_id'))
                                                ac_bg = 'rgba(255, 183, 77, 0.08)' if is_ac_allocated else 'rgba(255, 255, 255, 0.02)'
                                                
                                                with ui.card().classes('w-full q-pa-xs px-2 no-shadow rounded-md').style(f'background: {ac_bg}; border: 1px solid rgba(255,183,77,0.2);'):
                                                    with ui.row().classes('w-full justify-between items-center no-wrap'):
                                                        with ui.column().classes('gap-0'):
                                                            ui.label(ac['nome']).classes('text-xs text-grey-2 font-medium')
                                                            ui.label('(Acompanhante)').classes('text-[9px] text-amber-4 italic')
                                                        
                                                        if is_ac_allocated:
                                                            with ui.row().classes('items-center gap-1'):
                                                                ui.badge(f"Assento {ac['assento_id']}").props('color=amber text-color=black').classes('text-[9px]')
                                                                ui.button(icon='cancel', on_click=lambda ac=ac: remove_guest_allocation(ac)).props('unelevated color=danger dense flat round').classes('text-xs')
                                                        else:
                                                            ui.badge('Sem Assento').props('color=grey-8').classes('text-[9px]')
            else:
                with ui.column().classes('w-full items-center justify-center q-py-xl text-grey'):
                    ui.icon('person_off', size='3rem')
                    ui.label('Nenhuma autoridade ou convidado encontrado.').classes('text-sm q-mt-xs')

    # Métodos utilitários do painel
    def select_event(event_id):
        state.selected_event_id = event_id
        render_content.refresh()

    def toggle_mode(mode):
        state.edit_mode = mode
        render_content.refresh()

    def update_search(query):
        state.search_query = query
        render_content.refresh()

    def update_filter_category(cat):
        state.filter_category = cat
        render_content.refresh()

    def update_filter_unallocated(val):
        state.filter_only_unallocated = val
        render_content.refresh()

    def update_filter_seat_status(val):
        state.filter_seat_status = val
        render_content.refresh()

    def create_event(nome, data, local, layout_tipo, rows, cols, ref_top, ref_bottom, custom_sectors=None):
        if not nome or not data:
            ui.notify('Nome e Data do evento são obrigatórios.', color='warning')
            return
            
        db = get_service_db_connection() or get_db_connection()
        if db:
            layout_data = {
                'tipo': layout_tipo,
                'rows': int(rows),
                'cols': int(cols),
                'ref_top': ref_top or 'PALCO PRINCIPAL',
                'ref_bottom': ref_bottom or 'ENTRADA / FACHADA',
                'blocked_seats': [],
                'sectors': custom_sectors or []
            }
            try:
                res = db.table('jade_eventos').insert({
                    'nome': nome.upper(),
                    'data_evento': data,
                    'local': local or '',
                    'layout_json': json.dumps(layout_data)
                }).execute()
                
                ui.notify('Evento criado com sucesso!', color='success')
                if res.data:
                    state.selected_event_id = res.data[0]['id']
                render_content.refresh()
            except Exception as e:
                ui.notify(f"Erro ao criar evento: {e}", color='red')

    def delete_event(event):
        db = get_service_db_connection() or get_db_connection()
        if db:
            try:
                db.table('jade_convidados').delete().eq('evento_id', event['id']).execute()
                db.table('jade_eventos').delete().eq('id', event['id']).execute()
                ui.notify('Evento e dados associados excluídos.', color='success')
                state.selected_event_id = None
                render_content.refresh()
            except Exception as e:
                ui.notify(f"Erro ao excluir evento: {e}", color='red')

    def save_guest(guest_id, data, event_id):
        db = get_service_db_connection() or get_db_connection()
        if db:
            try:
                if guest_id:
                    db.table('jade_convidados').update(data).eq('id', guest_id).execute()
                    ui.notify('Dados do convidado atualizados.', color='success')
                    sync_companions(guest_id, data['nome'], data['max_acompanhantes'], event_id, data['categoria'])
                else:
                    data['evento_id'] = event_id
                    res = db.table('jade_convidados').insert(data).execute()
                    ui.notify('Convidado adicionado à lista.', color='success')
                    if res.data:
                        new_id = res.data[0]['id']
                        sync_companions(new_id, data['nome'], data['max_acompanhantes'], event_id, data['categoria'])
                render_content.refresh()
            except Exception as e:
                ui.notify(f"Erro ao salvar convidado: {e}", color='red')

    def delete_guest(guest):
        db = get_service_db_connection() or get_db_connection()
        if db:
            try:
                db.table('jade_convidados').delete().eq('convidado_principal_id', guest['id']).execute()
                db.table('jade_convidados').delete().eq('id', guest['id']).execute()
                ui.notify(f"Convidado {guest['nome']} e acompanhantes removidos.", color='success')
                render_content.refresh()
            except Exception as e:
                ui.notify(f"Erro ao excluir convidado: {e}", color='red')

    def allocate_guest(guest_id, seat_id, event_id):
        db = get_service_db_connection() or get_db_connection()
        if db:
            try:
                db.table('jade_convidados').update({'assento_id': None}).eq('evento_id', event_id).eq('assento_id', seat_id).execute()
                db.table('jade_convidados').update({'assento_id': seat_id}).eq('id', guest_id).execute()
                ui.notify(f"Assento {seat_id} ocupado com sucesso.", color='success')
                render_content.refresh()
            except Exception as e:
                ui.notify(f"Erro ao alocar assento: {e}", color='red')

    def remove_guest_allocation(guest):
        db = get_service_db_connection() or get_db_connection()
        if db:
            try:
                db.table('jade_convidados').update({'assento_id': None}).eq('id', guest['id']).execute()
                ui.notify(f"Convidado {guest['nome']} removido do assento {guest['assento_id']}.", color='success')
                render_content.refresh()
            except Exception as e:
                ui.notify(f"Erro ao remover alocação: {e}", color='red')

    def swap_guests(guest_a, guest_b_id, event_id):
        db = get_service_db_connection() or get_db_connection()
        if db:
            try:
                seat_a = guest_a.get('assento_id')
                res_b = db.table('jade_convidados').select('*').eq('id', guest_b_id).execute()
                if not res_b.data:
                    return
                guest_b = res_b.data[0]
                seat_b = guest_b.get('assento_id')
                
                db.table('jade_convidados').update({'assento_id': seat_b}).eq('id', guest_a['id']).execute()
                db.table('jade_convidados').update({'assento_id': seat_a}).eq('id', guest_b['id']).execute()
                
                ui.notify("Troca de assentos efetuada.", color='success')
                render_content.refresh()
            except Exception as e:
                ui.notify(f"Erro ao realizar troca: {e}", color='red')

    def toggle_seat_block(seat_id, event, layout):
        db = get_service_db_connection() or get_db_connection()
        if db:
            blocked = layout.get('blocked_seats', [])
            if seat_id in blocked:
                blocked.remove(seat_id)
            else:
                blocked.append(seat_id)
                
            layout['blocked_seats'] = blocked
            new_layout_json = json.dumps(layout)
            
            try:
                db.table('jade_convidados').update({'assento_id': None}).eq('evento_id', event['id']).eq('assento_id', seat_id).execute()
                db.table('jade_eventos').update({'layout_json': new_layout_json}).eq('id', event['id']).execute()
                render_content.refresh()
            except Exception as e:
                ui.notify(f"Erro ao atualizar layout: {e}", color='red')

    def update_grid_size(event, layout, row_delta, col_delta):
        db = get_service_db_connection() or get_db_connection()
        if db:
            r = max(1, min(50, layout.get('rows', 5) + row_delta))
            c = max(1, min(100, layout.get('cols', 8) + col_delta))
            
            layout['rows'] = r
            layout['cols'] = c
            
            # Se colunas foram alteradas, reajusta o limite do último setor para cobrir o grid
            sectors = layout.get('sectors', [])
            if sectors and col_delta != 0:
                sectors[-1]['end_col'] = max(sectors[-1]['start_col'], c)
                layout['sectors'] = sectors

            new_layout_json = json.dumps(layout)
            try:
                db.table('jade_eventos').update({'layout_json': new_layout_json}).eq('id', event['id']).execute()
                ui.notify(f"Grid ajustado para {r} fileiras × {c} colunas.", color='success')
                render_content.refresh()
            except Exception as e:
                ui.notify(f"Erro ao alterar dimensões do grid: {e}", color='red')

    def open_production_checklist_dialog(event, convidados):
        """Checklist de produção independente de mapa de assentos — lista TODOS os confirmados."""
        
        RANK_W = {
            'AE': 1, 'VA': 2, 'CA': 3, 'CMG': 4, 'CF': 5, 'CC': 6, 'CT': 7,
            '1TEN': 8, '2TEN': 9, 'GM': 10, 'SO': 11, '1SG': 12, '2SG': 13,
            '3SG': 14, 'CB': 15, 'SD': 16, 'MN': 16,
        }
        def rank_key(c):
            ant = c.get('numero_antiguidade')
            if ant is not None:
                try:
                    return (0, int(ant), c['nome'])
                except:
                    pass
            p = (c.get('posto_graduacao') or '').upper().replace('.','').replace(' ','').strip()
            return (1, RANK_W.get(p, 90), c['nome'])

        # Status placa legível
        STATUS_LABELS = {
            'pendente': ('🟡 PENDENTE', 'amber-9'),
            'em_producao': ('🔵 EM PRODUÇÃO', 'blue-9'),
            'impressa': ('🟢 IMPRESSA', 'green-9'),
            'reimpressao': ('🔴 REIMPRESSÃO', 'red-9'),
            'entregue': ('✅ ENTREGUE', 'teal-9'),
            'nao_necessaria': ('⬜ N/A', 'grey-7'),
        }

        filter_state = {'mode': 'all'}

        with ui.dialog() as diag, ui.card().classes('q-pa-lg').style('min-width: 850px; max-width: 96vw; max-height: 92vh; overflow-y: auto;'):
            with ui.row().classes('w-full justify-between items-center q-mb-sm no-wrap'):
                with ui.row().classes('items-center gap-2'):
                    ui.icon('checklist', color='amber-5', size='md')
                    with ui.column().classes('gap-0'):
                        ui.label('CHECKLIST DE PRODUÇÃO DE PLACAS').classes('text-md font-bold text-white cyber-title')
                        ui.label(f"Solenidade: {event.get('nome','N/I')} | {len([c for c in convidados if not c.get('convidado_principal_id')])} convidados principais").classes('text-[10px] text-grey-4')
                ui.button(icon='close', on_click=diag.close).props('flat round dense color=grey-4')

            # Filtros rápidos
            with ui.row().classes('w-full gap-2 q-mb-sm wrap'):
                def make_filter_btn(label, mode, color='grey-8'):
                    def click():
                        filter_state['mode'] = mode
                        render_list.refresh()
                    ui.button(label, on_click=click).props(f'unelevated color={color} dense').classes('text-xs')
                make_filter_btn('📋 Todos', 'all', 'grey-8')
                make_filter_btn('🟡 Pendentes de Impressão', 'pending', 'amber-9')
                make_filter_btn('🔵 Em Produção', 'producing', 'blue-9')
                make_filter_btn('🟢 Impressas', 'printed', 'green-9')
                make_filter_btn('📍 Sem Assento Alocado', 'no_seat', 'deep-purple-9')

            # Ações em massa
            with ui.row().classes('w-full gap-2 q-mb-md wrap items-center border-b border-white/10 q-pb-sm'):
                ui.label('Ações em massa:').classes('text-xs text-grey-4 font-bold')
                async def mass_to_production():
                    db = get_service_db_connection() or get_db_connection()
                    if db:
                        to_upd = [c for c in convidados if c.get('status_placa') == 'pendente']
                        for c in to_upd:
                            try:
                                db.table('jade_convidados').update({'status_placa': 'em_producao'}).eq('id', c['id']).execute()
                            except:
                                pass
                        ui.notify(f'🔵 {len(to_upd)} placas enviadas para produção!', color='info')
                        render_content.refresh()
                        render_list.refresh()
                async def mass_mark_printed():
                    db = get_service_db_connection() or get_db_connection()
                    if db:
                        to_upd = [c for c in convidados if c.get('status_placa') == 'em_producao']
                        for c in to_upd:
                            try:
                                db.table('jade_convidados').update({'status_placa': 'impressa'}).eq('id', c['id']).execute()
                            except:
                                pass
                        ui.notify(f'🟢 {len(to_upd)} placas marcadas como impressas!', color='positive')
                        render_content.refresh()
                        render_list.refresh()
                ui.button('▶ Enviar Pendentes → Produção', on_click=mass_to_production).props('unelevated color=blue-9 dense').classes('text-xs')
                ui.button('✅ Marcar Em Prod. → Impressas', on_click=mass_mark_printed).props('unelevated color=green-9 dense').classes('text-xs')

            # Lista principal
            @ui.refreshable
            def render_list():
                mode = filter_state['mode']
                # Filtrar e ordenar
                filtered = []
                for c in convidados:
                    sp = c.get('status_placa', 'nao_necessaria')
                    sc = c.get('status_confirmacao', 'pendente')
                    if mode == 'all' and sp != 'nao_necessaria':
                        filtered.append(c)
                    elif mode == 'pending' and sp == 'pendente':
                        filtered.append(c)
                    elif mode == 'producing' and sp == 'em_producao':
                        filtered.append(c)
                    elif mode == 'printed' and sp in ('impressa', 'entregue'):
                        filtered.append(c)
                    elif mode == 'no_seat' and not c.get('assento_id') and sp != 'nao_necessaria':
                        filtered.append(c)

                filtered.sort(key=rank_key)

                if not filtered:
                    with ui.column().classes('w-full items-center justify-center q-py-xl text-grey-5 gap-2'):
                        ui.icon('inventory_2', size='2.5rem')
                        ui.label('Nenhum item neste filtro.').classes('text-xs')
                    return

                with ui.column().classes('w-full gap-1 max-h-[440px] overflow-y-auto q-pr-xs'):
                    for idx, c in enumerate(filtered, 1):
                        sp = c.get('status_placa', 'nao_necessaria')
                        is_acomp = bool(c.get('convidado_principal_id'))
                        placa_label, placa_color = STATUS_LABELS.get(sp, ('❓', 'grey'))
                        assento = c.get('assento_id') or None
                        ant_num = c.get('numero_antiguidade', '')

                        bg_col = 'rgba(0,229,255,0.03)' if not is_acomp else 'rgba(255,255,255,0.01)'
                        border_col = '#00e5ff' if not is_acomp else '#4b5563'

                        with ui.row().classes('w-full items-center no-wrap q-pa-sm rounded-lg gap-2').style(
                            f'background:{bg_col}; border-left:3px solid {border_col}; border-top:1px solid rgba(255,255,255,0.03);'
                        ):
                            # Nº / Antiguidade
                            ui.label(f"{ant_num or idx:02}").classes('text-[9px] font-mono text-grey-5 shrink-0').style('min-width:22px;')

                            # Nome e posto
                            with ui.column().classes('gap-0 col no-wrap').style('min-width:0;'):
                                nome_display = f"{c.get('posto_graduacao') or ''} {c['nome']}".strip().upper()
                                ui.label(nome_display).classes('text-[11px] font-bold text-white truncate')
                                sub = '(Acompanhante)' if is_acomp else (c.get('cargo_funcao') or c.get('categoria') or '')
                                if sub:
                                    ui.label(sub).classes('text-[8px] text-grey-5 truncate')

                            # Assento
                            if assento:
                                ui.badge(f"ASS. {assento}", color='cyan-9').classes('text-[8px] shrink-0')
                            else:
                                ui.badge('SEM ASSENTO', color='deep-purple-9').classes('text-[8px] shrink-0')

                            # Status badge
                            ui.badge(placa_label, color=placa_color).classes('text-[8px] shrink-0')

                            # Botões de transição rápida
                            def make_status_btn(guest_id, new_status, label, color):
                                async def click():
                                    _db = get_service_db_connection() or get_db_connection()
                                    if _db:
                                        _db.table('jade_convidados').update({'status_placa': new_status}).eq('id', guest_id).execute()
                                        ui.notify(f'{label}', color='positive', timeout=1500)
                                        render_content.refresh()
                                        render_list.refresh()
                                return click

                            if sp == 'pendente':
                                ui.button('▶ Produção', on_click=make_status_btn(c['id'], 'em_producao', '🔵 Enviado para produção', 'blue')).props('unelevated color=blue-9 dense').classes('text-[9px] shrink-0')
                            elif sp == 'em_producao':
                                ui.button('✅ Impressa', on_click=make_status_btn(c['id'], 'impressa', '🟢 Marcada como impressa', 'green')).props('unelevated color=green-9 dense').classes('text-[9px] shrink-0')
                            elif sp == 'impressa':
                                ui.button('📍 Posicionada', on_click=make_status_btn(c['id'], 'entregue', '✅ Placa posicionada no assento!', 'teal')).props('unelevated color=teal-9 dense').classes('text-[9px] shrink-0')
                            elif sp == 'entregue':
                                ui.label('✅ POSICIONADA').classes('text-[9px] text-teal-4 font-black shrink-0')
                            
                            # Reimpressão para qualquer placa
                            ui.button('🔁 Reimprimir', on_click=make_status_btn(c['id'], 'reimpressao', '🔴 Solicitada reimpressão', 'red')).props('unelevated color=deep-orange dense').classes('text-[9px] shrink-0').tooltip('Solicitar nova impressão desta placa')

                            # Aumentar Quantidade / Criar Acompanhante
                            if not is_acomp:
                                async def add_companion_plate(principal_id, main_nome, main_posto):
                                    db = get_service_db_connection() or get_db_connection()
                                    if db:
                                        res_ac = db.table('jade_convidados').select('*').eq('convidado_principal_id', principal_id).execute()
                                        ac_list = res_ac.data or []
                                        count_ac = len(ac_list) + 1
                                        db.table('jade_convidados').insert({
                                            'evento_id': event['id'],
                                            'nome': f"ACOMP. {main_nome} ({count_ac})",
                                            'posto_graduacao': main_posto,
                                            'convidado_principal_id': principal_id,
                                            'status_placa': 'pendente',
                                            'presenca_confirmada': True
                                        }).execute()
                                        ui.notify(f'➕ Nova placa de acompanhante registrada para {main_nome}!', color='positive')
                                        render_content.refresh()
                                        render_list.refresh()

                                ui.button('➕ Placa Extra', on_click=lambda p_id=c['id'], m_n=c['nome'], m_p=c.get('posto_graduacao',''): add_companion_plate(p_id, m_n, m_p)).props('unelevated color=cyan text-color=black dense bold').classes('text-[9px] shrink-0').tooltip('Aumentar quantidade de placas reservadas')

            render_list()
            with ui.row().classes('w-full justify-end q-mt-md'):
                ui.button('Fechar', on_click=diag.close).props('unelevated color=grey-8 dense')
        diag.open()

    def open_seniority_checklist_dialog(event, convidados):
        # 1. Carrega dados do layout
        layout = {}
        try:
            layout = json.loads(event['layout_json']) if event.get('layout_json') else {}
        except:
            pass
        layout_rows = layout.get('rows', 5)
        layout_cols = layout.get('cols', 8)

        # 2. Peso de antiguidade de posto/graduação
        def get_rank_weight(rank_str):
            if not rank_str:
                return 999
            rank = str(rank_str).upper().replace('.', '').replace(' ', '').strip()
            if rank in ('AE', 'ALMIRANTEDEESQUADRA'): return 1
            if rank in ('VA', 'VICEALMIRANTE'): return 2
            if rank in ('CA', 'CONTRAALMIRANTE'): return 3
            if rank in ('CMG', 'CAPITAODEMAREGUERRA'): return 4
            if rank in ('CF', 'CAPITAODEFRAGATA'): return 5
            if rank in ('CC', 'CAPITAODECORVETA'): return 6
            if rank in ('CT', 'CAPITAOTENENTE'): return 7
            if any(x in rank for x in ('1TEN', '1ºTEN', '1SOTEN', '1TENENTE', '1ºTENENTE')): return 8
            if any(x in rank for x in ('2TEN', '2ºTEN', '2SOTEN', '2TENENTE', '2ºTENENTE')): return 9
            if rank in ('GM', 'GUARDAMARINHA'): return 10
            if rank in ('SO', 'SUBOFICIAL', 'SUB-OFICIAL'): return 11
            if any(x in rank for x in ('1SG', '1ºSG', '1SGT', '1ºSGT', '1ºSARGENTO', '1SARGENTO')): return 12
            if any(x in rank for x in ('2SG', '2ºSG', '2SGT', '2ºSGT', '2ºSARGENTO', '2SARGENTO')): return 13
            if any(x in rank for x in ('3SG', '3ºSG', '3SGT', '3ºSGT', '3ºSARGENTO', '3SARGENTO', 'SG', 'SARGENTO')): return 14
            if rank in ('CB', 'CABO'): return 15
            if rank in ('SD', 'SOLDADO', 'MN', 'MARINHEIRO'): return 16
            return 90

        # Filtra e ordena convidados principais (exclui acompanhantes diretos na listagem principal)
        main_guests = [c for c in convidados if not c.get('convidado_principal_id')]
        sorted_by_seniority = sorted(
            main_guests,
            key=lambda x: (get_rank_weight(x.get('posto_graduacao')), x['nome'].strip().upper())
        )

        # Análise de quebra de precedência
        def get_seat_prestige(seat_id):
            if not seat_id:
                return 0
            try:
                r_lbl, c_str = seat_id.split('-')
                r_idx = 0
                for r in range(100):
                    if get_row_label(r) == r_lbl:
                        r_idx = r
                        break
                c_idx = int(c_str)
                row_score = (50 - r_idx) * 100
                col_score = 50 - abs(c_idx - (layout_cols + 1) / 2.0) * 10
                return row_score + col_score
            except:
                return 0

        # Atribui prestígio temporário
        for c in sorted_by_seniority:
            c['_prestige'] = get_seat_prestige(c.get('assento_id'))

        warnings = []
        for i in range(len(sorted_by_seniority)):
            c_senior = sorted_by_seniority[i]
            if get_rank_weight(c_senior.get('posto_graduacao')) >= 90:
                continue
            for j in range(i + 1, len(sorted_by_seniority)):
                c_junior = sorted_by_seniority[j]
                if get_rank_weight(c_junior.get('posto_graduacao')) >= 90:
                    continue
                if c_junior['_prestige'] > c_senior['_prestige']:
                    senior_seat = c_senior.get('assento_id') or 'Não Alocado'
                    junior_seat = c_junior.get('assento_id')
                    warnings.append({
                        'senior': f"{c_senior.get('posto_graduacao') or ''} {c_senior['nome']}".strip().upper(),
                        'senior_seat': senior_seat,
                        'junior': f"{c_junior.get('posto_graduacao') or ''} {c_junior['nome']}".strip().upper(),
                        'junior_seat': junior_seat
                    })

        with ui.dialog() as diag, ui.card().classes('q-pa-lg bg-slate-900 border border-cyan-500/40 rounded-xl').style('min-width: 650px; max-width: 95vw; max-height: 90vh; overflow-y: auto;'):
            with ui.row().classes('w-full justify-between items-center q-mb-md no-wrap'):
                with ui.row().classes('items-center gap-2'):
                    ui.icon('verified', color='cyan-5', size='md')
                    with ui.column().classes('gap-0'):
                        ui.label('CHECKLIST DE ANTIGUIDADE & PRECEDÊNCIA').classes('text-md font-bold text-white cyber-title')
                        ui.label('Validação automática de alocação por ordem de precedência militar').classes('text-[10px] text-grey-4')
                ui.button(icon='close', on_click=diag.close).props('flat round dense color=grey-4')

            # ── QUADRO DE ANÁLISE DE PRECEDÊNCIA ──
            with ui.card().classes('w-full q-pa-sm q-mb-md no-shadow border border-white/5').style('background: rgba(255,255,255,0.02);'):
                if warnings:
                    with ui.column().classes('w-full gap-1'):
                        ui.label('⚠️ ANOMALIAS DETECTADAS NA PRECEDÊNCIA:').classes('text-xs font-black text-amber-5 tracking-wider')
                        ui.label('Militar mais moderno alocado em cadeira de maior prestígio ou mais antigo sem assento:').classes('text-[10px] text-grey-4 q-mb-xs')
                        with ui.column().classes('w-full gap-1 max-h-36 overflow-y-auto'):
                            for w in warnings:
                                with ui.row().classes('w-full items-center justify-between no-wrap gap-2 text-[10px] py-1 border-b border-white/5'):
                                    ui.label(w['senior']).classes('text-white font-bold truncate col')
                                    ui.badge(w['senior_seat'], color='grey-7').classes('text-[8px] font-bold')
                                    ui.icon('arrow_forward', color='amber', size='xs')
                                    ui.label(w['junior']).classes('text-amber-3 truncate col')
                                    ui.badge(w['junior_seat'], color='cyan-9').classes('text-[8px] font-bold')
                else:
                    with ui.row().classes('w-full items-center gap-2 py-2 justify-center text-green-4'):
                        ui.icon('check_circle', size='sm')
                        ui.label('ORDEM DE PRECEDÊNCIA OK! Nenhum militar mais moderno está em assento superior.').classes('text-xs font-bold')

            # ── LISTA GERAL DE PRECEDÊNCIA ──
            ui.label('LISTA GERAL DE CONVIDADOS (ORDENADOS POR ANTIGUIDADE)').classes('text-[10px] font-bold text-cyan tracking-wider q-mb-xs')
            
            with ui.column().classes('w-full gap-1.5 scroll-container max-h-[380px] overflow-y-auto q-pr-xs'):
                for idx, c in enumerate(sorted_by_seniority, 1):
                    rank_w = get_rank_weight(c.get('posto_graduacao'))
                    is_mil = (rank_w < 90)
                    bg_color = 'rgba(0, 229, 255, 0.03)' if is_mil else 'rgba(255, 255, 255, 0.01)'
                    border_style = 'border-left: 3px solid #00e5ff;' if is_mil else 'border-left: 3px solid #6b7280;'
                    
                    with ui.row().classes('w-full items-center justify-between no-wrap q-pa-sm rounded-lg').style(
                        f'background: {bg_color}; {border_style}; border-top: 1px solid rgba(255,255,255,0.03);'
                    ):
                        with ui.row().classes('items-center gap-2 col-grow no-wrap'):
                            ui.label(f"{idx:02d}").classes('text-[10px] font-mono text-grey-5 shrink-0')
                            with ui.column().classes('gap-0 truncate'):
                                name_lbl = f"{c.get('posto_graduacao') or ''} {c['nome']}".strip().upper()
                                ui.label(name_lbl).classes('text-xs font-bold text-white truncate')
                                ui.label(c.get('cargo_funcao') or c.get('categoria') or 'Convidado').classes('text-[9px] text-grey-4 truncate')
                                
                        seat_val = c.get('assento_id')
                        if seat_val:
                            ui.badge(f"ASSENTO {seat_val}", color='green-9').classes('text-[9px] font-black tracking-wider')
                        else:
                            ui.badge("NÃO ALOCADO", color='amber-9').classes('text-[9px] font-black tracking-wider')
            
            with ui.row().classes('w-full justify-end q-mt-md'):
                ui.button('Fechar', on_click=diag.close).props('unelevated color=grey-8 dense')
        diag.open()

    # --- MODAIS E DIÁLOGOS DE GERENCIAMENTO DE SETORES E EVENTOS ---
    def open_create_event_dialog():
        with ui.dialog() as diag, ui.card().classes('q-pa-lg').style('min-width: 520px; max-width: 90vw;'):
            ui.label('📅 Novo Evento com Gestão de Setores').classes('text-md font-bold text-cyan cyber-title q-mb-xs')
            ui.label('Cadastre a solenidade e configure as dimensões e nomes dos setores').classes('text-xs text-grey-4 q-mb-md')
            
            nome = ui.input('Nome do Evento / Solenidade').props('dark outlined dense w-full')
            data = ui.input('Data do Evento').props('type=date dark outlined dense w-full')
            local = ui.input('Local (ex: Auditório Principal)').props('dark outlined dense w-full')
            
            with ui.row().classes('w-full gap-2'):
                ref_top = ui.input('Referência Superior', value='PALCO PRINCIPAL').props('dark outlined dense').classes('col')
                ref_bottom = ui.input('Referência Inferior', value='ENTRADA / FACHADA').props('dark outlined dense').classes('col')
            
            with ui.row().classes('w-full gap-2'):
                rows = ui.number('Linhas (Grid Total)', value=6, min=1, max=25, step=1).props('dark outlined dense').classes('col')
                cols = ui.number('Colunas (Grid Total)', value=12, min=1, max=30, step=1).props('dark outlined dense').classes('col')
                
            layout_tipo = ui.select(
                options={'auditorio': 'Auditório (Fileiras em Setores)', 'mesas': 'Mesas Redondas'}, 
                value='auditorio'
            ).props('dark outlined dense w-full')
            
            ui.separator().classes('q-my-sm')
            ui.label('⚙️ Nomes dos Setores Customizados (Opcional):').classes('text-xs text-amber font-bold')
            sec1 = ui.input('Setor 1 (Esquerda/Alpha)', value='SETOR ALPHA (ESQUERDA)').props('dark outlined dense w-full')
            sec2 = ui.input('Setor 2 (Centro/Nobre)', value='SETOR NOBRE (CENTRO)').props('dark outlined dense w-full')
            sec3 = ui.input('Setor 3 (Direita/Bravo)', value='SETOR BRAVO (DIREITA)').props('dark outlined dense w-full')

            with ui.row().classes('w-full justify-end q-mt-md gap-2'):
                ui.button('Cancelar', on_click=diag.close).props('unelevated color=grey-8 dense')
                
                def criar_evento_custom():
                    c_total = int(cols.value)
                    custom_sectors = [
                        {'name': sec1.value.strip().upper(), 'start_col': 1, 'end_col': max(1, c_total // 3)},
                        {'name': sec2.value.strip().upper(), 'start_col': max(1, c_total // 3) + 1, 'end_col': max(1, (c_total * 2) // 3)},
                        {'name': sec3.value.strip().upper(), 'start_col': max(1, (c_total * 2) // 3) + 1, 'end_col': c_total}
                    ]
                    create_event(nome.value, data.value, local.value, layout_tipo.value, rows.value, cols.value, ref_top.value, ref_bottom.value, custom_sectors)
                    diag.close()

                ui.button('Criar Evento', on_click=criar_evento_custom).props('unelevated color=primary text-color=black dense bold')
                
        diag.open()

    def open_edit_event_dialog(event, layout):
        sectors_list = list(layout.get('sectors', []))
        c_total = layout.get('cols', 12)

        if not sectors_list:
            sectors_list = [
                {'name': 'SETOR ALPHA (ESQUERDA)', 'start_col': 1, 'end_col': max(1, c_total // 3)},
                {'name': 'SETOR NOBRE (CENTRO)', 'start_col': max(1, c_total // 3) + 1, 'end_col': max(1, (c_total * 2) // 3)},
                {'name': 'SETOR BRAVO (DIREITA)', 'start_col': max(1, (c_total * 2) // 3) + 1, 'end_col': c_total}
            ]

        with ui.dialog() as diag, ui.card().classes('q-pa-lg').style('min-width: 650px; max-width: 95vw; max-height: 88vh; overflow-y: auto;'):
            ui.label('📝 Configuração do Layout e Gestão de Setores').classes('text-md font-bold text-cyan cyber-title q-mb-xs')
            ui.label('Adicione, exclua ou renomeie setores e configure os limites de colunas').classes('text-xs text-grey-4 q-mb-md')
            
            nome = ui.input('Nome do Evento / Solenidade', value=event['nome']).props('dark outlined dense w-full')
            data = ui.input('Data do Evento', value=event['data_evento']).props('type=date dark outlined dense w-full')
            local = ui.input('Local', value=event.get('local') or '').props('dark outlined dense w-full')
            
            with ui.row().classes('w-full gap-2'):
                ref_top = ui.input('Referência Superior', value=layout.get('ref_top', 'PALCO PRINCIPAL')).props('dark outlined dense').classes('col')
                ref_bottom = ui.input('Referência Inferior', value=layout.get('ref_bottom', 'ENTRADA / FACHADA')).props('dark outlined dense').classes('col')
            
            with ui.row().classes('w-full gap-2 q-mt-xs'):
                rows_input = ui.number('Fileiras Totais (Linhas)', value=layout.get('rows', 5), min=1, max=50, step=1).props('dark outlined dense').classes('col')
                cols_input = ui.number('Colunas Totais (Largura)', value=layout.get('cols', 12), min=1, max=100, step=1).props('dark outlined dense').classes('col')

            # Prepara setores para edição exibindo a Quantidade de Fileiras e Colunas individuais
            for s in sectors_list:
                width = (s.get('end_col', 1) - s.get('start_col', 1)) + 1
                s['cols_count'] = max(1, width)
                if 'rows_count' not in s:
                    s['rows_count'] = layout.get('rows', 5)

            with ui.row().classes('w-full justify-between items-center q-mb-xs'):
                ui.label('📍 SETORES DO AUDITÓRIO (Fileiras e Colunas Individuais)').classes('text-xs text-amber font-bold')
                
                def add_sector_item():
                    num_sectors = len(sectors_list)
                    default_r = int(rows_input.value or 5)
                    sectors_list.append({
                        'name': f"SETOR {num_sectors + 1}",
                        'rows_count': default_r,
                        'cols_count': 12
                    })
                    render_sectors_editor.refresh()

                ui.button('➕ Adicionar Setor', icon='add', on_click=add_sector_item).props('unelevated color=amber text-color=black dense').classes('text-xs')

            # Renderizador dinâmico de campos de setores com botão de exclusão e cálculo de quantitativos
            sector_inputs = []

            @ui.refreshable
            def render_sectors_editor():
                sector_inputs.clear()
                total_seats_calc = 0

                with ui.column().classes('w-full gap-2 q-my-xs'):
                    for idx, sec in enumerate(sectors_list):
                        r_val = int(sec.get('rows_count', int(rows_input.value or 5)))
                        c_val = int(sec.get('cols_count', 12))
                        sec_seats = r_val * c_val
                        total_seats_calc += sec_seats

                        with ui.card().classes('w-full q-pa-xs px-3 bg-black/40 border border-cyan-500/20 rounded-lg'):
                            with ui.row().classes('w-full items-center justify-between gap-2'):
                                s_name = ui.input(f'Nome do Setor {idx+1}', value=sec['name']).props('dark outlined dense').classes('col')
                                s_rows = ui.number('Fileiras', value=r_val, min=1, max=50).props('dark outlined dense').style('width: 90px;')
                                s_cols = ui.number('Colunas', value=c_val, min=1, max=100).props('dark outlined dense').style('width: 90px;')
                                
                                # Badge com o quantitativo de assentos do setor (Fileiras x Colunas)
                                with ui.column().classes('items-center gap-0').style('min-width: 100px;'):
                                    ui.label(f"🪑 {sec_seats} lugares").classes('text-xs font-bold text-cyan')
                                    ui.label(f"({r_val} fil × {c_val} col)").classes('text-[9px] text-grey-4 font-mono')

                                def remove_sector(i=idx):
                                    if len(sectors_list) > 1:
                                        sectors_list.pop(i)
                                        render_sectors_editor.refresh()
                                    else:
                                        ui.notify('O evento precisa de pelo menos 1 setor.', color='warning')

                                ui.button(icon='delete', on_click=remove_sector).props('unelevated color=danger dense flat round').classes('text-xs')
                                sector_inputs.append((s_name, s_rows, s_cols))

                    # Banner de Quantitativo Geral de Assentos do Evento
                    with ui.card().classes('w-full q-pa-sm bg-cyan-950/40 border border-cyan-500/40 rounded-lg q-mt-xs'):
                        with ui.row().classes('w-full justify-between items-center px-2'):
                            with ui.row().classes('items-center gap-2'):
                                ui.icon('event_seat', color='cyan', size='sm')
                                ui.label('QUANTITATIVO TOTAL DO AUDITÓRIO:').classes('text-xs font-bold text-white')
                            ui.badge(f"{total_seats_calc} ASSENTOS TOTAIS").props('color=cyan text-color=black bold').classes('text-xs q-px-sm')

            # Atualiza os cálculos ao alterar fileiras gerais
            rows_input.on('change', render_sectors_editor.refresh)
            render_sectors_editor()

            with ui.row().classes('w-full justify-end q-mt-md gap-2'):
                ui.button('Cancelar', on_click=diag.close).props('unelevated color=grey-8 dense')
                
                def salvar_alteracoes():
                    db = get_service_db_connection() or get_db_connection()
                    if db:
                        layout['ref_top'] = ref_top.value
                        layout['ref_bottom'] = ref_bottom.value
                        
                        # Converte a quantidade simples de fileiras e colunas de cada setor em posições sequenciais automáticas
                        new_sectors = []
                        current_col = 1
                        max_grid_rows = 1

                        for inp_name, inp_rows, inp_cols in sector_inputs:
                            if inp_name.value and inp_name.value.strip():
                                q_rows = max(1, int(inp_rows.value or 1))
                                q_cols = max(1, int(inp_cols.value or 1))
                                
                                if q_rows > max_grid_rows:
                                    max_grid_rows = q_rows

                                start_val = current_col
                                end_val = current_col + q_cols - 1
                                new_sectors.append({
                                    'name': inp_name.value.strip().upper(),
                                    'rows_count': q_rows,
                                    'start_col': start_val,
                                    'end_col': end_val
                                })
                                current_col = end_val + 1
                                
                        layout['sectors'] = new_sectors
                        layout['rows'] = max_grid_rows
                        layout['cols'] = current_col - 1

                        try:
                            updated_json = json.dumps(layout)
                            db.table('jade_eventos').update({
                                'nome': nome.value.upper(),
                                'data_evento': data.value,
                                'local': local.value or '',
                                'layout_json': updated_json
                            }).eq('id', event['id']).execute()
                            
                            # Atualiza atributos do objeto em memória
                            event['nome'] = nome.value.upper()
                            event['data_evento'] = data.value
                            event['local'] = local.value or ''
                            event['layout_json'] = updated_json
                            
                            # Reseta o seletor de setor e recarrega
                            state.selected_sector = 'Todos'
                            
                            ui.notify('Layout e Setores atualizados com sucesso!', color='success')
                            diag.close()
                            render_content.refresh()
                        except Exception as e:
                            ui.notify(f"Erro ao salvar: {e}", color='red')
                            
                ui.button('Salvar Alterações', on_click=salvar_alteracoes).props('unelevated color=primary text-color=black dense bold')
                
        diag.open()

    def confirm_delete_event(event):
        with ui.dialog() as diag, ui.card().classes('q-pa-md').style('min-width: 320px;'):
            ui.label('⚠️ Excluir Evento?').classes('text-md font-bold text-red q-mb-sm')
            ui.label(f'Tem certeza que deseja excluir o evento "{event["nome"]}"? Esta ação removerá definitivamente toda a lista de convidados e logs associados.').classes('text-xs text-grey-4 q-mb-md')
            
            with ui.row().classes('w-full justify-end gap-2'):
                ui.button('Cancelar', on_click=diag.close).props('unelevated color=grey-8 dense')
                ui.button(
                    'Confirmar Exclusão', 
                    on_click=lambda: [delete_event(event), diag.close()]
                ).props('unelevated color=danger dense')
        diag.open()

    def open_edit_guest_dialog(guest=None, event_id=None):
        title_text = '➕ Adicionar Convidado' if not guest else '📝 Editar Convidado'
        
        with ui.dialog() as diag, ui.card().classes('q-pa-md').style('min-width: 360px;'):
            ui.label(title_text).classes('text-md font-bold text-cyan q-mb-md')
            
            nome = ui.input('Nome do Convidado', value=guest['nome'] if guest else '').props('dark outlined dense w-full')
            posto = ui.input('Posto / Graduação (ex: CT, CF, Ten)', value=guest.get('posto_graduacao') or '' if guest else '').props('dark outlined dense w-full')
            cargo = ui.input('Cargo / Função (ex: Secretário)', value=guest.get('cargo_funcao') or '' if guest else '').props('dark outlined dense w-full')
            
            cat_list = ['VIP', 'Autoridade Civil', 'Autoridade Militar', 'Imprensa', 'Apoio', 'Geral']
            categoria = ui.select(
                options=cat_list, 
                value=guest.get('categoria', 'Geral') if guest else 'Geral'
            ).props('dark outlined dense w-full')
            
            acomps = ui.number('Acompanhantes', value=guest.get('max_acompanhantes', 0) if guest else 0, min=0, step=1).props('dark outlined dense w-full')
            
            with ui.row().classes('w-full justify-end q-mt-md gap-2'):
                ui.button('Cancelar', on_click=diag.close).props('unelevated color=grey-8 dense')
                
                def salvar():
                    reg = {
                        'nome': nome.value.upper(),
                        'posto_graduacao': posto.value or None,
                        'cargo_funcao': cargo.value or None,
                        'categoria': categoria.value,
                        'max_acompanhantes': int(acomps.value)
                    }
                    save_guest(guest['id'] if guest else None, reg, event_id or state.selected_event_id)
                    diag.close()
                    
                ui.button('Salvar', on_click=salvar).props('unelevated color=primary text-color=black dense')
        diag.open()

    def confirm_delete_guest(guest):
        with ui.dialog() as diag, ui.card().classes('q-pa-md').style('min-width: 320px;'):
            ui.label('⚠️ Remover Convidado?').classes('text-md font-bold text-red q-mb-sm')
            ui.label(f'Tem certeza que deseja remover "{guest["nome"]}" da lista?').classes('text-xs text-grey-4 q-mb-md')
            
            with ui.row().classes('w-full justify-end gap-2'):
                ui.button('Cancelar', on_click=diag.close).props('unelevated color=grey-8 dense')
                ui.button(
                    'Remover', 
                    on_click=lambda: [delete_guest(guest), diag.close()]
                ).props('unelevated color=danger dense')
        diag.open()

    def open_allocate_seat_dialog(seat_id, convidados, event_id):
        def get_category_priority(cat):
            cat_upper = str(cat or '').upper()
            if 'VIP' in cat_upper: return 0
            if 'MILITAR' in cat_upper: return 1
            if 'CIVIL' in cat_upper: return 2
            if 'IMPRENSA' in cat_upper: return 3
            if 'APOIO' in cat_upper: return 4
            return 5

        sorted_convidados = sorted(
            convidados,
            key=lambda c: (get_category_priority(c.get('categoria', 'Geral')), c.get('posto_graduacao') or '', c['nome'])
        )
        
        with ui.dialog() as diag, ui.card().classes('q-pa-md').style('min-width: 420px; max-height: 550px;'):
            ui.label(f'Alocar Assento {seat_id}').classes('text-md font-bold text-cyan q-mb-xs')
            ui.label('Selecione um convidado na lista para alocar imediatamente:').classes('text-xs text-grey-4 q-mb-md')
            
            search_input = ui.input(placeholder='Filtrar por nome ou cargo...').props('dark outlined dense clearable w-full q-mb-md')
            
            @ui.refreshable
            def render_dialog_guests():
                query = search_input.value.lower() if search_input.value else ""
                filtered = sorted_convidados
                if query:
                    filtered = [
                        c for c in sorted_convidados
                        if query in c['nome'].lower() or
                        (c.get('cargo_funcao') and query in c['cargo_funcao'].lower()) or
                        (c.get('posto_graduacao') and query in c['posto_graduacao'].lower())
                    ]
                
                with ui.column().classes('w-full gap-2 scroll-container q-py-xs').style('overflow-y: auto; max-height: 320px;'):
                    if filtered:
                        for c in filtered:
                            is_seated = bool(c.get('assento_id'))
                            card_bg = 'rgba(0, 229, 255, 0.08)' if is_seated else 'rgba(255, 255, 255, 0.02)'
                            card_border = 'rgba(0, 229, 255, 0.25)' if is_seated else 'rgba(255, 255, 255, 0.06)'
                            text_style = 'opacity-70' if is_seated else ''
                            
                            with ui.card().classes('w-full q-pa-xs px-2 no-shadow rounded-lg cursor-pointer transition-all hover:bg-cyan-500/20').style(
                                f'background: {card_bg}; border: 1px solid {card_border}; gap: 0;'
                            ).on('click', lambda c_id=c['id']: [allocate_guest(c_id, seat_id, event_id), diag.close()]):
                                with ui.row().classes('w-full justify-between items-center no-wrap'):
                                    with ui.column().classes('gap-0'):
                                        nome_exibicao = f"{c.get('posto_graduacao') or ''} {c['nome']}".strip()
                                        ui.label(nome_exibicao).classes(f'text-xs font-bold text-white {text_style}')
                                        
                                        sub = c.get('cargo_funcao') or 'Sem cargo/função'
                                        ui.label(f"[{c.get('categoria', 'Geral').upper()}] {sub}").classes('text-[9px] text-grey-4')
                                    
                                    if is_seated:
                                        ui.badge(f"Assento {c['assento_id']}").props('color=cyan text-color=black').classes('text-[8px]')
                    else:
                        with ui.column().classes('w-full items-center justify-center q-py-md text-grey'):
                            ui.label('Nenhum convidado disponível').classes('text-xs')
                            
            search_input.on('value-change', render_dialog_guests.refresh)
            render_dialog_guests()
            
            with ui.row().classes('w-full justify-end q-mt-md'):
                ui.button('Cancelar', on_click=diag.close).props('unelevated color=grey-8 dense')
                
        diag.open()

    def open_seat_actions_dialog(seat_id, guest, convidados, event_id):
        with ui.dialog() as diag, ui.card().classes('q-pa-md').style('min-width: 400px; max-height: 600px;'):
            ui.label(f'Ações do Assento {seat_id}').classes('text-md font-bold text-cyan q-mb-xs')
            
            nome_completo = f"{guest.get('posto_graduacao') or ''} {guest['nome']}".strip()
            ui.label(nome_completo).classes('text-sm font-bold text-white')
            ui.label(guest.get('cargo_funcao') or guest.get('categoria') or 'Convidado').classes('text-xs text-grey-4 q-mb-md')
            
            ui.button(
                'Liberar / Desalocar Assento', 
                icon='block', 
                on_click=lambda: [remove_guest_allocation(guest), diag.close()]
            ).props('unelevated color=danger w-full q-mb-sm dense')
            
            other_allocated = [c for c in convidados if c.get('assento_id') and c['id'] != guest['id']]
            if other_allocated:
                ui.separator().classes('q-my-sm')
                ui.label('Permutar (Swap) com outro assento ocupado:').classes('text-[11px] text-grey q-mb-xs')
                
                swap_options = {
                    c['id']: f"{c['assento_id']} - {c.get('posto_graduacao') or ''} {c['nome']}".strip()
                    for c in other_allocated
                }
                swap_target = ui.select(swap_options).props('dark outlined dense w-full')
                
                ui.button(
                    'Confirmar Troca', 
                    icon='swap_horiz', 
                    on_click=lambda: [swap_guests(guest, swap_target.value, event_id), diag.close()]
                ).props('unelevated color=primary text-color=black w-full q-mt-xs dense')
                
            with ui.row().classes('w-full justify-end q-mt-md'):
                ui.button('Fechar', on_click=diag.close).props('unelevated color=grey-8 dense')
                
        diag.open()

    # ═══════════════════════════════════════════════════════════════
    # FASE 2: DIALOG DE CONFIRMAÇÃO DE PRESENÇAS EM MASSA
    # ═══════════════════════════════════════════════════════════════
    def open_mass_confirmation_dialog(event, convidados):
        """Dialog para confirmar/recusar presenças em massa, com checkboxes por categoria."""
        
        # Estado local do dialog
        selected_ids = set()
        filter_status = {'value': 'todos'}
        
        with ui.dialog().classes('q-dialog--maximized') as diag, ui.card().classes('w-full').style(
            f'background: {THEME["bg_panel"]}; max-width: 1200px; max-height: 95vh;'
        ):
            # Cabeçalho
            with ui.row().classes('w-full items-center justify-between q-mb-sm'):
                with ui.column().classes('gap-0'):
                    ui.label('✅ CONFIRMAÇÃO DE PRESENÇAS EM MASSA').classes('text-lg font-bold text-cyan')
                    ui.label(f'Evento: {event.get("nome", "N/I")} — {event.get("data_evento", "")}').classes('text-xs text-grey-4')
                ui.button(icon='close', on_click=diag.close).props('flat round dense text-color=grey')

            # Agrupar convidados por categoria (excluir acompanhantes — eles seguem o principal)
            main_guests = [c for c in convidados if not c.get('convidado_principal_id')]
            categories = {}
            for c in main_guests:
                cat = c.get('categoria', 'Sem Categoria') or 'Sem Categoria'
                if cat not in categories:
                    categories[cat] = []
                categories[cat].append(c)

            # Contadores resumo
            total = len(main_guests)
            count_conf = sum(1 for c in main_guests if c.get('status_confirmacao') == 'confirmado')
            count_ref = sum(1 for c in main_guests if c.get('status_confirmacao') == 'recusado')
            count_prov = sum(1 for c in main_guests if c.get('status_confirmacao') == 'provavel')
            count_pend = total - count_conf - count_ref - count_prov

            # Barra de resumo
            with ui.card().classes('w-full q-pa-xs no-shadow rounded-lg q-mb-sm').style('background: rgba(0,0,0,0.3); border: 1px solid rgba(255,255,255,0.1);'):
                with ui.row().classes('w-full items-center justify-between wrap gap-2'):
                    with ui.row().classes('gap-2 wrap'):
                        ui.badge(f'📋 Total: {total}').props('color=grey-7 text-color=white').classes('text-xs')
                        ui.badge(f'✅ Confirmados: {count_conf}').props('color=green text-color=white').classes('text-xs')
                        ui.badge(f'🟡 Prováveis: {count_prov}').props('color=amber text-color=black').classes('text-xs')
                        ui.badge(f'❓ Pendentes: {count_pend}').props('color=blue-grey text-color=white').classes('text-xs')
                        ui.badge(f'❌ Recusados: {count_ref}').props('color=red text-color=white').classes('text-xs')

            # Filtro de status
            @ui.refreshable
            def render_filter_bar():
                with ui.row().classes('w-full items-center gap-2 q-mb-sm wrap'):
                    ui.label('Filtrar:').classes('text-xs text-grey-4 font-bold')
                    filter_options = [
                        ('todos', '🌐 Todos', 'grey-7'),
                        ('confirmado', '✅ Confirmados', 'green'),
                        ('pendente', '❓ Pendentes', 'blue-grey'),
                        ('provavel', '🟡 Prováveis', 'amber'),
                        ('recusado', '❌ Recusados', 'red'),
                    ]
                    for fval, flabel, fcol in filter_options:
                        is_active = filter_status['value'] == fval
                        def make_filter_click(v=fval):
                            def _click():
                                filter_status['value'] = v
                                render_filter_bar.refresh()
                                render_guest_list.refresh()
                            return _click
                        ui.button(flabel, on_click=make_filter_click()).props(
                            f'dense {"unelevated" if is_active else "outline"} color={fcol} {"text-color=white" if is_active else ""}'
                        ).classes('text-[10px] q-px-xs')
                    
                    ui.separator().props('vertical').classes('q-mx-xs')
                    ui.label(f'🔲 {len(selected_ids)} selecionados').classes('text-[10px] text-grey-5 font-bold')

            render_filter_bar()

            # Lista de convidados agrupados por categoria
            @ui.refreshable
            def render_guest_list():
                with ui.scroll_area().classes('w-full').style('max-height: 55vh;'):
                    for cat_name, cat_guests in sorted(categories.items()):
                        # Filtrar conforme o filtro ativo
                        fval = filter_status['value']
                        if fval == 'todos':
                            filtered = cat_guests
                        elif fval == 'pendente':
                            filtered = [c for c in cat_guests if c.get('status_confirmacao', 'pendente') not in ('confirmado', 'recusado', 'provavel')]
                        else:
                            filtered = [c for c in cat_guests if c.get('status_confirmacao') == fval]
                        
                        if not filtered:
                            continue

                        with ui.expansion(
                            f'📁 {cat_name} ({len(filtered)} convidados)',
                            value=True
                        ).classes('w-full q-mb-xs text-white font-bold').style(
                            'background: rgba(0,60,80,0.3); border-radius: 8px;'
                        ):
                            # Botão selecionar todo o bloco
                            with ui.row().classes('w-full items-center justify-between q-mb-xs q-px-sm'):
                                def make_select_all(guests=filtered):
                                    def _click():
                                        all_ids = {str(g['id']) for g in guests}
                                        if all_ids.issubset(selected_ids):
                                            selected_ids.difference_update(all_ids)
                                        else:
                                            selected_ids.update(all_ids)
                                        render_guest_list.refresh()
                                        render_filter_bar.refresh()
                                    return _click
                                
                                all_selected = all(str(g['id']) in selected_ids for g in filtered)
                                ui.button(
                                    f'{"☑ Desmarcar" if all_selected else "☐ Selecionar"} Todo o Bloco',
                                    on_click=make_select_all()
                                ).props(f'dense flat text-color={"cyan" if not all_selected else "amber"}').classes('text-[10px]')

                            for g in filtered:
                                g_id = str(g['id'])
                                nome = g.get('nome', 'N/I')
                                posto = g.get('posto_graduacao', '') or ''
                                cargo = g.get('cargo_funcao', '') or ''
                                st = g.get('status_confirmacao', 'pendente') or 'pendente'
                                max_ac = g.get('max_acompanhantes', 0) or 0
                                
                                st_emoji = {'confirmado': '✅', 'recusado': '❌', 'provavel': '🟡', 'pendente': '❓'}.get(st, '❓')
                                st_color = {'confirmado': 'rgba(0,200,80,0.15)', 'recusado': 'rgba(200,0,0,0.15)', 'provavel': 'rgba(255,200,0,0.15)'}.get(st, 'rgba(255,255,255,0.03)')
                                
                                is_checked = g_id in selected_ids

                                with ui.row().classes('w-full items-center q-py-xs q-px-sm rounded-lg gap-2').style(
                                    f'background: {st_color}; border-bottom: 1px solid rgba(255,255,255,0.05);'
                                ):
                                    def make_toggle(gid=g_id):
                                        def _toggle(e):
                                            if e.value:
                                                selected_ids.add(gid)
                                            else:
                                                selected_ids.discard(gid)
                                            render_filter_bar.refresh()
                                        return _toggle

                                    ui.checkbox('', value=is_checked, on_change=make_toggle()).props('dense dark')
                                    ui.label(f'{st_emoji}').classes('text-sm')
                                    ui.label(f'{posto}').classes('text-[10px] text-amber font-bold').style('min-width: 60px;')
                                    ui.label(f'{nome}').classes('text-xs text-white font-bold flex-grow')
                                    if cargo:
                                        ui.label(f'{cargo[:40]}').classes('text-[9px] text-grey-5 truncate').style('max-width: 200px;')
                                    if max_ac > 0:
                                        ui.badge(f'+{max_ac} acomp.').props('color=blue-grey text-color=white').classes('text-[9px]')

            render_guest_list()

            # Barra de ações
            ui.separator().classes('q-my-sm')
            with ui.row().classes('w-full items-center justify-between wrap gap-2'):
                with ui.row().classes('gap-2 wrap'):
                    async def batch_update_status(new_status):
                        if not selected_ids:
                            ui.notify('⚠️ Nenhum convidado selecionado.', color='warning')
                            return
                        _db = get_service_db_connection() or get_db_connection()
                        if not _db:
                            ui.notify('❌ Banco indisponível.', color='negative')
                            return
                        
                        # Determinar status_placa com base no status de confirmação
                        if new_status == 'confirmado':
                            new_placa = 'pendente'  # Placa precisa ser impressa
                        elif new_status == 'provavel':
                            new_placa = 'pendente'  # Tier 2 — preventivo
                        elif new_status == 'recusado':
                            new_placa = 'nao_necessaria'
                        else:
                            new_placa = 'nao_necessaria'
                        
                        count_updated = 0
                        for gid in list(selected_ids):
                            try:
                                _db.table('jade_convidados').update({
                                    'status_confirmacao': new_status,
                                    'status_placa': new_placa
                                }).eq('id', int(gid)).execute()
                                count_updated += 1
                            except Exception:
                                # Fallback: coluna status_placa pode não existir ainda
                                try:
                                    _db.table('jade_convidados').update({
                                        'status_confirmacao': new_status
                                    }).eq('id', int(gid)).execute()
                                    count_updated += 1
                                except Exception as e2:
                                    print(f"[BATCH UPDATE ERR] {e2}")
                        
                        status_labels = {'confirmado': 'CONFIRMADOS', 'recusado': 'RECUSADOS', 'provavel': 'PROVÁVEIS', 'pendente': 'PENDENTES'}
                        ui.notify(f'✅ {count_updated} convidados marcados como {status_labels.get(new_status, new_status)}!', color='positive')
                        
                        # Notificar via Telegram se gerou novas placas pendentes
                        if new_status in ('confirmado', 'provavel') and count_updated > 0:
                            try:
                                from notifications_manager import notify_jade_production
                                notify_jade_production(event.get('nome', 'Solenidade'), count_updated, count_updated)
                            except Exception as n_err:
                                print(f"[JADE NOTIFY ERR] {n_err}")
                                
                        selected_ids.clear()
                        diag.close()
                        render_content.refresh()

                    ui.button('✅ Confirmar Selecionados', icon='check_circle', on_click=lambda: batch_update_status('confirmado')).props('unelevated color=green text-color=white dense').classes('text-xs q-px-sm')
                    ui.button('🟡 Marcar Prováveis (Tier 2)', icon='trending_up', on_click=lambda: batch_update_status('provavel')).props('unelevated color=amber text-color=black dense').classes('text-xs q-px-sm')
                    ui.button('❌ Recusar Selecionados', icon='cancel', on_click=lambda: batch_update_status('recusado')).props('unelevated color=red text-color=white dense').classes('text-xs q-px-sm')
                    ui.button('❓ Voltar a Pendente', icon='pending', on_click=lambda: batch_update_status('pendente')).props('outline color=grey dense').classes('text-xs q-px-sm')

                ui.button('Fechar', icon='close', on_click=diag.close).props('unelevated color=grey-8 dense').classes('text-xs')

        diag.open()

    # ═══════════════════════════════════════════════════════════════
    # FASE 3: IMPORTADOR INTELIGENTE DE EXCEL COM LEITURA DE CORES
    # ═══════════════════════════════════════════════════════════════
    def open_smart_excel_import_dialog(event):
        """Dialog com upload de arquivo .xlsx para importação inteligente de convidados com leitura de cores."""
        with ui.dialog() as diag, ui.card().classes('w-full').style(
            f'background: {THEME["bg_panel"]}; max-width: 650px;'
        ):
            ui.label('📥 IMPORTADOR INTELIGENTE JADE (EXCEL COM CORES)').classes('text-md font-bold text-cyan')
            ui.label(f'Solenidade Destino: {event.get("nome", "N/I")}').classes('text-xs text-grey-4 q-mb-sm')
            
            ui.markdown(
                "**Como funciona a importação inteligente:**\n"
                "• **Suporta o modelo oficial:** Planilhas `.xlsx` com blocos (ex: `LISTA DE CONVITE.xlsx`).\n"
                "• **Leitura de Cores:** Células **Verdes** na coluna *Confirmado* viram **CONFIRMADO** (placa na fila de produção). Células **Vermelhas** viram **RECUSADO**.\n"
                "• **Anti-duplicação:** Se a autoridade já existir na solenidade (mesmo Nome + Posto), os dados e acompanhantes são **atualizados sem duplicar**.\n"
                "• **Categorias:** Os blocos (Almirantado, CMGs FN, Civis, etc.) são identificados automaticamente."
            ).classes('text-xs text-grey-3 bg-black/30 q-pa-sm rounded-lg border border-white/10')
            
            log_container = ui.column().classes('w-full q-my-sm')

            async def handle_upload(e):
                import inspect
                file_obj = getattr(e, 'file', None)
                if not file_obj and hasattr(e, 'files') and e.files:
                    file_obj = e.files[0]
                
                if not file_obj:
                    ui.notify('❌ Nenhum arquivo de planilha detectado.', color='negative')
                    return

                file_bytes = file_obj.read() if hasattr(file_obj, 'read') else getattr(file_obj, 'content', None)
                if inspect.isawaitable(file_bytes):
                    file_bytes = await file_bytes
                elif hasattr(file_bytes, 'read'):
                    file_bytes = file_bytes.read()

                if not file_bytes:
                    ui.notify('❌ Arquivo de planilha vazio.', color='negative')
                    return

                file_name = getattr(file_obj, 'name', 'planilha.xlsx').lower()

                try:
                    import io
                    db = get_service_db_connection() or get_db_connection()
                    if not db:
                        ui.notify('❌ Banco de dados indisponível.', color='negative')
                        return
                    
                    event_id = event['id']
                    res_exist = db.table('jade_convidados').select('*').eq('evento_id', event_id).execute()
                    existing_list = res_exist.data if res_exist.data else []
                    
                    existing_map = {}
                    for item in existing_list:
                        if not item.get('convidado_principal_id'):
                            key = f"{(item.get('nome') or '').strip().upper()}|{(item.get('posto_graduacao') or '').strip().upper()}"
                            existing_map[key] = item

                    count_inserted = 0
                    count_updated = 0
                    count_conf = 0
                    count_rec = 0
                    count_pend = 0

                    wb = None
                    if file_name.endswith('.xlsx') or file_name.endswith('.xlsm'):
                        try:
                            import openpyxl
                            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
                        except Exception:
                            wb = None

                    if wb:
                        sheet = wb.active
                        current_category = "Geral"
                        for r in range(1, sheet.max_row + 1):
                            row_cells = [sheet.cell(row=r, column=c) for c in range(1, sheet.max_column + 1)]
                            row_vals = [c.value for c in row_cells]

                            if not any(v is not None for v in row_vals):
                                continue

                            c0 = str(row_vals[0]).strip() if len(row_vals) > 0 and row_vals[0] is not None else ""
                            c1 = str(row_vals[1]).strip() if len(row_vals) > 1 and row_vals[1] is not None else ""
                            c2 = str(row_vals[2]).strip() if len(row_vals) > 2 and row_vals[2] is not None else ""

                            if c0.upper() in ('QTD', 'QUANTIDADE') or c1.upper() in ('POSTO', 'GRAU'):
                                continue

                            if c0 and not c0.isdigit() and c0.upper() not in ('QTD', 'QUANTIDADE'):
                                current_category = c0
                                continue

                            nome = c2
                            if not nome or nome.upper() in ('NOME', 'CONVITE'):
                                continue

                            posto = c1
                            cargo = str(row_vals[3]).strip() if len(row_vals) > 3 and row_vals[3] is not None else ""
                            conf_val = row_vals[5] if len(row_vals) > 5 else None
                            conjuge_val = row_vals[6] if len(row_vals) > 6 else None
                            ant_val = row_vals[9] if len(row_vals) > 9 else None
                            try:
                                numero_antiguidade = int(float(str(ant_val))) if ant_val not in (None, '', 'None') else None
                            except (ValueError, TypeError):
                                numero_antiguidade = None

                            conf_fill = ""
                            if len(row_cells) > 5 and row_cells[5].fill and row_cells[5].fill.start_color:
                                conf_fill = str(row_cells[5].fill.start_color.rgb or "")

                            status_conf = "pendente"
                            status_placa = "nao_necessaria"

                            if any(g in conf_fill for g in ('00B050', '92D050', '00FF00')):
                                status_conf = "confirmado"
                                status_placa = "pendente"
                                count_conf += 1
                            elif any(rc in conf_fill for rc in ('FF0000', 'FA1717', 'C00000', 'FF5555')):
                                status_conf = "recusado"
                                status_placa = "nao_necessaria"
                                count_rec += 1
                            else:
                                if conf_val in (1, 2, '1', '2', 'SIM', 'Sim', 'sim', 'CONFIRMADO'):
                                    status_conf = "confirmado"
                                    status_placa = "pendente"
                                    count_conf += 1
                                else:
                                    count_pend += 1

                            max_acomp = 0
                            if isinstance(conf_val, int) and conf_val > 1:
                                max_acomp = conf_val - 1
                            elif str(conf_val).strip().isdigit() and int(str(conf_val).strip()) > 1:
                                max_acomp = int(str(conf_val).strip()) - 1
                            elif conjuge_val in ('E', '1', 1) or (isinstance(conjuge_val, str) and len(conjuge_val) > 2):
                                max_acomp = 1

                            key = f"{nome.strip().upper()}|{posto.strip().upper()}"
                            
                            conv_data = {
                                'evento_id': event_id,
                                'nome': nome,
                                'posto_graduacao': posto,
                                'cargo_funcao': cargo,
                                'categoria': current_category,
                                'status_confirmacao': status_conf,
                                'max_acompanhantes': max_acomp,
                            }
                            if numero_antiguidade is not None:
                                conv_data['numero_antiguidade'] = numero_antiguidade
                            
                            conv_data_with_placa = dict(conv_data)
                            conv_data_with_placa['status_placa'] = status_placa

                            if key in existing_map:
                                existing_id = existing_map[key]['id']
                                try:
                                    db.table('jade_convidados').update(conv_data_with_placa).eq('id', existing_id).execute()
                                except Exception:
                                    safe = {k: v for k, v in conv_data.items() if k != 'numero_antiguidade'}
                                    db.table('jade_convidados').update(safe).eq('id', existing_id).execute()
                                
                                sync_companions(existing_id, nome, max_acomp, event_id, current_category)
                                count_updated += 1
                            else:
                                res_ins = None
                                try:
                                    res_ins = db.table('jade_convidados').insert(conv_data_with_placa).execute()
                                except Exception:
                                    safe = {k: v for k, v in conv_data.items() if k != 'numero_antiguidade'}
                                    safe['status_placa'] = status_placa
                                    res_ins = db.table('jade_convidados').insert(safe).execute()
                                
                                if res_ins and res_ins.data:
                                    new_id = res_ins.data[0]['id']
                                    sync_companions(new_id, nome, max_acomp, event_id, current_category)
                                count_inserted += 1

                    else:
                        # Leitor Multi-Formato Universal via Pandas (.csv, .tsv, .txt, .ods, .xls)
                        bio_p = io.BytesIO(file_bytes)
                        df_p = None
                        if file_name.endswith('.csv') or file_name.endswith('.tsv') or file_name.endswith('.txt'):
                            for sep in [';', ',', '\t', '|']:
                                try:
                                    bio_p.seek(0)
                                    df_test = pd.read_csv(bio_p, sep=sep, encoding='utf-8-sig', dtype=str)
                                    if len(df_test.columns) > 1:
                                        df_p = df_test
                                        break
                                except Exception:
                                    pass
                            if df_p is None:
                                bio_p.seek(0)
                                df_p = pd.read_csv(bio_p, encoding='utf-8-sig', dtype=str)
                        elif file_name.endswith('.xls'):
                            try:
                                df_p = pd.read_excel(bio_p, engine='xlrd', dtype=str)
                            except Exception:
                                df_p = pd.read_excel(bio_p, dtype=str)
                        elif file_name.endswith('.ods'):
                            try:
                                df_p = pd.read_excel(bio_p, engine='odf', dtype=str)
                            except Exception:
                                df_p = pd.read_excel(bio_p, dtype=str)
                        else:
                            df_p = pd.read_excel(bio_p, dtype=str)

                        if df_p is not None:
                            # Mapeia colunas variantes automaticamente
                            col_map = {}
                            for col in df_p.columns:
                                c_clean = str(col).strip().upper()
                                if c_clean in ('NOME', 'NOME COMPLETO', 'AUTORIDADE', 'NOME DA AUTORIDADE', 'NOME DO CONVIDADO', 'CONVIDADO'):
                                    col_map[col] = 'Nome'
                                elif c_clean in ('POSTO', 'POSTO/GRADUAÇÃO', 'POSTO/GRADUACAO', 'POSTO / GRADUAÇÃO', 'GRADUAÇÃO', 'GRADUACAO', 'POSTO_GRADUACAO'):
                                    col_map[col] = 'Posto/Graduacao'
                                elif c_clean in ('CARGO', 'CARGO/FUNÇÃO', 'CARGO/FUNCAO', 'FUNÇÃO', 'FUNCAO', 'TÍTULO', 'TITULO', 'CARGO_FUNCAO'):
                                    col_map[col] = 'Cargo/Função'
                                elif c_clean in ('CATEGORIA', 'SETOR', 'BLOCO', 'GRUPO'):
                                    col_map[col] = 'Categoria'
                                elif c_clean in ('MAX ACOMPANHANTES', 'ACOMPANHANTES', 'ACOMP', 'N_ACOMPANHANTES', 'QTD ACOMPANHANTES', 'QTD_ACOMP'):
                                    col_map[col] = 'Max Acompanhantes'
                                elif c_clean in ('ANTIGUIDADE', 'Nº ANTIGUIDADE', 'NUMERO_ANTIGUIDADE'):
                                    col_map[col] = 'Antiguidade'
                            
                            df_p = df_p.rename(columns=col_map)
                            
                            if 'Nome' not in df_p.columns and len(df_p.columns) > 0:
                                df_p = df_p.rename(columns={df_p.columns[0]: 'Nome'})

                            for _, row in df_p.iterrows():
                                nome_p = str(row.get('Nome', '')).strip().upper()
                                if not nome_p or nome_p == 'NAN':
                                    continue
                                posto_p = str(row.get('Posto/Graduacao', '')).strip() if pd.notna(row.get('Posto/Graduacao')) and str(row.get('Posto/Graduacao')) != 'nan' else ''
                                cargo_p = str(row.get('Cargo/Função', '')).strip() if pd.notna(row.get('Cargo/Função')) and str(row.get('Cargo/Função')) != 'nan' else ''
                                cat_p = str(row.get('Categoria', 'Geral')).strip() if pd.notna(row.get('Categoria')) and str(row.get('Categoria')) != 'nan' else 'Geral'
                                try:
                                    acomp_p = int(row.get('Max Acompanhantes', 0)) if pd.notna(row.get('Max Acompanhantes')) else 0
                                except Exception:
                                    acomp_p = 0

                                ant_p = None
                                try:
                                    if pd.notna(row.get('Antiguidade')):
                                        ant_p = int(float(str(row.get('Antiguidade'))))
                                except Exception:
                                    ant_p = None

                                key_p = f"{nome_p}|{posto_p}"
                                guest_data = {
                                    'evento_id': event_id,
                                    'nome': nome_p,
                                    'posto_graduacao': posto_p,
                                    'cargo_funcao': cargo_p,
                                    'categoria': cat_p,
                                    'status_confirmacao': 'confirmado',
                                    'status_placa': 'pendente',
                                    'max_acompanhantes': acomp_p
                                }
                                if ant_p is not None:
                                    guest_data['numero_antiguidade'] = ant_p

                                if key_p in existing_map:
                                    e_id = existing_map[key_p]['id']
                                    try:
                                        db.table('jade_convidados').update(guest_data).eq('id', e_id).execute()
                                    except Exception:
                                        safe = {k: v for k, v in guest_data.items() if k != 'numero_antiguidade'}
                                        db.table('jade_convidados').update(safe).eq('id', e_id).execute()
                                    sync_companions(e_id, nome_p, acomp_p, event_id, cat_p)
                                    count_updated += 1
                                else:
                                    res_p = None
                                    try:
                                        res_p = db.table('jade_convidados').insert(guest_data).execute()
                                    except Exception:
                                        safe = {k: v for k, v in guest_data.items() if k != 'numero_antiguidade'}
                                        res_p = db.table('jade_convidados').insert(safe).execute()
                                    if res_p and res_p.data:
                                        new_id = res_p.data[0]['id']
                                        sync_companions(new_id, nome_p, acomp_p, event_id, cat_p)
                                    count_inserted += 1

                    ui.notify(f'✅ Importação concluída! {count_inserted} novos, {count_updated} atualizados.', color='positive')
                    
                    log_container.clear()
                    with log_container:
                        with ui.card().classes('w-full q-pa-sm no-shadow rounded-lg').style('background: rgba(0,255,150,0.1); border: 1px solid rgba(0,255,150,0.3);'):
                            ui.label('📊 RELATÓRIO DE IMPORTAÇÃO CONCLUÍDA').classes('text-xs font-bold text-green')
                            ui.label(f'• Registros Processados: {count_inserted + count_updated}').classes('text-xs text-white')
                            ui.label(f'• 🆕 Novos Convidados Inseridos: {count_inserted}').classes('text-xs text-green font-bold')
                            ui.label(f'• 🔄 Atualizados (Merge Anti-Duplicação): {count_updated}').classes('text-xs text-cyan font-bold')
                    
                    render_content.refresh()

                except Exception as ex:
                    print(f"[SMART IMPORT ERR] {ex}")
                    ui.notify(f'❌ Erro ao processar planilha: {ex}', color='negative')

            ui.upload(on_upload=handle_upload, auto_upload=True).props('accept=.xlsx,.xls,.xlsm,.csv,.tsv,.txt,.ods dark').classes('w-full q-my-sm')
            
            with ui.row().classes('w-full justify-end q-mt-sm'):
                ui.button('Fechar', icon='close', on_click=diag.close).props('unelevated color=grey-8 dense').classes('text-xs')

        diag.open()

    # ═══════════════════════════════════════════════════════════════
    # FASE 5: CADASTRO MESTRE DE AUTORIDADES (REUTILIZÁVEL)
    # ═══════════════════════════════════════════════════════════════
    def open_master_authorities_dialog(event):
        """Dialog para gerenciar e importar autoridades da base mestre permanente."""
        with ui.dialog().classes('q-dialog--maximized') as diag, ui.card().classes('w-full').style(
            f'background: {THEME["bg_panel"]}; max-width: 1100px; max-height: 90vh;'
        ):
            with ui.row().classes('w-full items-center justify-between q-mb-sm'):
                with ui.column().classes('gap-0'):
                    ui.label('🏛️ CADASTRO MESTRE DE AUTORIDADES').classes('text-lg font-bold text-indigo')
                    ui.label(f'Base permanente de convidados | Solenidade Ativa: {event.get("nome", "N/I")}').classes('text-xs text-grey-4')
                ui.button(icon='close', on_click=diag.close).props('flat round dense text-color=grey')

            db = get_service_db_connection() or get_db_connection()
            master_list = []
            if db:
                try:
                    res_m = db.table('jade_autoridades_base').select('*').order('categoria', desc=False).execute()
                    master_list = res_m.data if res_m.data else []
                except Exception as e_m:
                    print(f"[MASTER FETCH ERR] {e_m}")

            # Se a base mestre estiver vazia, sugerir sincronizar os convidados do evento ativo
            with ui.row().classes('w-full items-center justify-between bg-black/30 q-pa-sm rounded-lg border border-white/10 q-mb-md wrap gap-2'):
                ui.label(f'📊 {len(master_list)} autoridades cadastradas na Base Mestre').classes('text-xs font-bold text-white')
                
                with ui.row().classes('gap-2 wrap'):
                    async def save_current_to_master():
                        if not db:
                            return
                        res_c = db.table('jade_convidados').select('*').eq('evento_id', event['id']).execute()
                        convs = res_c.data if res_c.data else []
                        if not convs:
                            ui.notify('⚠️ Nenhum convidado no evento ativo para exportar.', color='warning')
                            return
                        
                        count_added = 0
                        for c in convs:
                            if c.get('convidado_principal_id'):
                                continue
                            nome = (c.get('nome') or '').strip()
                            posto = (c.get('posto_graduacao') or '').strip()
                            cat = c.get('categoria', 'Geral')
                            cargo = c.get('cargo_funcao', '')
                            
                            res_chk = db.table('jade_autoridades_base').select('id').eq('nome', nome).eq('posto_graduacao', posto).execute()
                            if not res_chk.data:
                                try:
                                    db.table('jade_autoridades_base').insert({
                                        'nome': nome,
                                        'posto_graduacao': posto,
                                        'cargo_funcao': cargo,
                                        'categoria': cat,
                                        'max_acompanhantes': c.get('max_acompanhantes', 0)
                                    }).execute()
                                    count_added += 1
                                except Exception as err_m:
                                    print(f"[SAVE MASTER ERR] {err_m}")
                        
                        ui.notify(f'✅ {count_added} novas autoridades salvas no Cadastro Mestre!', color='positive')
                        diag.close()
                        render_content.refresh()

                    ui.button('💾 Salvar Convidados Deste Evento no Mestre', icon='save', on_click=save_current_to_master).props('unelevated color=indigo dense').classes('text-xs')

            categories = {}
            for m in master_list:
                cat = m.get('categoria', 'Geral') or 'Geral'
                if cat not in categories:
                    categories[cat] = []
                categories[cat].append(m)

            selected_master_ids = set()

            with ui.scroll_area().classes('w-full').style('max-height: 55vh;'):
                if not master_list:
                    with ui.column().classes('w-full items-center justify-center q-py-xl gap-2 text-grey-4'):
                        ui.icon('account_balance', size='3rem', color='indigo')
                        ui.label('O Cadastro Mestre está vazio. Clique no botão acima para exportar os convidados do evento atual para a Base Mestre.').classes('text-xs text-center max-w-md')
                else:
                    for cat_name, items in sorted(categories.items()):
                        with ui.expansion(f'📁 {cat_name} ({len(items)} autoridades)', value=True).classes('w-full q-mb-xs text-white font-bold').style('background: rgba(63,81,181,0.15); border-radius: 8px;'):
                            for item in items:
                                m_id = str(item['id'])
                                is_sel = m_id in selected_master_ids
                                
                                def make_master_toggle(mid=m_id):
                                    def _toggle(e):
                                        if e.value:
                                            selected_master_ids.add(mid)
                                        else:
                                            selected_master_ids.discard(mid)
                                    return _toggle

                                with ui.row().classes('w-full items-center q-py-xs q-px-sm rounded-lg gap-2 border-b border-white/5'):
                                    ui.checkbox('', value=is_sel, on_change=make_master_toggle()).props('dense dark')
                                    ui.label(f"{item.get('posto_graduacao', '') or ''}").classes('text-[10px] text-amber font-bold').style('min-width: 60px;')
                                    ui.label(f"{item.get('nome', '')}").classes('text-xs text-white font-bold flex-grow')
                                    if item.get('cargo_funcao'):
                                        ui.label(f"{item['cargo_funcao'][:40]}").classes('text-[9px] text-grey-4 truncate').style('max-width: 200px;')

            ui.separator().classes('q-my-sm')
            with ui.row().classes('w-full justify-between items-center'):
                async def import_selected_from_master():
                    if not selected_master_ids:
                        ui.notify('⚠️ Nenhuma autoridade selecionada.', color='warning')
                        return
                    if not db:
                        return
                    
                    event_id = event['id']
                    count_imported = 0
                    for mid in list(selected_master_ids):
                        m_item = next((x for x in master_list if str(x['id']) == mid), None)
                        if m_item:
                            nome = m_item['nome']
                            posto = m_item.get('posto_graduacao', '')
                            cargo = m_item.get('cargo_funcao', '')
                            cat = m_item.get('categoria', 'Geral')
                            max_ac = m_item.get('max_acompanhantes', 0)
                            
                            res_chk = db.table('jade_convidados').select('id').eq('evento_id', event_id).eq('nome', nome).eq('posto_graduacao', posto).execute()
                            if not res_chk.data:
                                try:
                                    res_ins = db.table('jade_convidados').insert({
                                        'evento_id': event_id,
                                        'nome': nome,
                                        'posto_graduacao': posto,
                                        'cargo_funcao': cargo,
                                        'categoria': cat,
                                        'max_acompanhantes': max_ac,
                                        'status_confirmacao': 'pendente',
                                        'status_placa': 'nao_necessaria'
                                    }).execute()
                                    if res_ins.data:
                                        sync_companions(res_ins.data[0]['id'], nome, max_ac, event_id, cat)
                                    count_imported += 1
                                except Exception as err_imp:
                                    print(f"[IMPORT FROM MASTER ERR] {err_imp}")
                    
                    ui.notify(f'✅ {count_imported} autoridades importadas do Cadastro Mestre para o evento!', color='positive')
                    selected_master_ids.clear()
                    diag.close()
                    render_content.refresh()

                ui.button('📥 Puxar Selecionadas para Este Evento', icon='download', on_click=import_selected_from_master).props('unelevated color=indigo text-color=white dense').classes('text-xs')
                ui.button('Fechar', icon='close', on_click=diag.close).props('unelevated color=grey-8 dense').classes('text-xs')

        diag.open()


    # ═══════════════════════════════════════════════════════════════
    # FASE 6: PLACA EXPRESS (GERAÇÃO E IMPRESSÃO EM 1 CLIQUE)
    # ═══════════════════════════════════════════════════════════════
    def open_direct_print_preview_dialog(event, guest_list):
        """Abre diretamente a janela de impressão limpa via iframe de mesma origem (1 clique)."""
        js_print_code = """
        (function() {
            var area = document.querySelector('.print-area-express');
            if (!area) { area = document.querySelector('.print-area'); }
            if (!area) { window.print(); return; }

            var oldIframe = document.getElementById('jade_express_iframe');
            if (oldIframe) { oldIframe.remove(); }

            var iframe = document.createElement('iframe');
            iframe.id = 'jade_express_iframe';
            iframe.style.position = 'fixed';
            iframe.style.right = '0';
            iframe.style.bottom = '0';
            iframe.style.width = '0';
            iframe.style.height = '0';
            iframe.style.border = '0';
            document.body.appendChild(iframe);

            var cssStyles = Array.from(document.querySelectorAll('style, link[rel="stylesheet"]'))
                                 .map(s => s.outerHTML).join('\\n');

            var doc = iframe.contentWindow.document;
            doc.open();
            doc.write(`
                <!DOCTYPE html>
                <html>
                <head>
                    <title>JADE - Placa Express</title>
                    ${cssStyles}
                    <style>
                        @page { size: A4 portrait; margin: 5mm; }
                        * { -webkit-print-color-adjust: exact !important; print-color-adjust: exact !important; }
                        body { margin: 0 !important; padding: 4mm 6mm !important; background: #ffffff !important; color: #000000 !important; font-family: Arial, sans-serif !important; }
                        .print-hide, .q-header, .q-drawer, .q-footer { display: none !important; }
                        .print-area-express { display: block !important; position: static !important; width: 100% !important; visibility: visible !important; }
                        .prisma-card-a4-slot { height: 66mm !important; max-height: 66mm !important; border: 1.5pt solid #1a1a1a !important; outline: 0.5pt solid #1a1a1a !important; outline-offset: -2.5mm !important; margin-bottom: 4.5mm !important; page-break-inside: avoid !important; background: #ffffff !important; color: #000000 !important; display: flex !important; flex-direction: column !important; justify-content: center !important; align-items: center !important; position: relative !important; box-sizing: border-box !important; }
                        .prisma-conteudo-central { display: flex !important; flex-direction: column !important; align-items: center !important; justify-content: center !important; text-align: center !important; width: 100% !important; }
                        .prisma-texto-reservado { font-weight: 900 !important; letter-spacing: 3px !important; text-transform: uppercase !important; color: #1f4e79 !important; font-size: 20pt !important; margin-bottom: 2px !important; }
                        .prisma-posto-extenso { font-weight: bold !important; text-transform: uppercase !important; letter-spacing: 1.5px !important; font-size: 18pt !important; margin-bottom: 2px !important; }
                        .prisma-nome-autoridade { font-weight: 900 !important; text-transform: uppercase !important; font-size: 32pt !important; line-height: 1.05 !important; }
                        img { max-width: 100% !important; display: inline-block !important; visibility: visible !important; opacity: 1 !important; -webkit-print-color-adjust: exact !important; print-color-adjust: exact !important; }
                    </style>
                </head>
                <body>
                    <div class="print-area-express">
                        ${area.innerHTML}
                    </div>
                </body>
                </html>
            `);
            doc.close();

            setTimeout(function() {
                iframe.contentWindow.focus();
                iframe.contentWindow.print();
            }, 1200);
        })();
        """

        # Carregar brasão padrão CGCFN em base64 offline
        brasao_cfn_b64 = url_to_base64('assets/brasao_cgcfn.png')

        with ui.dialog() as print_diag, ui.card().classes('q-pa-md').style('min-width: 820px; max-width: 96vw; max-height: 90vh; overflow-y: auto; background: #fff; color: #000;'):
            with ui.row().classes('w-full justify-between items-center q-mb-md print-hide'):
                ui.label('⚡ IMPRESSÃO DIRETA DE EMERGÊNCIA (1 CLIQUE)').classes('text-md font-bold text-deep-orange')
                with ui.row().classes('items-center gap-2'):
                    ui.button('🖨️ DISPARAR IMPRESSORA AGORA', icon='print', on_click=lambda: ui.run_javascript(js_print_code)).props('unelevated color=deep-orange text-color=white bold').classes('text-xs')
                    ui.button('Fechar', on_click=print_diag.close).props('unelevated color=grey-8 dense').classes('text-xs')

            # Auto-dispara iframe de impressão após 500ms (aguarda DOM renderizar)
            ui.timer(0.5, lambda: ui.run_javascript(js_print_code), once=True)

            with ui.column().classes('w-full print-area-express gap-4'):
                for c in guest_list:
                    is_acomp = bool(c.get('convidado_principal_id'))
                    posto = c.get('posto_graduacao') or ''
                    almirantado_info = parse_almirantado_stars(posto)
                    nome_limpo = clean_authority_name(c['nome'])
                    assento = c.get('assento_id', '')

                    # QR Code 100% offline em Base64
                    qr_b64 = gen_qr_base64(f"JADE|{event.get('id','')}|{c['id']}|{assento}")

                    # Insígnia em base64 se disponível
                    rank_logo_b64 = url_to_base64(almirantado_info['png_asset']) if almirantado_info.get('png_asset') else None

                    with ui.element('div').classes('prisma-card-a4-slot').style('border: 1.5pt solid #1a1a1a; outline: 0.5pt solid #1a1a1a; outline-offset: -2.5mm; position: relative; margin-bottom: 12px; font-family: Arial, sans-serif;'):
                        # Canto Esquerdo: Brasão + Estrelas (Base64 offline)
                        with ui.element('div').style('position: absolute; top: 4mm; left: 6mm; z-index: 10; display: flex; flex-direction: column; align-items: flex-start;'):
                            ui.image(brasao_cfn_b64).style('width: 16mm; height: auto;')
                            if rank_logo_b64:
                                ui.image(rank_logo_b64).style('width: 16mm; height: auto; margin-top: 1mm;')
                            elif almirantado_info['stars']:
                                ui.label(almirantado_info['stars']).classes('prisma-estrelas-esquerda')

                        # Canto Direito: QR Code (Base64 offline)
                        with ui.element('div').style('position: absolute; top: 3mm; right: 5mm; z-index: 10; font-size: 7px; text-align: center;'):
                            ui.image(qr_b64).style('width: 13mm; height: 13mm;')
                            if assento:
                                ui.label(assento).classes('text-[7px] font-mono text-black font-bold')

                        # Centro: Reservado + Posto + Nome
                        with ui.element('div').classes('prisma-conteudo-central'):
                            if is_acomp:
                                ui.label('RESERVADO').classes('prisma-texto-reservado')
                                if almirantado_info['title']:
                                    ui.label(almirantado_info['title']).classes('prisma-posto-extenso')
                                elif posto:
                                    ui.label(posto.upper()).classes('prisma-posto-extenso')
                                ui.label(nome_limpo).classes('prisma-nome-autoridade')
                            else:
                                if almirantado_info['title']:
                                    ui.label(almirantado_info['title']).classes('prisma-posto-extenso')
                                elif posto:
                                    ui.label(posto.upper()).classes('prisma-posto-extenso')
                                ui.label(nome_limpo).classes('prisma-nome-autoridade')

        print_diag.open()



    def open_express_plate_dialog(event, convidados, layout):
        """Dialog de busca tática ultra-rápida e confecção de placas expressas avulsas de última hora."""
        search_input = {'text': ''}
        
        with ui.dialog() as diag, ui.card().classes('w-full').style(f'background: {THEME["bg_panel"]}; max-width: 720px;'):
            with ui.row().classes('w-full items-center justify-between q-mb-xs'):
                with ui.column().classes('gap-0'):
                    ui.label('⚡ PLACA EXPRESS (CONFECÇÃO E IMPRESSÃO DE 1 CLIQUE)').classes('text-md font-bold text-deep-orange cyber-title')
                    ui.label('Busca rápida ou criação avulsa para confecção e disparo imediato de impressora').classes('text-[11px] text-grey-4')
                ui.button(icon='close', on_click=diag.close).props('flat round dense text-color=grey')

            with ui.tabs().classes('w-full text-cyan') as express_tabs:
                tab_search = ui.tab('🔍 Buscar e Imprimir Direto')
                tab_create = ui.tab('➕ Nova Placa Avulsa de Emergência')

            with ui.tab_panels(express_tabs, value=tab_search).classes('w-full bg-transparent'):
                # TAB 1: BUSCAR E IMPRIMIR DIRETO
                with ui.tab_panel(tab_search):
                    ui.input(
                        placeholder='🔍 Digite o nome ou posto da autoridade...',
                        on_change=lambda e: [search_input.update({'text': e.value}), results_container.refresh()]
                    ).props('dark outlined dense autofocus').classes('w-full q-mb-sm text-xs')

                    @ui.refreshable
                    def results_container():
                        query = search_input['text'].strip().lower()
                        if not query:
                            ui.label('💡 Digite acima para buscar entre as autoridades já cadastradas.').classes('text-xs text-grey-5 q-py-md text-center w-full')
                            return

                        matches = [
                            c for c in convidados
                            if query in (c.get('nome') or '').lower()
                            or query in (c.get('posto_graduacao') or '').lower()
                            or query in (c.get('cargo_funcao') or '').lower()
                        ]

                        if not matches:
                            ui.label(f'Nenhum cadastro encontrado para "{query}". Utilize a aba "➕ Nova Placa Avulsa" acima para criar na hora!').classes('text-xs text-amber font-bold text-center w-full q-py-md')
                            return

                        with ui.column().classes('w-full gap-1 max-h-[350px] overflow-y-auto q-pr-xs'):
                            for m in matches[:15]:
                                g_id = m['id']
                                g_nome = m.get('nome', 'N/I')
                                g_posto = m.get('posto_graduacao', '') or ''
                                g_cargo = m.get('cargo_funcao', '') or ''
                                g_assento = m.get('assento_id', '') or 'Sem Assento'
                                is_acomp = bool(m.get('convidado_principal_id'))
                                
                                with ui.card().classes('w-full q-pa-xs no-shadow rounded-lg border border-white/10').style('background: rgba(255,255,255,0.03);'):
                                    with ui.row().classes('w-full items-center justify-between wrap gap-1'):
                                        with ui.column().classes('gap-0 flex-grow'):
                                            with ui.row().classes('items-center gap-1'):
                                                if is_acomp:
                                                    ui.badge('RESERVADO').props('color=amber-9 text-color=white').classes('text-[9px]')
                                                ui.label(f"{g_posto} {g_nome}").classes('text-xs font-bold text-white')
                                            ui.label(f"{g_cargo} • 🪑 Assento: {g_assento}").classes('text-[10px] text-cyan')

                                        with ui.row().classes('items-center gap-1'):
                                            # Botão ➕ Placa Extra (Adicionar Acompanhante)
                                            if not is_acomp:
                                                async def add_extra_companion(principal_id, main_nome, main_posto):
                                                    _db = get_service_db_connection() or get_db_connection()
                                                    if _db:
                                                        res_ac = _db.table('jade_convidados').select('*').eq('convidado_principal_id', principal_id).execute()
                                                        count_ac = len(res_ac.data or []) + 1
                                                        ins_res = _db.table('jade_convidados').insert({
                                                            'evento_id': event['id'],
                                                            'nome': f"ACOMP. {main_nome} ({count_ac})",
                                                            'posto_graduacao': main_posto,
                                                            'convidado_principal_id': principal_id,
                                                            'status_placa': 'impressa',
                                                            'presenca_confirmada': True
                                                        }).execute()
                                                        ui.notify(f'➕ Placa extra criada para {main_nome}!', color='positive')
                                                        render_content.refresh()
                                                        if ins_res and ins_res.data:
                                                            open_direct_print_preview_dialog(event, [ins_res.data[0]])

                                                ui.button('➕ Extra', on_click=lambda p_id=g_id, m_n=g_nome, m_p=g_posto: add_extra_companion(p_id, m_n, m_p)).props('unelevated color=cyan text-color=black dense bold').classes('text-[10px] q-px-xs').tooltip('Criar e imprimir placa de acompanhante extra na hora')

                                            # Botão 🖨️ IMPRIMIR DIRETO (1 CLIQUE)
                                            async def make_print_express_direct(guest=m):
                                                _db = get_service_db_connection() or get_db_connection()
                                                if _db:
                                                    try:
                                                        _db.table('jade_convidados').update({'status_placa': 'impressa'}).eq('id', guest['id']).execute()
                                                    except Exception:
                                                        pass
                                                open_direct_print_preview_dialog(event, [guest])

                                            ui.button('🖨️ IMPRIMIR AGORA', icon='print', on_click=make_print_express_direct).props('unelevated color=deep-orange text-color=white dense bold').classes('text-[10px] q-px-xs').tooltip('Disparar impressora imediatamente (1 clique)')
                                            
                                            # Opção opcional de abrir o estúdio completo
                                            ui.button('⚙️', on_click=lambda guest=m: open_print_cards_dialog(event, [guest], layout)).props('flat round dense color=grey-4 size=xs').tooltip('Abrir configurações avançadas de estúdio')

                    results_container()

                # TAB 2: NOVA IMPRESSÃO AVULSA DE EMERGÊNCIA
                with ui.tab_panel(tab_create):
                    with ui.column().classes('w-full gap-2 q-pa-xs'):
                        ui.label('Cadastre uma nova autoridade avulsa para confecção e disparo imediato:').classes('text-xs text-grey-4')

                        with ui.row().classes('w-full gap-2 wrap'):
                            sel_posto_express = ui.select(
                                options={
                                    'AE - Almirante de Esquadra': '⚓ AE - Almirante de Esquadra',
                                    'VA - Vice-Almirante': '⚓ VA - Vice-Almirante',
                                    'CA - Contra-Almirante': '⚓ CA - Contra-Almirante',
                                    'CMG - Capitão de Mar e Guerra': '🎖️ CMG - Capitão de Mar e Guerra',
                                    'CF - Capitão de Fragata': '🎖️ CF - Capitão de Fragata',
                                    'CC - Capitão de Corveta': '🎖️ CC - Capitão de Corveta',
                                    'CT - Capitão-Tenente': '🎖️ CT - Capitão-Tenente',
                                    'Desembargador(a)': '🏛️ Desembargador(a)',
                                    'Senador(a) / Deputado(a)': '🏛️ Senador(a) / Deputado(a)',
                                    'Juiz(a) de Direito': '🏛️ Juiz(a) de Direito',
                                    'Senhor / Senhora': '👤 Senhor / Senhora',
                                    'AUTORIDADE': '🎖️ Outra Autoridade'
                                },
                                value='AE - Almirante de Esquadra'
                            ).props('dark outlined dense').classes('col')

                            input_nome_express = ui.input(placeholder='Nome Completo da Autoridade...').props('dark outlined dense').classes('col')

                        with ui.row().classes('w-full gap-2 wrap items-center'):
                            input_assento_express = ui.input(placeholder='Assento (opcional, ex: A-1)...').props('dark outlined dense').style('width: 180px;')
                            sel_acomp_express = ui.select(
                                options={0: '0 Acompanhantes', 1: '1 Acompanhante Extra', 2: '2 Acompanhantes Extras', 3: '3 Acompanhantes Extras'},
                                value=0
                            ).props('dark outlined dense').classes('col')

                        async def create_and_print_avulso():
                            nome_val = input_nome_express.value.strip().upper()
                            if not nome_val:
                                ui.notify('⚠️ Digite o Nome da Autoridade!', color='warning')
                                return
                            posto_val = sel_posto_express.value
                            assento_val = input_assento_express.value.strip().upper()
                            num_acomp = int(sel_acomp_express.value or 0)

                            _db = get_service_db_connection() or get_db_connection()
                            if not _db:
                                ui.notify('❌ Banco de dados indisponível.', color='negative')
                                return

                            try:
                                main_ins = _db.table('jade_convidados').insert({
                                    'evento_id': event['id'],
                                    'nome': nome_val,
                                    'posto_graduacao': posto_val,
                                    'assento_id': assento_val if assento_val else None,
                                    'status_confirmacao': 'confirmado',
                                    'status_placa': 'impressa',
                                    'max_acompanhantes': num_acomp
                                }).execute()

                                print_list = []
                                if main_ins and main_ins.data:
                                    main_obj = main_ins.data[0]
                                    print_list.append(main_obj)
                                    main_id = main_obj['id']

                                    # Cria acompanhantes
                                    for i in range(num_acomp):
                                        ac_ins = _db.table('jade_convidados').insert({
                                            'evento_id': event['id'],
                                            'nome': f"ACOMP. {nome_val} ({i+1}/{num_acomp})",
                                            'posto_graduacao': posto_val,
                                            'convidado_principal_id': main_id,
                                            'status_confirmacao': 'confirmado',
                                            'status_placa': 'impressa'
                                        }).execute()
                                        if ac_ins and ac_ins.data:
                                            print_list.append(ac_ins.data[0])

                                ui.notify(f'⚡ {nome_val} cadastrado com sucesso!', color='positive')
                                render_content.refresh()
                                diag.close()
                                open_direct_print_preview_dialog(event, print_list)
                            except Exception as ex_av:
                                ui.notify(f'Erro ao cadastrar: {ex_av}', color='negative')

                        ui.button('🖨️ GERAR E IMPRIMIR PLACA EXPRESSA AGORA (1 CLIQUE)', icon='flash_on', on_click=create_and_print_avulso).props('unelevated color=deep-orange text-color=white bold').classes('w-full q-mt-sm')

            with ui.row().classes('w-full justify-end q-mt-sm'):
                ui.button('Fechar', icon='close', on_click=diag.close).props('unelevated color=grey-8 dense').classes('text-xs')

        diag.open()

def gerar_pdf_placas_reportlab(event, convidados, current_model, only_confirmed, print_config):
    """Gera PDF vetorial em milímetros A4 nativamente no servidor Python via ReportLab."""
    try:
        import io, re, base64, os
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image as RLImage
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=10, rightMargin=10, topMargin=10, bottomMargin=10)
        elements = []
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            'PrismaTitle', parent=styles['Normal'],
            fontName='Helvetica-Bold', fontSize=14, leading=16,
            textColor=colors.HexColor('#1f4e79'), alignment=TA_CENTER
        )
        posto_style = ParagraphStyle(
            'PrismaPosto', parent=styles['Normal'],
            fontName='Helvetica-Bold', fontSize=15, leading=18,
            textColor=colors.HexColor('#000000'), alignment=TA_CENTER
        )
        nome_style = ParagraphStyle(
            'PrismaNome', parent=styles['Normal'],
            fontName='Helvetica-Bold', fontSize=24, leading=28,
            textColor=colors.HexColor('#000000'), alignment=TA_CENTER
        )
        sub_style = ParagraphStyle(
            'PrismaSub', parent=styles['Normal'],
            fontName='Helvetica', fontSize=10, leading=12,
            textColor=colors.HexColor('#444444'), alignment=TA_CENTER
        )

        logo_path = 'assets/brasao_cgcfn.png'
        rl_logo = None
        if os.path.exists(logo_path):
            try: rl_logo = RLImage(logo_path, width=32, height=32)
            except Exception: pass

        # Coleta e ordena convidados
        all_cards = []
        seen_ids = set()
        for c in (convidados or []):
            if c['id'] not in seen_ids:
                if only_confirmed and c.get('status_placa') not in ('pendente', 'em_producao', 'impressa', 'reimpressao'):
                    continue
                seen_ids.add(c['id'])
                all_cards.append(c)

        def sort_key_assento(c):
            ass = str(c.get('assento_id', '')).upper().strip()
            match = re.match(r'([A-Z]+)-?(\d+)', ass)
            if match:
                row, num = match.groups()
                return (row, int(num))
            return (ass if ass else 'ZZZ', 0)

        all_cards.sort(key=sort_key_assento)
        if not all_cards:
            return None

        items_per_sheet = 4
        total_pages = (len(all_cards) + items_per_sheet - 1) // items_per_sheet

        for p_idx in range(total_pages):
            batch = all_cards[p_idx * items_per_sheet : (p_idx + 1) * items_per_sheet]
            table_data = []
            
            for c in batch:
                is_acomp = bool(c.get('convidado_principal_id'))
                posto = (c.get('posto_graduacao') or '').strip()
                almirantado_info = parse_almirantado_stars(posto)
                nome_limpo = clean_authority_name(c['nome'])
                termo_reservado = print_config.get('termo_convidado', 'RESERVADO')
                
                qr_rl_img = None
                try:
                    import qrcode
                    qr_obj = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=3, border=1)
                    qr_obj.add_data(f"JADE|{event.get('id','')}|{c['id']}|{c.get('assento_id','')}")
                    qr_obj.make(fit=True)
                    img_q = qr_obj.make_image(fill_color="black", back_color="white")
                    q_buf = io.BytesIO()
                    img_q.save(q_buf, format="PNG")
                    q_buf.seek(0)
                    qr_rl_img = RLImage(q_buf, width=36, height=36)
                except Exception: pass

                cell_content = []
                if is_acomp:
                    cell_content.append(Paragraph(termo_reservado, title_style))
                    cell_content.append(Spacer(1, 4))

                tit_str = almirantado_info['title'] or posto.upper()
                if tit_str:
                    cell_content.append(Paragraph(tit_str, posto_style))
                    cell_content.append(Spacer(1, 4))

                cell_content.append(Paragraph(nome_limpo, nome_style))
                
                if c.get('assento_id'):
                    cell_content.append(Spacer(1, 4))
                    cell_content.append(Paragraph(f"ASSENTO: {c.get('assento_id')}", sub_style))

                inner_row = []
                inner_row.append([rl_logo] if rl_logo else [''])
                inner_row.append(cell_content)
                inner_row.append([qr_rl_img] if qr_rl_img else [''])

                inner_table = Table(
                    [[inner_row[0][0], inner_row[1], inner_row[2][0]]],
                    colWidths=[40, 480, 45]
                )
                inner_table.setStyle(TableStyle([
                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                    ('ALIGN', (0,0), (0,0), 'LEFT'),
                    ('ALIGN', (1,0), (1,0), 'CENTER'),
                    ('ALIGN', (2,0), (2,0), 'RIGHT'),
                ]))

                table_data.append([inner_table])

            while len(table_data) < 4:
                table_data.append([''])

            page_table = Table(table_data, colWidths=[570], rowHeights=[185, 185, 185, 185])
            page_table.setStyle(TableStyle([
                ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#1a1a1a')),
                ('INNERGRID', (0,0), (-1,-1), 1, colors.HexColor('#1a1a1a')),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('TOPPADDING', (0,0), (-1,-1), 10),
                ('BOTTOMPADDING', (0,0), (-1,-1), 10),
            ]))

            elements.append(page_table)
            if p_idx < total_pages - 1:
                elements.append(PageBreak())

        doc.build(elements)
        buf.seek(0)
        return buf.getvalue()
    except Exception as pdf_e:
        print(f"[PDF REPORTLAB GERAR ERR] {pdf_e}")
        return None


def open_print_cards_dialog(event, convidados, layout):
    rows_count = layout.get('rows', 5)

    allocated_by_row = {}
    for r in range(rows_count):
        row_label = get_row_label(r)
        allocated_by_row[row_label] = [c for c in convidados if (c.get('assento_id') or '').startswith(f"{row_label}-")]

    # Estado e Persistência do Estúdio de Impressão
    saved_cfg = {}
    if isinstance(layout, dict):
        saved_cfg = layout.get('print_config', {})
    elif isinstance(event, dict) and event.get('layout_json'):
        try:
            import json
            l_dict = json.loads(event.get('layout_json') or '{}')
            saved_cfg = l_dict.get('print_config', {})
        except Exception:
            saved_cfg = {}

    print_config = {
        'model': saved_cfg.get('model', 'prisma_a4_4slots'),
        'items_per_sheet': int(saved_cfg.get('items_per_sheet', 4)),
        'chk_only_confirmed': bool(saved_cfg.get('chk_only_confirmed', False)),
        'chk_logo': bool(saved_cfg.get('chk_logo', True)),
        'chk_qr': bool(saved_cfg.get('chk_qr', True)),
        'chk_rank': bool(saved_cfg.get('chk_rank', True)),
        'chk_border': bool(saved_cfg.get('chk_border', True)),
        'header_line1': str(saved_cfg.get('header_line1', 'MARINHA DO BRASIL')),
        'header_line2': str(saved_cfg.get('header_line2', event.get('nome', 'SOLENIDADE').upper())),
        'termo_convidado': str(saved_cfg.get('termo_convidado', 'RESERVADO')),
        'brasao_pos': str(saved_cfg.get('brasao_pos', 'esquerda')),
        'origin_logo_l': str(saved_cfg.get('origin_logo_l', 'bucket')),
        'logo_preset_l': str(saved_cfg.get('logo_preset_l', 'cfn')),
        'upload_brasao_left': str(saved_cfg.get('upload_brasao_left', '')),
        'qr_pos': str(saved_cfg.get('qr_pos', 'direita')),
        'origin_logo_r': str(saved_cfg.get('origin_logo_r', 'bucket')),
        'logo_preset_r': str(saved_cfg.get('logo_preset_r', 'mb')),
        'upload_brasao_right': str(saved_cfg.get('upload_brasao_right', '')),
        'origin_bg': str(saved_cfg.get('origin_bg', 'none')),
        'bg_preset': str(saved_cfg.get('bg_preset', 'cfn')),
        'template_bg_url': str(saved_cfg.get('template_bg_url', '')),
        'logo_width': float(saved_cfg.get('logo_width', 16)),
        'logo_pos_x': float(saved_cfg.get('logo_pos_x', 6)),
        'logo_pos_y': float(saved_cfg.get('logo_pos_y', 4)),
        'qr_size': float(saved_cfg.get('qr_size', 13)),
        'qr_pos_x': float(saved_cfg.get('qr_pos_x', 5)),
        'qr_pos_y': float(saved_cfg.get('qr_pos_y', 3)),
        'font_nome': float(saved_cfg.get('font_nome', 22)),
        'font_posto': float(saved_cfg.get('font_posto', 13)),
        'font_reservado': float(saved_cfg.get('font_reservado', 18)),
    }

    # Configurações Padrão de Prisma
    CONFIG_DEFAULT = {
        'formato_folha': 'A4_PORTRAIT_4_SLOTS',
        'exibir_borda_dupla': True,
        'brasao_padrao': 'assets/brasao_cgcfn.png',
        'gap_linhas_mm': 6,
        'termo_convidado': 'RESERVADO',
        'fonte_familia': 'Montserrat, sans-serif',
    }

    # Insígnias oficiais por Posto/Graduação (texto rico com indicador visual)
    RANK_INSIGNIAS = {
        'AE':  {'stars': '★★★★', 'title': 'ALMIRANTE DE ESQUADRA',       'color': '#FFD700'},
        'VA':  {'stars': '★★★',  'title': 'VICE-ALMIRANTE',              'color': '#FFD700'},
        'CA':  {'stars': '★★',   'title': 'CONTRA-ALMIRANTE',            'color': '#FFD700'},
        'CMG': {'stars': '★',    'title': 'CAPITÃO DE MAR E GUERRA',     'color': '#C0C0C0'},
        'CF':  {'stars': '⚓',   'title': 'CAPITÃO DE FRAGATA',          'color': '#C0C0C0'},
        'CC':  {'stars': '⚓',   'title': 'CAPITÃO DE CORVETA',          'color': '#C0C0C0'},
        'CT':  {'stars': '⚓',   'title': 'CAPITÃO-TENENTE',             'color': '#B0B0B0'},
        '1TEN': {'stars': '▬',   'title': '1º TENENTE',                  'color': '#B0B0B0'},
        '2TEN': {'stars': '▬',   'title': '2º TENENTE',                  'color': '#B0B0B0'},
        'SO':  {'stars': '◆',    'title': 'SUBOFICIAL',                  'color': '#CD7F32'},
        '1SG': {'stars': '▲▲▲',  'title': '1º SARGENTO',                 'color': '#CD7F32'},
        '2SG': {'stars': '▲▲',   'title': '2º SARGENTO',                 'color': '#CD7F32'},
        '3SG': {'stars': '▲',    'title': '3º SARGENTO',                 'color': '#CD7F32'},
        'CB':  {'stars': '∨∨',   'title': 'CABO',                        'color': '#808080'},
        'SD':  {'stars': '∨',    'title': 'SOLDADO',                     'color': '#808080'},
        'MN':  {'stars': '∨',    'title': 'MARINHEIRO',                  'color': '#808080'},
        'Dr.': {'stars': '⚖️',   'title': 'AUTORIDADE CIVIL',            'color': '#4A90D9'},
        'Min.':{'stars': '🏛️',   'title': 'MINISTRO DE ESTADO',          'color': '#9B59B6'},
        'Dep.':{'stars': '🏛️',   'title': 'DEPUTADO',                    'color': '#27AE60'},
        'Sen.':{'stars': '🏛️',   'title': 'SENADOR',                     'color': '#2980B9'},
        'Gen.':{'stars': '★★★★', 'title': 'GENERAL DE EXÉRCITO',         'color': '#FFD700'},
        'Cel.':{'stars': '★',    'title': 'CORONEL',                     'color': '#C0C0C0'},
        'TC':  {'stars': '★',    'title': 'TENENTE-CORONEL',             'color': '#C0C0C0'},
        'Maj': {'stars': '★',    'title': 'MAJOR',                       'color': '#C0C0C0'}
    }


    import os
    import re

    def clean_authority_name(raw_name):
        if not raw_name:
            return ""
        name = str(raw_name).strip()
        name = re.sub(r'^(ACOMP\.|ACOMPANHANTE)\s*', '', name, flags=re.IGNORECASE)
        name = re.sub(r'\s*\(\d+(/\d+)?\)$', '', name).strip()
        return name.upper()

    def get_rank_logo_asset(posto_str):
        if not posto_str:
            return None
        p = str(posto_str).upper().strip()
        
        sigla = None
        if any(k in p for k in ['ESQUADRA', 'SQUADRA', 'AE', 'ALMIRANTE DE ESQUADRA']):
            sigla = 'AE'
        elif any(k in p for k in ['VICE', 'VADM', 'V-ADM', 'VA', 'VICE-ALMIRANTE']):
            sigla = 'VA'
        elif any(k in p for k in ['CONTRA', 'CALTE', 'C-ADM', 'CA', 'CONTRA-ALMIRANTE']):
            sigla = 'CA'
        elif any(k in p for k in ['MAR E GUERRA', 'CMG']):
            sigla = 'CMG'
        elif any(k in p for k in ['FRAGATA', 'CF']):
            sigla = 'CF'
        elif any(k in p for k in ['CORVETA', 'CC']):
            sigla = 'CC'
        elif any(k in p for k in ['TENENTE', 'CT']):
            sigla = 'CT'
        else:
            sigla = p.split()[0] if p else None

        if not sigla:
            return None

        sigla_clean = re.sub(r'\W+', '', sigla).upper()


            # 1. Procura no bucket 'logos' do Supabase por ex: AE.png, AE.PNG, ae.png
        try:
            from database import list_supabase_storage_files
            bucket_files = list_supabase_storage_files("logos")
            for f in bucket_files:
                fname = f.get('name', '')
                fname_no_ext = os.path.splitext(fname)[0].upper()
                if fname_no_ext == sigla_clean:
                    return f.get('url')
        except Exception as b_err:
            print(f"[RANK LOGO BUCKET ERR] {b_err}")

        # 2. Procura localmente em assets/insignias/
        possible_names = [f"{sigla_clean.lower()}.png", f"{sigla_clean}.png", f"{sigla_clean.lower()}.jpg", f"{sigla_clean}.jpg"]
        for p_name in possible_names:
            local_p = os.path.join('assets', 'insignias', p_name)
            if os.path.exists(local_p):
                return local_p

        return None

    def parse_almirantado_stars(posto_str):
        if not posto_str:
            return {'eh_almirante': False, 'stars': '', 'title': '', 'color': '#000000', 'png_asset': None}
        p = str(posto_str).upper().strip()
        
        png_path = get_rank_logo_asset(p)
        
        if any(k in p for k in ['ESQUADRA', 'SQUADRA', 'AE', 'ALMIRANTE DE ESQUADRA']):
            return {'eh_almirante': True, 'stars': '★ ★ ★ ★', 'title': 'ALMIRANTE DE ESQUADRA', 'color': '#000000', 'png_asset': png_path}
        elif any(k in p for k in ['VICE', 'VADM', 'V-ADM', 'VA', 'VICE-ALMIRANTE']):
            return {'eh_almirante': True, 'stars': '★ ★ ★', 'title': 'VICE-ALMIRANTE', 'color': '#000000', 'png_asset': png_path}
        elif any(k in p for k in ['CONTRA', 'CALTE', 'C-ADM', 'CA', 'CONTRA-ALMIRANTE']):
            return {'eh_almirante': True, 'stars': '★ ★', 'title': 'CONTRA-ALMIRANTE', 'color': '#000000', 'png_asset': png_path}

        return {'eh_almirante': False, 'stars': '', 'title': p, 'color': '#000000', 'png_asset': png_path}


        with ui.dialog() as diag, ui.card().classes('q-pa-lg').style('min-width: 860px; max-width: 96vw; max-height: 92vh; overflow-y: auto;'):
            ui.label(f"🖨️ ESTÚDIO DE IMPRESSÃO DE PLACAS & CREDENCIAIS JADE").classes('text-md font-bold text-cyan cyber-title q-mb-xs')
            ui.label("Personalize Modelos, Templates de Fundo, Brasões, Insígnias e Posicionamento").classes('text-xs text-grey-4 q-mb-md')

            # ═══════════════════════════════════════════════════════════════
            # PAINEL DE CONFIGURAÇÃO EXPANDIDO (print-hide)
            # ═══════════════════════════════════════════════════════════════
            with ui.card().classes('w-full q-pa-md bg-black/50 border border-cyan-500/30 rounded-xl q-mb-md print-hide'):

                # ── Linha 1: Modelo + Placas por Folha + Toggles ──
                with ui.row().classes('w-full items-center gap-3 wrap'):
                    ui.label('Modelo:').classes('text-xs text-grey-3 font-bold')
                    model_select = ui.select(
                        options={
                            'prisma_a4_4slots': '🏛️ Prisma Institucional A4 (4 por Folha - Moldura Dupla)',
                            'prisma_dobravel_v': '📐 Prisma Dobrável V (Frente & Verso Invertido 180°)',
                            'jade_horizontal_a4': '📋 Placa Horizontal CGCFN (Padrão Oficial)',
                            'cadeira_a4': '📄 Placa de Cadeira (A4 Padrão)',
                            'mesa_a5_dobravel': '🏷️ Placa de Mesa Dobrável (A5)',
                            'credencial': '🪪 Credencial / Crachá de Peito',
                            'template_custom': '🎨 Template Customizado (Imagem de Fundo)'
                        },
                        value=print_config['model']
                    ).props('dark outlined dense').style('min-width: 310px;')

                    ui.label('Placas / Folha A4:').classes('text-xs text-grey-3 font-bold q-ml-xs')
                    input_items_per_sheet = ui.select(
                        options={1: '1 por Folha A4', 2: '2 por Folha A4', 4: '4 por Folha A4 (Padrão)', 6: '6 por Folha A4', 8: '8 por Folha A4'},
                        value=print_config['items_per_sheet']
                    ).props('dark outlined dense').style('width: 170px;')

                    with ui.row().classes('items-center gap-2'):
                        chk_only_confirmed = ui.checkbox('Só Confirmados / Fila Ativa', value=print_config['chk_only_confirmed']).props('dark dense').classes('text-xs text-amber-3')
                        chk_logo = ui.checkbox('Brasão MB', value=print_config['chk_logo']).props('dark dense').classes('text-xs text-grey-3')
                        chk_qr = ui.checkbox('QR Code', value=print_config['chk_qr']).props('dark dense').classes('text-xs text-grey-3')
                        chk_rank = ui.checkbox('Insígnia de Posto', value=print_config['chk_rank']).props('dark dense').classes('text-xs text-grey-3')
                        chk_border = ui.checkbox('Borda Dupla', value=print_config['chk_border']).props('dark dense').classes('text-xs text-grey-3')

                ui.separator().classes('q-my-sm').style('border-color: rgba(255,255,255,0.08);')

                # ── Linha 2: Cabeçalho e Título Editáveis ──
                with ui.row().classes('w-full gap-2'):
                    with ui.column().classes('col gap-0'):
                        ui.label('Linha 1 do Cabeçalho:').classes('text-[10px] text-grey-5')
                        input_header1 = ui.input(value=print_config['header_line1']).props('dark outlined dense').classes('w-full')
                    with ui.column().classes('col gap-0'):
                        ui.label('Linha 2 (Título do Evento):').classes('text-[10px] text-grey-5')
                        input_header2 = ui.input(value=print_config['header_line2']).props('dark outlined dense').classes('w-full')
                    with ui.column().classes('col-3 gap-0'):
                        ui.label('Termo Convidado:').classes('text-[10px] text-grey-5')
                        input_termo_conv = ui.input(value=print_config['termo_convidado']).props('dark outlined dense').classes('w-full')

                ui.separator().classes('q-my-sm').style('border-color: rgba(255,255,255,0.08);')

                # ── Funções de Integração com Supabase Storage Bucket 'logos' ──
                from database import list_supabase_storage_files, upload_file_to_supabase_storage

                def get_dynamic_logo_options():
                    opts = {
                        'cfn': '⚓ CFN (Fuzileiros Navais)',
                        'mb':  '⚓ Marinha do Brasil'
                    }
                    try:
                        storage_files = list_supabase_storage_files('logos')
                        for f in storage_files:
                            fname = f.get('name', '')
                            furl = f.get('url', '')
                            if fname and furl:
                                opts[furl] = f"☁️ {fname}"
                    except Exception as err:
                        print(f"[LOGOS BUCKET FETCH ERR] {err}")
                    opts['custom'] = '🔗 URL Customizada'
                    return opts

                def refresh_logo_options():
                    new_opts = get_dynamic_logo_options()
                    sel_logo_preset.options = new_opts
                    sel_logo_preset.update()

                def open_logo_upload_dialog():
                    with ui.dialog() as upload_diag, ui.card().classes('q-pa-md').style('min-width: 480px; background: #0c1829; border: 1px solid #00e5ff; border-radius: 12px;'):
                        ui.label('📤 UPLOAD DE NOVO LOGO / BRASÃO').classes('text-sm font-bold text-cyan cyber-title q-mb-xs')
                        ui.label('Envie um arquivo PNG/JPG para o Supabase Storage Bucket (logos)').classes('text-xs text-grey-4 q-mb-sm')

                        async def handle_logo_upload(e):
                            try:
                                import inspect
                                import asyncio
                                import os
                                file_obj = getattr(e, 'file', None)
                                if not file_obj:
                                    ui.notify('❌ Nenhum arquivo detectado no upload.', color='negative')
                                    return
                                
                                content = file_obj.read()
                                if inspect.isawaitable(content):
                                    content = await content
                                
                                fname = getattr(file_obj, 'name', 'logo.png')
                                content_type = getattr(file_obj, 'content_type', 'image/png') or 'image/png'
                                
                                fname_lower = fname.lower()
                                os.makedirs('assets/insignias', exist_ok=True)
                                local_path = os.path.join('assets', 'insignias', fname_lower)
                                try:
                                    with open(local_path, 'wb') as f_out:
                                        f_out.write(content)
                                except Exception as f_err:
                                    print(f"[LOCAL SAVE ERR] {f_err}")

                                public_url = await asyncio.to_thread(upload_file_to_supabase_storage, content, fname, content_type, 'logos')
                                if public_url:
                                    ui.notify(f'✅ Logo "{fname}" enviado para o Supabase com sucesso!', color='positive')
                                    refresh_logo_options()
                                    sel_logo_preset.value = public_url
                                    upload_diag.close()
                                    preview_container.refresh()
                                else:
                                    ui.notify(f'✅ Salvo localmente em assets/insignias/{fname_lower}', color='warning')
                                    refresh_logo_options()
                                    upload_diag.close()
                                    preview_container.refresh()
                            except Exception as u_err:
                                print(f"[LOGO UPLOAD ERR] {u_err}")
                                ui.notify(f'❌ Erro no upload: {u_err}', color='negative')

                        ui.upload(on_upload=handle_logo_upload, auto_upload=True, max_files=1).props('accept=.png,.jpg,.jpeg,.svg dark dense').classes('w-full q-my-sm')

                        with ui.row().classes('w-full justify-end q-mt-sm'):
                            ui.button('Cancelar', on_click=upload_diag.close).props('unelevated color=grey-8 dense').classes('text-xs')

                    upload_diag.open()

                # ── Linha 3: Brasão Principal (Esquerdo) ──
                with ui.row().classes('w-full gap-3 items-end wrap bg-cyan-950/20 q-pa-sm rounded-lg border border-cyan-500/10'):
                    with ui.column().classes('gap-0'):
                        ui.label('Posição dos Brasões:').classes('text-[10px] text-grey-5')
                        sel_brasao_pos = ui.select(
                            options={
                                'esquerda': '◀ Brasão à Esquerda',
                                'ambos': '◀ Esquerda + Direita ▶',
                                'centro': '● Brasão Centralizado',
                                'nenhum': '✕ Sem Brasão'
                            },
                            value=print_config['brasao_pos']
                        ).props('dark outlined dense').style('min-width: 170px;')

                    with ui.column().classes('gap-0'):
                        ui.label('Origem Brasão Principal:').classes('text-[10px] text-grey-5')
                        sel_origin_logo_l = ui.select(
                            options={'bucket': '☁️ Bucket Supabase', 'url': '🔗 URL Externa Customizada'},
                            value=print_config['origin_logo_l']
                        ).props('dark outlined dense').style('min-width: 170px;')

                    with ui.column().classes('gap-0 col'):
                        ui.label('Brasão Principal (Bucket Supabase):').classes('text-[10px] text-grey-5')
                        with ui.row().classes('items-center gap-1 w-full'):
                            sel_logo_preset = ui.select(
                                options=get_dynamic_logo_options(),
                                value=print_config['logo_preset_l']
                            ).props('dark outlined dense').classes('col')
                            
                            ui.button(icon='refresh', on_click=refresh_logo_options).props('flat round dense color=cyan text-color=white').classes('text-xs').tooltip('Recarregar logos do Supabase')
                            ui.button('➕ Upload', icon='cloud_upload', on_click=open_logo_upload_dialog).props('unelevated color=cyan text-color=black dense bold').classes('text-xs')

                    with ui.column().classes('gap-0 col'):
                        ui.label('URL Direta Brasão Principal (se URL Externa):').classes('text-[10px] text-grey-5')
                        upload_brasao_left = ui.input(value=print_config['upload_brasao_left'], placeholder='https://.../brasao.png').props('dark outlined dense').classes('w-full')

                # ── Linha 4: Brasão Secundário (Direito) & QR Code ──
                with ui.row().classes('w-full gap-3 items-end wrap bg-cyan-950/20 q-pa-sm rounded-lg border border-cyan-500/10 q-mt-xs'):
                    with ui.column().classes('gap-0'):
                        ui.label('Posição QR Code:').classes('text-[10px] text-grey-5')
                        sel_qr_pos = ui.select(
                            options={
                                'direita': '▶ Canto Direito',
                                'esquerda': '◀ Canto Esquerdo',
                                'centro_baixo': '▼ Centro Inferior'
                            },
                            value=print_config['qr_pos']
                        ).props('dark outlined dense').style('min-width: 170px;')

                    with ui.column().classes('gap-0'):
                        ui.label('Origem Brasão Direito:').classes('text-[10px] text-grey-5')
                        sel_origin_logo_r = ui.select(
                            options={'bucket': '☁️ Bucket Supabase', 'url': '🔗 URL Externa Customizada'},
                            value=print_config['origin_logo_r']
                        ).props('dark outlined dense').style('min-width: 170px;')

                    with ui.column().classes('gap-0 col'):
                        ui.label('Brasão Direito (Bucket Supabase):').classes('text-[10px] text-grey-5')
                        sel_logo_r_preset = ui.select(
                            options=get_dynamic_logo_options(),
                            value=print_config['logo_preset_r']
                        ).props('dark outlined dense').classes('w-full')

                    with ui.column().classes('gap-0 col'):
                        ui.label('URL Direta Brasão Direito (se URL Externa):').classes('text-[10px] text-grey-5')
                        upload_brasao_right = ui.input(value=print_config['upload_brasao_right'], placeholder='https://.../brasao_direita.png').props('dark outlined dense').classes('w-full')

                # ── Linha 5: Template de Fundo (Background) ──
                with ui.row().classes('w-full gap-3 items-end wrap bg-cyan-950/20 q-pa-sm rounded-lg border border-cyan-500/10 q-mt-xs'):
                    with ui.column().classes('gap-0'):
                        ui.label('Origem Imagem de Fundo:').classes('text-[10px] text-amber-4 font-bold')
                        sel_origin_bg = ui.select(
                            options={'none': '🎨 Fundo Padrão (Sem Imagem)', 'bucket': '☁️ Bucket Supabase', 'url': '🔗 URL Externa Customizada'},
                            value=print_config['origin_bg']
                        ).props('dark outlined dense').style('min-width: 210px;')

                    with ui.column().classes('gap-0 col'):
                        ui.label('Imagem de Fundo (Bucket Supabase):').classes('text-[10px] text-amber-4 font-bold')
                        sel_bg_preset = ui.select(
                            options=get_dynamic_logo_options(),
                            value=print_config['bg_preset']
                        ).props('dark outlined dense').classes('w-full')

                    with ui.column().classes('gap-0 col'):
                        ui.label('URL Direta Imagem de Fundo (se URL Externa):').classes('text-[10px] text-amber-4 font-bold')
                        input_template_bg = ui.input(value=print_config['template_bg_url'], placeholder='https://.../fundo.png').props('dark outlined dense').classes('w-full')

                ui.separator().classes('q-my-sm').style('border-color: rgba(255,255,255,0.08);')

                # ── Linha 5: Ajustes Finais de Escala, Tamanho e Posicionamento X/Y ──
                with ui.expansion('📐 Ajustes Finais de Escala, Tamanho (mm/pt) e Posicionamento Fino (X/Y)', icon='tune').classes('w-full bg-cyan-950/40 border border-cyan-500/20 rounded-lg text-cyan text-xs q-mb-sm'):
                    with ui.row().classes('w-full gap-3 q-pa-sm wrap items-end'):
                        with ui.column().classes('gap-0'):
                            ui.label('Largura Logo (mm):').classes('text-[10px] text-grey-4')
                            input_logo_width = ui.number(value=print_config['logo_width'], min=5, max=60, step=1).props('dark outlined dense').style('width: 110px;')
                        with ui.column().classes('gap-0'):
                            ui.label('Posição Logo X (mm):').classes('text-[10px] text-grey-4')
                            input_logo_pos_x = ui.number(value=print_config['logo_pos_x'], min=0, max=60, step=1).props('dark outlined dense').style('width: 110px;')
                        with ui.column().classes('gap-0'):
                            ui.label('Posição Logo Y (mm):').classes('text-[10px] text-grey-4')
                            input_logo_pos_y = ui.number(value=print_config['logo_pos_y'], min=0, max=40, step=1).props('dark outlined dense').style('width: 110px;')

                        with ui.column().classes('gap-0'):
                            ui.label('Tamanho QR Code (mm):').classes('text-[10px] text-grey-4')
                            input_qr_size = ui.number(value=print_config['qr_size'], min=5, max=40, step=1).props('dark outlined dense').style('width: 120px;')
                        with ui.column().classes('gap-0'):
                            ui.label('Posição QR Code X (mm):').classes('text-[10px] text-grey-4')
                            input_qr_pos_x = ui.number(value=print_config['qr_pos_x'], min=0, max=60, step=1).props('dark outlined dense').style('width: 120px;')
                        with ui.column().classes('gap-0'):
                            ui.label('Posição QR Code Y (mm):').classes('text-[10px] text-grey-4')
                            input_qr_pos_y = ui.number(value=print_config['qr_pos_y'], min=0, max=40, step=1).props('dark outlined dense').style('width: 120px;')

                        with ui.column().classes('gap-0'):
                            ui.label('Fonte Nome (pt):').classes('text-[10px] text-grey-4')
                            input_font_nome = ui.number(value=print_config['font_nome'], min=10, max=40, step=1).props('dark outlined dense').style('width: 100px;')
                        with ui.column().classes('gap-0'):
                            ui.label('Fonte Posto (pt):').classes('text-[10px] text-grey-4')
                            input_font_posto = ui.number(value=print_config['font_posto'], min=8, max=30, step=1).props('dark outlined dense').style('width: 100px;')
                        with ui.column().classes('gap-0'):
                            ui.label('Fonte Reservado (pt):').classes('text-[10px] text-grey-4')
                            input_font_reservado = ui.number(value=print_config['font_reservado'], min=10, max=36, step=1).props('dark outlined dense').style('width: 110px;')

                def collect_current_print_config():
                    return {
                        'model': model_select.value,
                        'items_per_sheet': input_items_per_sheet.value,
                        'chk_only_confirmed': chk_only_confirmed.value,
                        'chk_logo': chk_logo.value,
                        'chk_qr': chk_qr.value,
                        'chk_rank': chk_rank.value,
                        'chk_border': chk_border.value,
                        'header_line1': input_header1.value or '',
                        'header_line2': input_header2.value or '',
                        'termo_convidado': input_termo_conv.value or 'RESERVADO',
                        'brasao_pos': sel_brasao_pos.value,
                        'origin_logo_l': sel_origin_logo_l.value,
                        'logo_preset_l': sel_logo_preset.value,
                        'upload_brasao_left': upload_brasao_left.value or '',
                        'qr_pos': sel_qr_pos.value,
                        'origin_logo_r': sel_origin_logo_r.value,
                        'logo_preset_r': sel_logo_r_preset.value,
                        'upload_brasao_right': upload_brasao_right.value or '',
                        'origin_bg': sel_origin_bg.value,
                        'bg_preset': sel_bg_preset.value,
                        'template_bg_url': input_template_bg.value or '',
                        'logo_width': input_logo_width.value or 16,
                        'logo_pos_x': input_logo_pos_x.value or 6,
                        'logo_pos_y': input_logo_pos_y.value or 4,
                        'qr_size': input_qr_size.value or 13,
                        'qr_pos_x': input_qr_pos_x.value or 5,
                        'qr_pos_y': input_qr_pos_y.value or 3,
                        'font_nome': input_font_nome.value or 22,
                        'font_posto': input_font_posto.value or 13,
                        'font_reservado': input_font_reservado.value or 18,
                    }

                def save_print_config_to_event(notify_user=True):
                    try:
                        import json
                        from database import get_db_connection
                        _db = get_db_connection()

                        cfg = collect_current_print_config()
                        target_layout = layout if isinstance(layout, dict) else {}
                        target_layout['print_config'] = cfg
                        
                        new_layout_str = json.dumps(target_layout, ensure_ascii=False)
                        _db.table('jade_eventos').update({'layout_json': new_layout_str}).eq('id', event['id']).execute()
                        if notify_user:
                            ui.notify('💾 Configurações salvas como padrão deste evento!', color='positive', icon='check_circle')
                    except Exception as s_err:
                        print(f"[SAVE PRINT CONFIG ERR] {s_err}")
                        if notify_user:
                            ui.notify(f'❌ Erro ao salvar configurações: {s_err}', color='negative')

                def on_update_and_preview():
                    save_print_config_to_event(notify_user=False)
                    preview_container.refresh()

                with ui.row().classes('w-full justify-between items-center q-mt-sm'):
                    ui.label('💡 As configurações salvas são gravadas como padrão deste evento.').classes('text-[11px] text-cyan-3 italic')
                    with ui.row().classes('gap-2'):
                        ui.button('💾 Salvar Configurações no Evento', icon='save', on_click=lambda: save_print_config_to_event(notify_user=True)).props('unelevated color=emerald text-color=white dense bold').classes('text-xs').tooltip('Salva o modelo, brasões, títulos e fontes como padrão permanente deste evento')
                        ui.button('🔄 Atualizar Pré-Visualização', icon='refresh', on_click=on_update_and_preview).props('unelevated color=cyan text-color=black dense bold').classes('text-xs')

            # ═══════════════════════════════════════════════════════════════
            # CSS DE IMPRESSÃO
            # ═══════════════════════════════════════════════════════════════
            print_css = """
            <style>
            @media print {
                body * { visibility: hidden !important; background: white !important; color: black !important; }
                .print-area, .print-area * { visibility: visible !important; }
                .print-area { position: absolute !important; left: 0 !important; top: 0 !important; width: 100% !important; }
                .print-hide { display: none !important; }
                .page-break { page-break-after: always !important; }
                .jade-card { border: 2px solid #000 !important; margin-bottom: 16px !important; padding: 16px !important; background: #fff !important; color: #000 !important; border-radius: 8px !important; }
                .jade-card-template { margin-bottom: 16px !important; page-break-inside: avoid !important; }
                .jade-card-mesa { border: 2px dashed #666 !important; padding: 12px !important; background: #fff !important; color: #000 !important; margin-bottom: 16px !important; }
                .jade-card-cred { border: 1px solid #000 !important; width: 300px !important; height: 420px !important; padding: 12px !important; margin: 8px !important; float: left !important; }
                /* Placa Horizontal Oficial CGCFN */
                .jade-horizontal-sheet { page-break-after: always !important; }
                .jade-horizontal-grid { display: grid !important; grid-template-columns: 1fr 1fr !important; gap: 12mm !important; padding: 8mm !important; }
                .jade-placa-horiz {
                    background: #ffffff !important; color: #1a1a6e !important;
                    border: 2px solid #1a1a6e !important; border-radius: 4px !important;
                    width: 100% !important; height: 72mm !important; overflow: hidden !important;
                    page-break-inside: avoid !important; position: relative !important;
                }
                .jade-cut-line {
                    border: 1px dashed #aaa !important;
                    margin: 3mm 0 !important;
                }
            }
            /* Preview screen */
            .jade-placa-horiz {
                background: #ffffff;
                color: #1a1a6e;
                border: 2px solid #1a1a6e;
                border-radius: 4px;
                position: relative;
                overflow: hidden;
                display: flex;
                flex-direction: row;
                align-items: center;
                min-height: 110px;
                gap: 0;
            }
            .jade-placa-horiz .jade-placa-left {
                display: flex;
                flex-direction: column;
                align-items: center;
                justify-content: center;
                padding: 12px 10px 12px 16px;
                gap: 6px;
                min-width: 90px;
            }
            .jade-placa-horiz .jade-placa-center {
                flex: 1;
                display: flex;
                flex-direction: column;
                align-items: center;
                justify-content: center;
                text-align: center;
                padding: 12px 8px;
                gap: 4px;
            }
            .jade-placa-horiz .jade-placa-right {
                display: flex;
                flex-direction: column;
                align-items: center;
                justify-content: center;
                padding: 12px 12px 12px 4px;
                gap: 4px;
            }
            .jade-star { font-size: 22px; color: #1a1a6e; line-height: 1.1; }
            .jade-reservado { font-size: 15px; font-weight: 900; color: #8B0000; letter-spacing: 2px; text-transform: uppercase; }
            .jade-posto-ext { font-size: 11px; font-weight: 700; color: #1a1a6e; letter-spacing: 1px; text-transform: uppercase; }
            .jade-nome-guerra { font-size: 22px; font-weight: 900; color: #1a1a6e; letter-spacing: 1.5px; text-transform: uppercase; }
            .jade-assento-badge { font-size: 9px; font-weight: 800; color: #555; border: 1px solid #ccc; padding: 2px 6px; border-radius: 3px; margin-top: 4px; }
            .jade-horizontal-grid {
                display: grid;
                grid-template-columns: 1fr 1fr;
                gap: 16px;
                padding: 8px 0;
            }
            /* Container de 1 dos 4 slots por folha A4 (210mm x 74.25mm) */
            .prisma-card-a4-slot {
                width: 100%;
                max-width: 210mm;
                height: 74.25mm;
                box-sizing: border-box;
                padding: 4mm 8mm;
                position: relative;
                background-color: #ffffff;
                color: #000000;
                border: 1.5pt solid #1a1a1a;
                outline: 0.5pt solid #1a1a1a;
                outline-offset: -2.5mm;
                display: flex;
                align-items: center;
                justify-content: center;
                page-break-inside: avoid;
                margin-bottom: 3mm;
            }
            .prisma-canto-esquerdo {
                position: absolute;
                top: 4mm;
                left: 6mm;
                display: flex;
                flex-direction: column;
                align-items: flex-start;
                z-index: 10;
            }
            .prisma-brasao-om {
                width: 14mm;
                height: auto;
                margin-bottom: 1.5mm;
            }
            .prisma-estrelas-esquerda {
                font-size: 18pt;
                font-weight: bold;
                color: #000000;
                letter-spacing: 2px;
                line-height: 1;
            }
            .prisma-conteudo-central {
                width: 100%;
                height: 100%;
                display: flex;
                flex-direction: column;
                justify-content: center;
                align-items: center;
                text-align: center;
                padding-left: 26mm;
                padding-right: 26mm;
                gap: 5mm;
            }
            .prisma-texto-reservado {
                font-size: 18pt;
                font-weight: 800;
                text-transform: uppercase;
                color: #000000;
                line-height: 1;
                margin: 0;
                white-space: nowrap;
            }
            .prisma-posto-extenso {
                font-size: 13pt;
                font-weight: 600;
                text-transform: uppercase;
                color: #333333;
                line-height: 1;
                margin: 0;
                white-space: nowrap;
            }
            .prisma-nome-autoridade {
                font-size: 22pt;
                font-weight: 800;
                text-transform: uppercase;
                color: #000000;
                line-height: 1;
                margin: 0;
                white-space: nowrap;
            }
            .jade-template-card {
                position: relative;
                width: 100%;
                min-height: 180px;
                background-size: cover;
                background-position: center;
                background-repeat: no-repeat;
                border-radius: 12px;
                overflow: hidden;
            }
            .jade-template-overlay {
                position: relative;
                z-index: 2;
                padding: 16px;
                min-height: 180px;
                display: flex;
                flex-direction: column;
                justify-content: space-between;
                background: rgba(0,0,0,0.45);
            }
            .jade-insignia-badge {
                display: inline-flex;
                align-items: center;
                gap: 4px;
                padding: 2px 8px;
                border-radius: 4px;
                font-size: 10px;
                font-weight: 800;
                letter-spacing: 1px;
            }
            </style>
            """
            ui.add_head_html(print_css)

            # ═══════════════════════════════════════════════════════════════
            # ÁREA DE PRÉ-VISUALIZAÇÃO
            # ═══════════════════════════════════════════════════════════════
            @ui.refreshable
            def preview_container():
                current_model = model_select.value
                h1 = input_header1.value or 'MARINHA DO BRASIL'
                h2 = input_header2.value or ''
                brasao_pos = sel_brasao_pos.value
                qr_pos = sel_qr_pos.value
                show_logo = chk_logo.value
                show_qr = chk_qr.value
                show_rank = chk_rank.value
                only_confirmed = chk_only_confirmed.value
                bg_url = ''
                try:
                    bg_url = input_template_bg.value.strip() if hasattr(input_template_bg, 'value') else ''
                except Exception:
                    bg_url = ''
                SUPABASE_LOGOS_BUCKET_URL = "https://ruabgndnhgdverqlgvef.supabase.co/storage/v1/object/public/logos"

                def resolve_logo_url(inp_val):
                    if not inp_val:
                        return ""
                    val = str(inp_val).strip()
                    if val.startswith('http://') or val.startswith('https://'):
                        return val
                    return f"{SUPABASE_LOGOS_BUCKET_URL}/{val.lstrip('/')}"

                LOGO_URLS = {
                    'cgcfn': url_to_base64('assets/brasao_cgcfn.png') or f"{SUPABASE_LOGOS_BUCKET_URL}/brasao_cgcfn.png",
                    'mb':    url_to_base64('assets/brasao_marinha.png') or f"{SUPABASE_LOGOS_BUCKET_URL}/brasao_marinha.png",
                    'cfn':   url_to_base64('assets/brasao_cgcfn.png'),
                }

                def resolve_asset_image(origin_type, bucket_val, custom_val):
                    if origin_type == 'bucket' and bucket_val:
                        if bucket_val in LOGO_URLS:
                            return LOGO_URLS[bucket_val]
                        return url_to_base64(resolve_logo_url(bucket_val))
                    elif origin_type == 'url' and custom_val:
                        return url_to_base64(custom_val.strip())
                    return ""

                origin_l = sel_origin_logo_l.value if hasattr(sel_origin_logo_l, 'value') else 'bucket'
                preset_l = sel_logo_preset.value if hasattr(sel_logo_preset, 'value') else 'cfn'
                custom_l = upload_brasao_left.value if hasattr(upload_brasao_left, 'value') else ''
                resolved_logo_url = resolve_asset_image(origin_l, preset_l, custom_l) or url_to_base64('assets/brasao_cgcfn.png')


                origin_r = sel_origin_logo_r.value if hasattr(sel_origin_logo_r, 'value') else 'bucket'
                preset_r = sel_logo_r_preset.value if hasattr(sel_logo_r_preset, 'value') else 'mb'
                custom_r = upload_brasao_right.value if hasattr(upload_brasao_right, 'value') else ''
                brasao_r_url = resolve_asset_image(origin_r, preset_r, custom_r)

                origin_bg = sel_origin_bg.value if hasattr(sel_origin_bg, 'value') else 'none'
                preset_bg = sel_bg_preset.value if hasattr(sel_bg_preset, 'value') else ''
                custom_bg = input_template_bg.value if hasattr(input_template_bg, 'value') else ''
                bg_url = resolve_asset_image(origin_bg, preset_bg, custom_bg)
                use_bg = bool(bg_url) or current_model == 'template_custom'

                # Escala e Posicionamento Fino
                logo_w = float(input_logo_width.value or 16)
                logo_x = float(input_logo_pos_x.value or 6)
                logo_y = float(input_logo_pos_y.value or 4)

                qr_size = float(input_qr_size.value or 13)
                qr_x = float(input_qr_pos_x.value or 5)
                qr_y = float(input_qr_pos_y.value or 3)

                f_nome = float(input_font_nome.value or 22)
                f_posto = float(input_font_posto.value or 13)
                f_reservado = float(input_font_reservado.value or 18)

                # Filtro de impressão: só confirmados ou todos com placa
                def should_print(c):
                    if only_confirmed:
                        return c.get('status_placa') in ('pendente', 'em_producao', 'impressa', 'reimpressao')
                    return True

                with ui.column().classes('w-full gap-4 print-area'):
                    # 1. Coleta TODOS os cartões (alocados + não alocados) em uma única lista contínua
                    all_cards = []
                    seen_ids = set()
                    main_guests_map = {g['id']: g for g in convidados if not g.get('convidado_principal_id')}
                    
                    # Se abriu o estúdio passando uma lista específica
                    if convidados and len(convidados) > 0 and isinstance(convidados, list):
                        for c in convidados:
                            if c['id'] not in seen_ids and should_print(c):
                                seen_ids.add(c['id'])
                                all_cards.append(c)
                    else:
                        for row_label, list_c_raw in allocated_by_row.items():
                            for c in list_c_raw:
                                if c['id'] not in seen_ids and should_print(c):
                                    seen_ids.add(c['id'])
                                    all_cards.append(c)

                    def sort_key_assento(c):
                        ass = str(c.get('assento_id', '')).upper().strip()
                        match = re.match(r'([A-Z]+)-?(\d+)', ass)
                        if match:
                            row, num = match.groups()
                            return (row, int(num))
                        return (ass if ass else 'ZZZ', 0)

                    all_cards.sort(key=sort_key_assento)

                    if not all_cards:
                        with ui.column().classes('w-full items-center justify-center q-py-xl gap-3'):
                            ui.icon('chair_alt', size='3rem', color='cyan-3')
                            ui.label('Nenhum convidado confirmado para impressão nesta solenidade.').classes('text-sm text-grey-4 text-center')
                        return

                    # 2. Pega a quantidade solicitada de placas por folha A4 (configurável pelo operador, padrão 4)
                    items_per_sheet = int(input_items_per_sheet.value or 4)
                    total_pages = (len(all_cards) + items_per_sheet - 1) // items_per_sheet

                    for page_idx in range(total_pages):
                        batch = all_cards[page_idx * items_per_sheet : (page_idx + 1) * items_per_sheet]
                        
                        with ui.column().classes('w-full gap-2 page-break q-mb-md'):
                            with ui.row().classes('w-full justify-between items-center bg-cyan-950/60 q-pa-sm rounded-lg border border-cyan-500/40 print-hide'):
                                first_seat = batch[0].get('assento_id') or 'Sem Assento'
                                last_seat = batch[-1].get('assento_id') or 'Sem Assento'
                                ui.label(f"📄 FOLHA A4 #{page_idx + 1} DE {total_pages} — {len(batch)} PLACAS ({first_seat} ATÉ {last_seat})").classes('text-sm font-bold text-cyan')
                                ui.badge(f"Folha {page_idx + 1} / {total_pages}").props('color=cyan text-color=black')

                            # ═══ MODELO: PRISMA INSTITUCIONAL A4 (4 por Folha) ═══
                            if current_model == 'prisma_a4_4slots':
                                termo_reservado = input_termo_conv.value or 'RESERVADO'
                                show_double_border = chk_border.value
                                with ui.column().classes('w-full gap-3'):
                                    for c in batch:
                                        is_acomp = bool(c.get('convidado_principal_id'))
                                        main_g = main_guests_map.get(c.get('convidado_principal_id')) if is_acomp else None
                                        posto = (c.get('posto_graduacao') or (main_g.get('posto_graduacao') if main_g else '') or '').strip()
                                        almirantado_info = parse_almirantado_stars(posto)
                                        nome_limpo = clean_authority_name(c['nome'])
                                        target_logo = resolved_logo_url or brasao_l_url
                                        
                                        border_style = 'border: 1.5pt solid #1a1a1a; outline: 0.5pt solid #1a1a1a; outline-offset: -2.5mm;' if show_double_border else 'border: 1.5pt solid #1a1a1a;'
                                        
                                        with ui.element('div').classes('prisma-card-a4-slot').style(border_style):
                                            # Canto Superior Esquerdo: Brasão + Estrelas / PNG Insígnias
                                            with ui.element('div').style(f'position: absolute; top: {logo_y}mm; left: {logo_x}mm; z-index: 10; display: flex; flex-direction: column; align-items: flex-start; shadow: none;'):
                                                if show_logo and target_logo:
                                                    ui.image(target_logo).style(f'width: {logo_w}mm; height: auto; object-fit: contain;')
                                                elif show_logo:
                                                    ui.label('⚓').style(f'font-size: {logo_w}px; color: #000;')
                                                
                                                # Insígnia/Estrelas impressas tanto para o titular quanto para os acompanhantes
                                                if show_rank:
                                                    if almirantado_info['png_asset']:
                                                        ui.image(almirantado_info['png_asset']).style(f'width: {logo_w}mm; height: auto; margin-top: 1mm;')
                                                    elif almirantado_info['stars']:
                                                        ui.label(almirantado_info['stars']).classes('prisma-estrelas-esquerda')

                                            # Canto Superior Direito: QR Code (Base64 offline)
                                            if show_qr:
                                                qr_url = gen_qr_base64(f"JADE|{event.get('id','')}|{c['id']}|{c.get('assento_id','')}")
                                                with ui.element('div').style(f'position: absolute; top: {qr_y}mm; right: {qr_x}mm; z-index: 10; font-size: 7px; text-align: center;'):
                                                    ui.image(qr_url).style(f'width: {qr_size}mm; height: {qr_size}mm;')
                                                    if c.get('assento_id'):
                                                        ui.label(f"{c.get('assento_id', '')}").classes('text-[7px] font-mono text-black font-bold')

                                            # Bloco Central - Totalmente Centralizado com Fontes Dinâmicas
                                            with ui.element('div').classes('prisma-conteudo-central'):
                                                if is_acomp:
                                                    ui.label(termo_reservado).classes('prisma-texto-reservado').style(f'font-size: {f_reservado}pt;')
                                                    if almirantado_info['title']:
                                                        ui.label(almirantado_info['title']).classes('prisma-posto-extenso').style(f'font-size: {f_posto}pt;')
                                                    elif posto:
                                                        ui.label(posto.upper()).classes('prisma-posto-extenso').style(f'font-size: {f_posto}pt;')
                                                    ui.label(nome_limpo).classes('prisma-nome-autoridade').style(f'font-size: {f_nome}pt;')
                                                else:
                                                    if almirantado_info['title']:
                                                        ui.label(almirantado_info['title']).classes('prisma-posto-extenso').style(f'font-size: {f_posto}pt;')
                                                    elif posto:
                                                        ui.label(posto.upper()).classes('prisma-posto-extenso').style(f'font-size: {f_posto}pt;')
                                                    ui.label(nome_limpo).classes('prisma-nome-autoridade').style(f'font-size: {f_nome}pt;')
                            # ═══ MODELO: PRISMA DOBRÁVEL V (FRENTE E VERSO INVERTIDO 180°) ═══
                            elif current_model == 'prisma_dobravel_v':
                                termo_reservado = input_termo_conv.value or 'RESERVADO'
                                with ui.column().classes('w-full gap-4'):
                                    for c in batch:
                                        is_acomp = bool(c.get('convidado_principal_id'))
                                        posto = c.get('posto_graduacao') or ''
                                        almirantado_info = parse_almirantado_stars(posto)
                                        nome_limpo = clean_authority_name(c['nome'])
                                        target_logo = resolved_logo_url or brasao_l_url
                                        
                                        # Geração local offline de QR Code
                                        import qrcode, io, base64
                                        qr_b64 = f"https://api.qrserver.com/v1/create-qr-code/?size=100x100&data={c['id']}"
                                        try:
                                            qr_obj = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=4, border=1)
                                            qr_obj.add_data(f"JADE|{event.get('id','')}|{c['id']}|{c.get('assento_id','')}")
                                            qr_obj.make(fit=True)
                                            img_q = qr_obj.make_image(fill_color="black", back_color="white")
                                            buf_q = io.BytesIO()
                                            img_q.save(buf_q, format="PNG")
                                            qr_b64 = f"data:image/png;base64,{base64.b64encode(buf_q.getvalue()).decode('utf-8')}"
                                        except Exception:
                                            pass

                                        with ui.element('div').classes('w-full border border-black rounded-lg bg-white text-black q-pa-xs').style('min-height: 170mm; display: flex; flex-direction: column; justify-content: space-between; position: relative;'):
                                            # Verso da Dobra (Texto Invertido 180° para quem olha de frente)
                                            with ui.element('div').style('transform: rotate(180deg); display: flex; flex-direction: column; align-items: center; justify-content: center; padding: 10mm 4mm;'):
                                                if show_logo and target_logo:
                                                    ui.image(target_logo).style(f'width: {logo_w}mm; height: auto; object-fit: contain; margin-bottom: 2mm;')
                                                if is_acomp:
                                                    ui.label(termo_reservado).classes('prisma-texto-reservado').style(f'font-size: {f_reservado}pt;')
                                                if almirantado_info['title']:
                                                    ui.label(almirantado_info['title']).classes('prisma-posto-extenso').style(f'font-size: {f_posto}pt;')
                                                elif posto:
                                                    ui.label(posto.upper()).classes('prisma-posto-extenso').style(f'font-size: {f_posto}pt;')
                                                ui.label(nome_limpo).classes('prisma-nome-autoridade').style(f'font-size: {f_nome}pt;')

                                            # Linha de Dobra do Papel A4
                                            ui.html('<div style="border-top: 2px dashed #000; text-align: center; margin: 4px 0;"><span style="background:#fff; padding: 0 8px; font-size: 8px; color: #666; font-family: monospace;">✂️ LINHA DE DOBRA DO PRISMA DE MESA (V-SHAPE)</span></div>')

                                            # Frente da Dobra (Texto Normal 0° para quem olha do corredor)
                                            with ui.element('div').style('display: flex; flex-direction: column; align-items: center; justify-content: center; padding: 10mm 4mm; position: relative;'):
                                                if show_logo and target_logo:
                                                    ui.image(target_logo).style(f'width: {logo_w}mm; height: auto; object-fit: contain; margin-bottom: 2mm;')
                                                if is_acomp:
                                                    ui.label(termo_reservado).classes('prisma-texto-reservado').style(f'font-size: {f_reservado}pt;')
                                                if almirantado_info['title']:
                                                    ui.label(almirantado_info['title']).classes('prisma-posto-extenso').style(f'font-size: {f_posto}pt;')
                                                elif posto:
                                                    ui.label(posto.upper()).classes('prisma-posto-extenso').style(f'font-size: {f_posto}pt;')
                                                ui.label(nome_limpo).classes('prisma-nome-autoridade').style(f'font-size: {f_nome}pt;')

                                                if show_qr:
                                                    with ui.element('div').style('position: absolute; right: 4mm; bottom: 4mm; text-align: center;'):
                                                        ui.image(qr_b64).style(f'width: {qr_size}mm; height: {qr_size}mm;')
                                                        if c.get('assento_id'):
                                                            ui.label(f"{c.get('assento_id', '')}").classes('text-[7px] font-mono text-black font-bold')

                            # ═══ MODELO: PLACA HORIZONTAL CGCFN (PADRÃO OFICIAL) ═══
                            elif current_model == 'jade_horizontal_a4':
                                with ui.element('div').classes('jade-horizontal-grid w-full'):
                                    for c in batch:
                                        is_acomp = bool(c.get('convidado_principal_id'))
                                        posto_abrev = (c.get('posto_graduacao') or '').strip()
                                        insignia_data = RANK_INSIGNIAS.get(posto_abrev, None)
                                        posto_ext = insignia_data['title'] if insignia_data else posto_abrev
                                        stars = insignia_data['stars'] if insignia_data else ''
                                        nome_guerra = c['nome'].strip().upper()
                                        assento = c.get('assento_id') or ''

                                        import urllib.parse
                                        qr_data = urllib.parse.quote(f"JADE|{event.get('id','')}|{c['id']}|{posto_abrev}|{c['nome']}|{assento}")
                                        qr_url = f"https://api.qrserver.com/v1/create-qr-code/?size=80x80&data={qr_data}&color=1a1a6e&bgcolor=ffffff"

                                        logo_html = ''
                                        if show_logo and resolved_logo_url:
                                            logo_html = f'<img src="{resolved_logo_url}" style="width:{logo_w}mm;height:auto;object-fit:contain;" />'
                                        else:
                                            logo_html = f'<span style="font-size:{logo_w}px;">⚓</span>'

                                        stars_html = ''
                                        if stars and show_rank:
                                            star_list = [s for s in stars]
                                            stars_html = ''.join(f'<span class="jade-star">{s}</span>' for s in star_list)

                                        reservado_html = '<div class="jade-reservado">RESERVADO</div>' if is_acomp else ''
                                        assento_html = f'<div class="jade-assento-badge">ASSENTO {assento}</div>' if assento else ''
                                        qr_html = f'<img src="{qr_url}" style="width:{qr_size}mm;height:{qr_size}mm;background:#fff;border-radius:3px;border:1px solid #ccc;" /><div style="font-size:7px;color:#888;font-family:monospace;">{assento}</div>' if show_qr else ''

                                        ui.html(f'''
                                        <div class="jade-placa-horiz">
                                            <div class="jade-placa-left">
                                                {logo_html}
                                                <div style="display:flex;flex-direction:column;align-items:center;gap:1px;margin-top:4px;">
                                                    {stars_html}
                                                </div>
                                            </div>
                                            <div class="jade-placa-center">
                                                {reservado_html}
                                                <div class="jade-posto-ext" style="font-size:{f_posto}pt;">{posto_ext}</div>
                                                <div class="jade-nome-guerra" style="font-size:{f_nome}pt;">{nome_guerra}</div>
                                                {assento_html}
                                            </div>
                                            <div class="jade-placa-right">
                                                {qr_html}
                                            </div>
                                        </div>
                                        ''')

                            # ═══ MODELO: TEMPLATE CUSTOMIZADO ═══
                            elif current_model == 'template_custom' or use_bg:
                                with ui.grid(columns='1 sm:grid-cols-2').classes('w-full gap-4'):
                                    for c in sorted(list_c, key=lambda x: x.get('assento_id', '')):
                                        is_acomp = bool(c.get('convidado_principal_id'))
                                        posto = c.get('posto_graduacao') or ''
                                        insignia_data = RANK_INSIGNIAS.get(posto, None)
                                        qr_url = f"https://api.qrserver.com/v1/create-qr-code/?size=120x120&data={c['id']}&color=000000&bgcolor=ffffff"

                                        bg_style = f"background-image: url('{bg_url}');" if bg_url else "background: linear-gradient(135deg, #0a1628 0%, #1a2744 50%, #0d1b2a 100%);"

                                        with ui.card().classes('w-full jade-card-template jade-template-card').style(f'{bg_style} border: 2px solid rgba(0,229,255,0.3); border-radius: 12px;'):
                                            with ui.column().classes('jade-template-overlay'):
                                                # TOPO: Brasões + Cabeçalho
                                                with ui.row().classes('w-full justify-between items-start'):
                                                    if show_logo and brasao_pos in ('esquerda', 'ambos'):
                                                        if brasao_l_url:
                                                            ui.image(brasao_l_url).classes('w-10 h-10 rounded')
                                                        else:
                                                            ui.label('⚓').classes('text-2xl')
                                                    with ui.column().classes('col items-center gap-0'):
                                                        ui.label(h1).classes('text-[10px] font-black text-cyan tracking-[3px] text-center')
                                                        if h2:
                                                            ui.label(h2).classes('text-[9px] font-bold text-amber text-center')
                                                        if show_logo and brasao_pos == 'centro':
                                                            if brasao_l_url:
                                                                ui.image(brasao_l_url).classes('w-8 h-8 rounded q-mt-xs')
                                                            else:
                                                                ui.label('⚓').classes('text-xl')
                                                    if show_logo and brasao_pos in ('ambos',):
                                                        if brasao_r_url:
                                                            ui.image(brasao_r_url).classes('w-10 h-10 rounded')
                                                        else:
                                                            ui.label('⚓').classes('text-2xl')

                                                # MEIO: Insígnia + Nome + Assento
                                                with ui.column().classes('w-full items-center gap-1 q-my-sm'):
                                                    if show_rank and insignia_data:
                                                        ui.html(
                                                            f'<span class="jade-insignia-badge" style="background:{insignia_data["color"]}22; color:{insignia_data["color"]}; border: 1px solid {insignia_data["color"]}44;">'
                                                            f'{insignia_data["stars"]} {insignia_data["title"]}</span>'
                                                        )
                                                    nome_c = f"{posto} {c['nome']}".strip()
                                                    ui.label(nome_c).classes('text-lg font-black text-white text-center leading-tight')
                                                    sub = '(Acompanhante Oficial)' if is_acomp else (c.get('cargo_funcao') or c.get('categoria') or '')
                                                    if sub:
                                                        ui.label(sub).classes('text-[10px] text-grey-3 font-bold')
                                                    ui.badge(f"ASSENTO {c['assento_id']}").props('color=cyan text-color=black bold').classes('text-xs q-mt-xs')

                                                # BASE: QR Code
                                                if show_qr:
                                                    qr_align = 'justify-end' if qr_pos == 'direita' else ('justify-start' if qr_pos == 'esquerda' else 'justify-center')
                                                    with ui.row().classes(f'w-full {qr_align} items-end'):
                                                        with ui.column().classes('items-center gap-0'):
                                                            ui.image(qr_url).classes('w-14 h-14 rounded bg-white p-1')
                                                            ui.label(f"ID:{c['id']}").classes('text-[7px] font-mono text-grey-5')

                            # ═══ MODELO: PLACA DE CADEIRA A4 ═══
                            elif current_model == 'cadeira_a4':
                                with ui.grid(columns='1 sm:grid-cols-2').classes('w-full gap-4'):
                                    for c in sorted(list_c, key=lambda x: x.get('assento_id', '')):
                                        is_acomp = bool(c.get('convidado_principal_id'))
                                        posto = c.get('posto_graduacao') or ''
                                        insignia_data = RANK_INSIGNIAS.get(posto, None)
                                        qr_url = f"https://api.qrserver.com/v1/create-qr-code/?size=120x120&data={c['id']}&color=000000&bgcolor=ffffff"

                                        with ui.card().classes('w-full q-pa-md jade-card bg-slate-900 border border-cyan-500/40 rounded-xl').style('border-left: 6px solid #00e5ff !important;'):
                                            with ui.row().classes('w-full justify-between items-start no-wrap'):
                                                with ui.column().classes('gap-0 col'):
                                                    # Brasão no topo
                                                    if show_logo:
                                                        with ui.row().classes('items-center gap-1'):
                                                            if brasao_l_url:
                                                                ui.image(brasao_l_url).classes('w-5 h-5')
                                                            ui.label(h1).classes('text-[10px] font-black text-cyan tracking-widest')

                                                    ui.badge(f"ASSENTO {c['assento_id']}").props('color=cyan text-color=black bold').classes('text-xs w-fit q-my-xs')

                                                    if show_rank and insignia_data:
                                                        ui.html(
                                                            f'<span class="jade-insignia-badge" style="background:{insignia_data["color"]}22; color:{insignia_data["color"]}; border: 1px solid {insignia_data["color"]}44;">'
                                                            f'{insignia_data["stars"]} {insignia_data["title"]}</span>'
                                                        )

                                                    nome_c = f"{posto} {c['nome']}".strip()
                                                    ui.label(nome_c).classes('text-md font-black text-white leading-tight q-mt-xs')
                                                    sub = '(Acompanhante Oficial)' if is_acomp else (c.get('cargo_funcao') or c.get('categoria') or 'Convidado de Honra')
                                                    ui.label(sub).classes('text-xs text-grey-4 font-bold q-mt-xs')
                                                    if h2:
                                                        ui.label(h2).classes('text-[9px] text-amber font-bold q-mt-xs')

                                                if show_qr:
                                                    with ui.column().classes('items-center gap-0'):
                                                        ui.image(qr_url).classes('w-16 h-16 rounded bg-white p-1')
                                                        ui.label(f"ID:{c['id']}").classes('text-[8px] font-mono text-grey-5')

                            # ═══ MODELO: MESA DOBRÁVEL A5 ═══
                            elif current_model == 'mesa_a5_dobravel':
                                with ui.grid(columns='1 sm:grid-cols-2').classes('w-full gap-4'):
                                    for c in sorted(list_c, key=lambda x: x.get('assento_id', '')):
                                        is_acomp = bool(c.get('convidado_principal_id'))
                                        posto = c.get('posto_graduacao') or ''
                                        nome_c = f"{posto} {c['nome']}".strip()
                                        sub = '(Acompanhante)' if is_acomp else (c.get('cargo_funcao') or 'Convidado')
                                        qr_url = f"https://api.qrserver.com/v1/create-qr-code/?size=100x100&data={c['id']}"

                                        with ui.card().classes('w-full q-pa-md jade-card-mesa bg-slate-900 border-2 border-dashed border-cyan-500/40 rounded-xl text-center'):
                                            ui.label('✂️ DOBRA DE MESA (FACE FRONTAL)').classes('text-[8px] text-grey-5 tracking-widest uppercase print-hide')
                                            if show_logo:
                                                ui.label(h1).classes('text-[9px] font-black text-cyan tracking-widest')
                                            ui.label(f"ASSENTO {c['assento_id']}").classes('text-sm font-black text-cyan')
                                            ui.label(nome_c).classes('text-lg font-black text-white q-my-xs')
                                            ui.label(sub).classes('text-xs text-grey-3 font-bold')

                                            ui.separator().classes('q-my-sm print-hide').style('border-color: rgba(255,255,255,0.1);')
                                            ui.label('✂️ DOBRA DE MESA (FACE TRASEIRA)').classes('text-[8px] text-grey-5 tracking-widest uppercase print-hide')
                                            with ui.row().classes('w-full justify-center items-center gap-2 q-mt-xs'):
                                                if show_qr:
                                                    ui.image(qr_url).classes('w-12 h-12 bg-white p-1 rounded')
                                                ui.label(f"FILEIRA {row_label} | {c['assento_id']}").classes('text-xs text-grey-4 font-mono font-bold')

                            # ═══ MODELO: CREDENCIAL / CRACHÁ ═══
                            elif current_model == 'credencial':
                                with ui.grid(columns='1 sm:grid-cols-2 md:grid-cols-3').classes('w-full gap-3'):
                                    for c in sorted(list_c, key=lambda x: x.get('assento_id', '')):
                                        is_acomp = bool(c.get('convidado_principal_id'))
                                        main_g = main_guests_map.get(c.get('convidado_principal_id')) if is_acomp else None
                                        posto = (c.get('posto_graduacao') or (main_g.get('posto_graduacao') if main_g else '') or '').strip()
                                        almirantado_info = parse_almirantado_stars(posto)
                                        nome_limpo = clean_authority_name(main_g.get('nome') if (is_acomp and main_g) else c['nome'])
                                        target_logo = resolved_logo_url or brasao_l_url
                                        insignia_data = RANK_INSIGNIAS.get(posto, None)
                                        qr_url = f"https://api.qrserver.com/v1/create-qr-code/?size=100x100&data={c['id']}"

                                        with ui.card().classes('w-full jade-card-cred bg-slate-900 border border-amber-500/40 rounded-2xl q-pa-md items-center text-center justify-between').style('height: 300px;'):
                                            with ui.column().classes('items-center gap-0 w-full'):
                                                if show_logo:
                                                    with ui.row().classes('items-center gap-1'):
                                                        if brasao_l_url:
                                                            ui.image(brasao_l_url).classes('w-6 h-6')
                                                        ui.label(h1).classes('text-[9px] font-black text-cyan tracking-widest')
                                                        if brasao_r_url and brasao_pos == 'ambos':
                                                            ui.image(brasao_r_url).classes('w-6 h-6')
                                                if h2:
                                                    ui.label(h2).classes('text-[10px] font-bold text-amber truncate w-full')
                                                ui.separator().classes('q-my-xs w-full').style('border-color: rgba(255,255,255,0.1);')

                                            with ui.column().classes('items-center gap-0 w-full'):
                                                ui.badge(f"FILEIRA {row_label} - {c['assento_id']}").props('color=cyan text-color=black bold').classes('text-xs q-mb-xs')
                                                if show_rank and insignia_data:
                                                    ui.html(
                                                        f'<span class="jade-insignia-badge" style="background:{insignia_data["color"]}22; color:{insignia_data["color"]}; border: 1px solid {insignia_data["color"]}44; font-size:8px;">'
                                                        f'{insignia_data["stars"]} {insignia_data["title"]}</span>'
                                                    )
                                                ui.label(nome_limpo).classes('text-sm font-black text-white leading-tight')
                                                sub = '(Acompanhante)' if is_acomp else (c.get('cargo_funcao') or c.get('categoria'))
                                                if sub:
                                                    ui.label(sub).classes('text-[10px] text-grey-4 font-bold')

                                            if show_qr:
                                                ui.image(qr_url).classes('w-14 h-14 bg-white p-1 rounded border border-cyan-500')

            model_select.on('update:model-value', lambda: preview_container.refresh())
            preview_container()

            js_clean_print_cards = """
            (function() {
                var area = document.querySelector('.print-area');
                if (!area) { window.print(); return; }

                var oldIframe = document.getElementById('jade_print_iframe');
                if (oldIframe) { oldIframe.remove(); }

                var iframe = document.createElement('iframe');
                iframe.id = 'jade_print_iframe';
                iframe.style.position = 'fixed';
                iframe.style.right = '0';
                iframe.style.bottom = '0';
                iframe.style.width = '0';
                iframe.style.height = '0';
                iframe.style.border = '0';
                document.body.appendChild(iframe);

                var doc = iframe.contentWindow.document;
                var cssStyles = Array.from(document.querySelectorAll('style, link[rel="stylesheet"]'))
                                     .map(s => s.outerHTML).join('\\n');

                doc.open();
                doc.write(`
                    <!DOCTYPE html>
                    <html>
                    <head>
                        <title>JADE - Impressão Oficial de Placas de Assento</title>
                        ${cssStyles}
                        <style>
                            @page { size: A4 portrait; margin: 0mm !important; }
                            * {
                                -webkit-print-color-adjust: exact !important;
                                print-color-adjust: exact !important;
                                color-adjust: exact !important;
                            }
                            body { margin: 0 !important; padding: 4mm 6mm !important; background: #ffffff !important; color: #000000 !important; font-family: Arial, sans-serif !important; -webkit-print-color-adjust: exact !important; print-color-adjust: exact !important; }
                            .print-hide, .q-header, .q-drawer, .q-footer { display: none !important; }
                            .print-area { display: block !important; position: static !important; width: 100% !important; visibility: visible !important; }
                            .prisma-card-a4-slot { height: 66mm !important; max-height: 66mm !important; border: 1.5pt solid #1a1a1a !important; outline: 0.5pt solid #1a1a1a !important; outline-offset: -2.5mm !important; margin-bottom: 4.5mm !important; page-break-inside: avoid !important; background: #ffffff !important; color: #000000 !important; display: flex !important; flex-direction: column !important; justify-content: center !important; align-items: center !important; position: relative !important; box-sizing: border-box !important; }
                            .prisma-conteudo-central { display: flex !important; flex-direction: column !items: center !important; justify-content: center !important; text-align: center !important; width: 100% !important; }
                            .prisma-texto-reservado { font-weight: 900 !important; letter-spacing: 3px !important; text-transform: uppercase !important; color: #1f4e79 !important; font-size: 20pt !important; margin-bottom: 2px !important; }
                            .prisma-posto-extenso { font-weight: bold !important; text-transform: uppercase !important; letter-spacing: 1.5px !important; font-size: 18pt !important; margin-bottom: 2px !important; }
                            .prisma-nome-autoridade { font-weight: 900 !important; text-transform: uppercase !important; font-size: 32pt !important; line-height: 1.05 !important; }
                            img { max-width: 100% !important; display: inline-block !important; visibility: visible !important; opacity: 1 !important; -webkit-print-color-adjust: exact !important; print-color-adjust: exact !important; }
                        </style>
                    </head>
                    <body>
                        <div class="print-area">
                            ${area.innerHTML}
                        </div>
                    </body>
                    </html>
                `);
                doc.close();

                function triggerPrintWhenReady() {
                    var images = doc.images;
                    var loaded = 0;
                    var total = images.length;

                    function doPrint() {
                        iframe.contentWindow.focus();
                        iframe.contentWindow.print();
                    }

                    if (total === 0) {
                        setTimeout(doPrint, 800);
                        return;
                    }

                    for (var i = 0; i < total; i++) {
                        if (images[i].complete) {
                            loaded++;
                        } else {
                            images[i].onload = images[i].onerror = function() {
                                loaded++;
                                if (loaded >= total) {
                                    setTimeout(doPrint, 500);
                                }
                            };
                        }
                    }

                    if (loaded >= total) {
                        setTimeout(doPrint, 800);
                    } else {
                        setTimeout(doPrint, 2500); // Fallback de segurança máximo
                    }
                }

                setTimeout(triggerPrintWhenReady, 300);
            })();
            """


            def on_trigger_print():
                save_print_config_to_event(notify_user=False)
                ui.run_javascript(js_clean_print_cards)

            def on_trigger_pdf():
                save_print_config_to_event(notify_user=False)
                try:
                    cfg_now = collect_current_print_config()
                    pdf_bytes = gerar_pdf_placas_reportlab(event, convidados, current_model, only_confirmed, cfg_now)
                    if pdf_bytes:
                        import base64
                        b64_pdf = base64.b64encode(pdf_bytes).decode('utf-8')
                        file_name = f"Placas_JADE_{event.get('nome_evento','Evento').replace(' ', '_')}.pdf"
                        js_download = f"""
                        var a = document.createElement('a');
                        a.href = 'data:application/pdf;base64,{b64_pdf}';
                        a.download = '{file_name}';
                        document.body.appendChild(a);
                        a.click();
                        document.body.removeChild(a);
                        """
                        ui.run_javascript(js_download)
                        ui.notify('📄 PDF Vetorial gerado com sucesso no servidor e baixado!', color='positive')
                    else:
                        ui.notify('Nenhum convidado selecionado para PDF.', color='warning')
                except Exception as pdf_err:
                    print(f"[PDF REPORTLAB ERR] {pdf_err}")
                    ui.notify(f"Erro ao gerar PDF no servidor: {pdf_err}. Usando fallback do navegador...", color='warning')
                    ui.run_javascript(js_pdf_export_cards)

            with ui.row().classes('w-full justify-between items-center q-mt-md print-hide'):
                ui.button('💾 Salvar Configurações no Evento', icon='save', on_click=lambda: save_print_config_to_event(notify_user=True)).props('unelevated color=emerald text-color=white bold dense').classes('text-xs').tooltip('Salva o modelo, brasões, títulos e fontes como padrão permanente deste evento')
                with ui.row().classes('gap-2'):
                    ui.button('Fechar', on_click=diag.close).props('unelevated color=grey-8 dense')
                    ui.button('📄 Baixar PDF Oficial (Servidor)', icon='picture_as_pdf', on_click=on_trigger_pdf).props('unelevated color=deep-purple text-color=white bold dense').tooltip('Gera arquivo PDF milimétrico vetorial A4 direto do servidor sem falhas de imagem')
                    ui.button('🖨️ Imprimir Placas Selecionadas', icon='print', on_click=on_trigger_print).props('unelevated color=cyan text-color=black bold dense')
        diag.open()


    # ═══════════════════════════════════════════════════════════════
    # FASE: PLANILHÃO DO EVENTO (Tabela Completa + CSV Export)
    # ═══════════════════════════════════════════════════════════════
    def open_event_spreadsheet_dialog(event, convidados):
        """Abre a planilha completa do evento com filtros, status RSVP e exportação CSV."""

        STATUS_CONF_LABELS = {
            'confirmado':  ('✅ Confirmado',   'positive'),
            'recusado':    ('❌ Recusado',      'negative'),
            'justificado': ('📝 Justificado',   'warning'),
            'provavel':    ('🕐 Provável',       'cyan'),
            'pendente':    ('⏳ Pendente',       'grey-6'),
        }
        STATUS_PLACA_LABELS = {
            'pendente':       ('🟡 Pendente',     'amber'),
            'em_producao':    ('🔵 Em Produção',  'blue'),
            'impressa':       ('🟢 Impressa',     'green'),
            'reimpressao':    ('🔴 Reimpressão',  'red'),
            'entregue':       ('⚪ Entregue',     'grey-4'),
            'nao_necessaria': ('➖ N/A',           'grey-7'),
        }

        filter_conf = {'value': 'todos'}
        filter_placa = {'value': 'todos'}
        filter_search = {'value': ''}

        with ui.dialog() as diag, ui.card().classes('q-pa-md').style(
            f'min-width: 980px; max-width: 98vw; max-height: 92vh; overflow-y: auto; background: {THEME["bg_panel"]};'
        ):
            # Header
            with ui.row().classes('w-full items-center justify-between q-mb-sm'):
                with ui.column().classes('gap-0'):
                    ui.label(f'📊 PLANILHÃO DO EVENTO').classes('text-lg font-bold text-teal cyber-title')
                    ui.label(f'{event.get("nome","")[:60]}  ·  {event.get("data_evento","")}'
                             ).classes('text-xs text-grey-4')
                ui.button(icon='close', on_click=diag.close).props('flat round dense text-color=grey')

            # Métricas
            titulares = [c for c in convidados if not c.get('convidado_principal_id')]
            acompanhantes = [c for c in convidados if c.get('convidado_principal_id')]
            n_conf = sum(1 for c in titulares if c.get('status_confirmacao') == 'confirmado')
            n_rec  = sum(1 for c in titulares if c.get('status_confirmacao') == 'recusado')
            n_pend = sum(1 for c in titulares if c.get('status_confirmacao') not in ('confirmado', 'recusado', 'justificado'))

            with ui.row().classes('w-full gap-2 q-mb-sm wrap'):
                for label, val, color in [
                    ('Total Titulares', len(titulares), 'cyan'),
                    ('Acompanhantes', len(acompanhantes), 'indigo'),
                    ('✅ Confirmados', n_conf, 'green'),
                    ('❌ Recusados', n_rec, 'red'),
                    ('⏳ Pendentes', n_pend, 'amber'),
                ]:
                    with ui.card().classes(f'q-pa-sm border-l-4 border-{color}').style(
                        f'background: rgba(255,255,255,0.04); border-left: 4px solid; border-left-color: var(--q-{color});'
                    ):
                        ui.label(str(val)).classes(f'text-2xl font-black text-{color}')
                        ui.label(label).classes('text-[10px] text-grey-4 font-bold uppercase tracking-wider')

            ui.separator().classes('q-my-xs')

            # Filtros
            with ui.row().classes('w-full items-center gap-3 q-mb-sm wrap'):
                search_inp = ui.input(placeholder='🔍 Buscar por nome, posto, assento...').props('dark outlined dense').classes('col-5')
                filt_conf_sel = ui.select(
                    options={'todos': 'Todos os Status RSVP', 'confirmado': '✅ Confirmados', 'recusado': '❌ Recusados', 'provavel': '🕐 Prováveis', 'pendente': '⏳ Pendentes'},
                    value='todos'
                ).props('dark outlined dense').style('min-width: 180px;').tooltip('Filtrar por Status de Confirmação')
                filt_placa_sel = ui.select(
                    options={'todos': 'Todos os Status de Placa', 'pendente': '🟡 Pendente', 'em_producao': '🔵 Em Produção', 'impressa': '🟢 Impressa', 'entregue': '⚪ Entregue'},
                    value='todos'
                ).props('dark outlined dense').style('min-width: 200px;').tooltip('Filtrar por Status de Placa')

            @ui.refreshable
            def render_table():
                busca = search_inp.value.lower().strip()
                sc = filt_conf_sel.value
                sp = filt_placa_sel.value

                lista = sorted(convidados, key=lambda c: (c.get('assento_id') or 'ZZZ', c.get('nome', '')))

                def matches(c):
                    if busca:
                        text = f"{c.get('nome','')} {c.get('posto_graduacao','')} {c.get('assento_id','')} {c.get('cargo_funcao','')}".lower()
                        if busca not in text:
                            return False
                    if sc != 'todos':
                        if sc == 'pendente':
                            conf = c.get('status_confirmacao', 'pendente')
                            if conf in ('confirmado', 'recusado', 'justificado', 'provavel'):
                                return False
                        elif c.get('status_confirmacao') != sc:
                            return False
                    if sp != 'todos' and c.get('status_placa', 'pendente') != sp:
                        return False
                    return True

                visible = [c for c in lista if matches(c)]

                if not visible:
                    with ui.column().classes('w-full items-center q-py-lg gap-2 text-grey-4'):
                        ui.icon('search_off', size='3rem')
                        ui.label('Nenhum convidado encontrado com esses filtros.').classes('text-sm')
                    return

                # Tabela header
                with ui.element('div').classes('w-full overflow-x-auto'):
                    with ui.element('table').classes('w-full').style(
                        'border-collapse: collapse; font-size: 12px; font-family: monospace;'
                    ):
                        # Cabeçalho
                        with ui.element('thead'):
                            with ui.element('tr').style('background: rgba(0,200,200,0.12); color: #80deea;'):
                                for col in ['#', 'Assento', 'Posto/Grad.', 'Nome Completo', 'Cargo/Função', 'Tipo', 'RSVP', 'Placa', 'Acomp.']:
                                    ui.element('th').style('padding: 6px 8px; text-align: left; border-bottom: 1px solid rgba(255,255,255,0.1); white-space: nowrap;').text(col)

                        # Linhas
                        with ui.element('tbody'):
                            for idx, c in enumerate(visible, 1):
                                is_acomp = bool(c.get('convidado_principal_id'))
                                sc_label, sc_color = STATUS_CONF_LABELS.get(
                                    c.get('status_confirmacao', 'pendente'),
                                    ('⏳ Pendente', 'grey-6')
                                )
                                sp_label, sp_color = STATUS_PLACA_LABELS.get(
                                    c.get('status_placa', 'pendente'),
                                    ('🟡 Pendente', 'amber')
                                )
                                row_bg = 'rgba(255,255,255,0.02)' if idx % 2 == 0 else 'transparent'
                                nome_clean = clean_authority_name(c.get('nome', ''))

                                with ui.element('tr').style(f'background: {row_bg}; vertical-align: middle;'):
                                    ui.element('td').style('padding: 5px 8px; color: #888; border-bottom: 1px solid rgba(255,255,255,0.04);').text(str(idx))
                                    ui.element('td').style('padding: 5px 8px; font-weight: bold; color: #4dd0e1; border-bottom: 1px solid rgba(255,255,255,0.04);').text(c.get('assento_id') or '—')
                                    ui.element('td').style('padding: 5px 8px; color: #ffd54f; border-bottom: 1px solid rgba(255,255,255,0.04); white-space: nowrap;').text(c.get('posto_graduacao') or '—')
                                    with ui.element('td').style('padding: 5px 8px; border-bottom: 1px solid rgba(255,255,255,0.04);'):
                                        ui.label(nome_clean).classes('text-white font-bold text-xs')
                                    ui.element('td').style('padding: 5px 8px; color: #aaa; border-bottom: 1px solid rgba(255,255,255,0.04);').text((c.get('cargo_funcao') or '')[:35] + '…' if len(c.get('cargo_funcao') or '') > 35 else (c.get('cargo_funcao') or '—'))
                                    with ui.element('td').style('padding: 5px 8px; border-bottom: 1px solid rgba(255,255,255,0.04);'):
                                        ui.badge('ACOMP.' if is_acomp else 'TITULAR').props(f'color={"indigo" if is_acomp else "cyan"} text-color=white').classes('text-[9px]')
                                    with ui.element('td').style('padding: 5px 8px; border-bottom: 1px solid rgba(255,255,255,0.04);'):
                                        ui.badge(sc_label).props(f'color={sc_color} text-color=white').classes('text-[9px]')
                                    with ui.element('td').style('padding: 5px 8px; border-bottom: 1px solid rgba(255,255,255,0.04);'):
                                        ui.badge(sp_label).props(f'color={sp_color} text-color=white').classes('text-[9px]')
                                    ui.element('td').style('padding: 5px 8px; text-align: center; border-bottom: 1px solid rgba(255,255,255,0.04); color: #81c784;').text(str(c.get('max_acompanhantes') or 0))

                ui.label(f'Exibindo {len(visible)} de {len(convidados)} registros').classes('text-[10px] text-grey-5 q-mt-xs')

            render_table()

            search_inp.on('update:model-value', lambda: render_table.refresh())
            filt_conf_sel.on('update:model-value', lambda: render_table.refresh())
            filt_placa_sel.on('update:model-value', lambda: render_table.refresh())

            ui.separator().classes('q-my-sm')

            # Botões de exportação
            async def export_csv():
                import csv, io as csvio
                output = csvio.StringIO()
                writer = csv.writer(output)
                writer.writerow(['Assento', 'Posto/Graduação', 'Nome Completo', 'Cargo/Função', 'Tipo', 'Status RSVP', 'Status Placa', 'Acompanhantes Previstos'])
                for c in sorted(convidados, key=lambda x: (x.get('assento_id') or 'ZZZ', x.get('nome', ''))):
                    sc_label, _ = STATUS_CONF_LABELS.get(c.get('status_confirmacao', 'pendente'), ('Pendente', ''))
                    sp_label, _ = STATUS_PLACA_LABELS.get(c.get('status_placa', 'pendente'), ('Pendente', ''))
                    writer.writerow([
                        c.get('assento_id') or '',
                        c.get('posto_graduacao') or '',
                        clean_authority_name(c.get('nome', '')),
                        c.get('cargo_funcao') or '',
                        'Acompanhante' if c.get('convidado_principal_id') else 'Titular',
                        sc_label,
                        sp_label,
                        str(c.get('max_acompanhantes') or 0),
                    ])
                csv_data = output.getvalue()
                encoded = base64.b64encode(csv_data.encode('utf-8-sig')).decode()
                nome_ev = (event.get('nome') or 'evento').replace(' ', '_')[:30]
                fname = f"planilhao_{nome_ev}_{event.get('data_evento','')}.csv"
                await ui.run_javascript(f'''
                    const a = document.createElement("a");
                    a.href = "data:text/csv;charset=utf-8-sig;base64,{encoded}";
                    a.download = "{fname}";
                    a.click();
                ''')
                ui.notify('📥 CSV exportado com sucesso!', color='positive', position='top')

            with ui.row().classes('w-full justify-end items-center gap-2'):
                ui.label(f'📋 {len(convidados)} registros totais').classes('text-xs text-grey-5 q-mr-auto')
                ui.button('🖨️ Imprimir Lista', icon='print', on_click=lambda: ui.run_javascript("window.print()")).props('unelevated color=light-blue-9 text-color=white dense bold').classes('text-xs')
                ui.button('📥 Exportar CSV', icon='download', on_click=export_csv).props('unelevated color=teal text-color=white dense bold').classes('text-xs')
                ui.button('Fechar', on_click=diag.close).props('unelevated color=grey-8 dense').classes('text-xs')

        diag.open()

    def open_tactical_scanner_dialog(event, convidados):

        scanned_history = []

        with ui.dialog() as diag, ui.card().classes('q-pa-lg').style('min-width: 620px; max-width: 95vw;'):
            ui.label('🔍 SCANNER & CONFERÊNCIA TÁTICA').classes('text-md font-bold text-amber cyber-title q-mb-xs')
            ui.label('Use a Câmera do Celular ou Leitor Físico para separar e bater a lista por Fileira').classes('text-xs text-grey-4 q-mb-md')
            
            with ui.row().classes('w-full gap-2 q-mb-md items-center'):
                scan_input = ui.input(placeholder='Bipe o QR Code ou digite ID/Nome/Assento (ex: G-5)...').props('dark outlined dense autofocus').classes('col')
                cam_btn = ui.button('📸 Câmera do Celular', icon='videocam').props('unelevated color=primary text-color=black dense').classes('text-xs')

            # Injeta biblioteca HTML5-QRCode no head
            ui.add_head_html('<script src="https://unpkg.com/html5-qrcode"></script>')

            # Container da Câmera do Celular (HTML5 QR Scanner)
            cam_container = ui.column().classes('w-full hidden q-mb-md border border-cyan-500/40 rounded-xl q-pa-sm bg-black/60')
            with cam_container:
                ui.label('📸 Aponta a câmera para o QR Code impresso no cartão:').classes('text-xs font-bold text-cyan text-center w-full q-mb-xs')
                ui.html('<div id="qr-reader" style="width:100%; max-width:400px; margin:0 auto; background:#000; border-radius:8px;"></div>').classes('w-full flex justify-center')
                ui.button('🛑 Fechar Câmera', on_click=lambda: cam_container.classes(add='hidden')).props('flat color=grey dense').classes('w-full text-xs q-mt-xs')

            def toggle_camera():
                cam_container.classes(remove='hidden')
                ui.run_javascript('''
                    if (window.html5QrcodeScanner) {
                        try { window.html5QrcodeScanner.clear(); } catch(e){}
                    }
                    window.html5QrcodeScanner = new Html5QrcodeScanner("qr-reader", { fps: 10, qrbox: {width: 250, height: 250} }, false);
                    window.html5QrcodeScanner.render(function(decodedText, decodedResult) {
                        let inp = document.querySelector('input[placeholder*="Bipe"]');
                        if (inp) {
                            inp.value = decodedText;
                            inp.dispatchEvent(new KeyboardEvent('keydown', {'key': 'Enter', 'keyCode': 13, 'bubbles': true}));
                        }
                    }, function(error) {});
                ''')

            cam_btn.on_click(toggle_camera)
            feedback_container = ui.column().classes('w-full')

            # Script de áudio para bipe sonoro tático
            audio_script = """
            function playBeep(type) {
                try {
                    let ctx = new (window.AudioContext || window.webkitAudioContext)();
                    let osc = ctx.createOscillator();
                    let gain = ctx.createGain();
                    osc.connect(gain);
                    gain.connect(ctx.destination);
                    if (type === 'success') {
                        osc.frequency.value = 880;
                        gain.gain.value = 0.12;
                        osc.start();
                        setTimeout(() => osc.stop(), 150);
                    } else if (type === 'duplicate') {
                        osc.frequency.value = 280;
                        gain.gain.value = 0.25;
                        osc.start();
                        setTimeout(() => osc.stop(), 320);
                    }
                } catch(e) {}
            }
            """
            ui.add_head_html(f"<script>{audio_script}</script>")

            def process_scan(val):
                if not val or not val.strip():
                    return
                query = val.strip().lower()
                scan_input.value = ''
                feedback_container.clear()
                
                # Busca convidado por ID exato do QR Code, Assento ou Nome
                matches = [
                    c for c in convidados
                    if str(c.get('id', '')) == query or
                    query == str(c.get('assento_id', '')).lower() or
                    query in c['nome'].lower() or
                    (c.get('posto_graduacao') and query in f"{c.get('posto_graduacao')} {c['nome']}".lower())
                ]

                with feedback_container:
                    if not matches:
                        ui.run_javascript("playBeep('duplicate')")
                        with ui.card().classes('w-full q-pa-md bg-red-950/80 border border-red-500 rounded-xl text-center'):
                            ui.icon('cancel', color='red', size='3rem')
                            ui.label('❌ CARTÃO / CONVIDADO NÃO ENCONTRADO!').classes('text-sm font-bold text-red-3')
                            ui.label(f"Nenhum assento ou convidado registrado para o código: '{val}'").classes('text-xs text-grey-4')
                    else:
                        target = matches[0]
                        seat = target.get('assento_id', 'NÃO ALOCADO')
                        row_name = seat.split('-')[0] if '-' in seat else 'N/A'
                        is_duplicate = target['id'] in scanned_history

                        if is_duplicate:
                            ui.run_javascript("playBeep('duplicate')")
                            with ui.card().classes('w-full q-pa-md bg-amber-950/80 border border-amber-500 rounded-xl text-center'):
                                ui.icon('warning', color='amber', size='3rem')
                                ui.label('⚠️ ATENÇÃO: CARTÃO JÁ CONFERIDO & SEPARADO!').classes('text-sm font-bold text-amber-3')
                                ui.label(f"{target['nome']} — Assento: {seat} (Fileira {row_name})").classes('text-xs text-grey-3 font-bold')
                                ui.label('Este cartão já passou pela triagem anteriormente.').classes('text-[11px] text-amber-4 q-mt-xs')
                        else:
                            scanned_history.append(target['id'])
                            ui.run_javascript("playBeep('success')")
                            
                            main_info = ""
                            if target.get('convidado_principal_id'):
                                p_main = next((x for x in convidados if x['id'] == target['convidado_principal_id']), None)
                                if p_main:
                                    main_info = f" (Acompanhante de {p_main['nome']})"

                            with ui.card().classes('w-full q-pa-md bg-green-950/80 border border-green-500 rounded-xl text-center'):
                                ui.icon('check_circle', color='green', size='3rem')
                                ui.label('✅ CONFERIDO & SEPARADO COM SUCESSO!').classes('text-sm font-bold text-green-3')
                                ui.label(f"{target.get('posto_graduacao') or ''} {target['nome']}{main_info}").classes('text-md font-bold text-white')
                                
                                with ui.row().classes('w-full justify-center items-center gap-2 q-mt-xs wrap'):
                                    ui.badge(f"COLOCAR NA FILEIRA {row_name}").props('color=cyan text-color=black bold').classes('text-sm q-px-sm')
                                    ui.badge(f"ASSENTO {seat}").props('color=green text-color=white bold').classes('text-sm q-px-sm')
                                    
                                    async def remap_seat_dialog(target_guest):
                                        with ui.dialog() as r_diag, ui.card().classes('q-pa-md bg-cyan-950 border border-cyan-400 rounded-xl').style('min-width: 320px;'):
                                            ui.label('🔀 REMAPEAR ASSENTO NO SALÃO').classes('text-xs font-bold text-cyan q-mb-xs')
                                            ui.label(f"Convidado: {target_guest['nome']}").classes('text-[11px] text-grey-3 q-mb-sm')
                                            new_seat_input = ui.input(label='Novo Código de Assento (ex: A-1, F-5)', value=target_guest.get('assento_id', '')).props('dark outlined dense').classes('w-full q-mb-md')
                                            
                                            async def confirm_remap():
                                                val_seat = new_seat_input.value.strip().upper()
                                                if val_seat:
                                                    _db = get_service_db_connection() or get_db_connection()
                                                    if _db:
                                                        _db.table('jade_convidados').update({'assento_id': val_seat}).eq('id', target_guest['id']).execute()
                                                        target_guest['assento_id'] = val_seat
                                                        ui.notify(f"✅ Assento remapeado para {val_seat}!", color='positive')
                                                        r_diag.close()
                                                        process_scan(target_guest['id'])
                                                        render_content.refresh()

                                            with ui.row().classes('w-full justify-end gap-2'):
                                                ui.button('Cancelar', on_click=r_diag.close).props('flat dense color=grey')
                                                ui.button('Salvar Remapeamento', on_click=confirm_remap).props('unelevated color=cyan text-color=black dense bold')
                                        r_diag.open()

                                    ui.button('🔀 Mudar Assento', on_click=lambda tg=target: remap_seat_dialog(tg)).props('unelevated color=amber text-color=black dense bold').classes('text-xs').tooltip('Trocar assento desta placa no salão instantaneamente')

            scan_input.on('keydown.enter', lambda: process_scan(scan_input.value))

            with ui.row().classes('w-full justify-between items-center q-mt-md'):
                ui.label(f"Total Conferidos: {len(scanned_history)} de {len([c for c in convidados if c.get('assento_id')])} cartões").classes('text-xs text-cyan font-bold')
                ui.button('Concluir Conferência', on_click=diag.close).props('unelevated color=grey-8 dense')
        diag.open()

    def open_field_assembly_report_dialog(event, convidados):
        """Gera relatório de prancheta de campo para a equipe de montagem de assentos no salão."""
        with ui.dialog() as diag, ui.card().classes('q-pa-lg').style('min-width: 800px; max-width: 95vw; max-height: 90vh; overflow-y: auto;'):
            with ui.row().classes('w-full justify-between items-center q-mb-md print-hide'):
                with ui.row().classes('items-center gap-2'):
                    ui.icon('assignment', size='md', color='cyan')
                    ui.label('📋 RELATÓRIO DE PRANCHETA DE MONTAGEM DO SALÃO').classes('text-md font-bold text-cyan cyber-title')
                with ui.row().classes('items-center gap-2'):
                    ui.button('🖨️ Imprimir Prancheta', icon='print', on_click=lambda: ui.run_javascript('window.print()')).props('unelevated color=cyan text-color=black dense bold').classes('text-xs')
                    ui.button(icon='close', on_click=diag.close).props('flat round dense color=grey-4')

            with ui.column().classes('w-full print-area q-pa-sm'):
                allocated = [c for c in convidados if c.get('assento_id')]
                
                def sort_key_assento(c):
                    ass = str(c.get('assento_id', '')).upper().strip()
                    match = re.match(r'([A-Z]+)-?(\d+)', ass)
                    if match:
                        row, num = match.groups()
                        return (row, int(num))
                    return (ass, 0)

                allocated.sort(key=sort_key_assento)

                rows_html = ""
                for idx, c in enumerate(allocated, 1):
                    assento = c.get('assento_id', 'N/I')
                    posto = c.get('posto_graduacao', '') or ''
                    nome = c.get('nome', '')
                    is_acomp = bool(c.get('convidado_principal_id'))
                    tipo_placa = "RESERVADO (Acompanhante)" if is_acomp else "Titular"
                    st = c.get('status_placa', 'pendente').upper()
                    bg_row = "#f9f9f9" if idx % 2 == 0 else "#ffffff"

                    rows_html += f'''
                    <tr style="background: {bg_row};">
                        <td style="padding: 6px; border: 1px solid #ccc; text-align: center; font-size: 14pt;">☐</td>
                        <td style="padding: 6px; border: 1px solid #ccc; font-weight: bold; color: #1f4e79;">{assento}</td>
                        <td style="padding: 6px; border: 1px solid #ccc; font-weight: bold;">{posto} {clean_authority_name(nome)}</td>
                        <td style="padding: 6px; border: 1px solid #ccc;">{tipo_placa}</td>
                        <td style="padding: 6px; border: 1px solid #ccc; font-size: 8pt;">{st}</td>
                    </tr>
                    '''

                ui.html(f'''
                <div style="font-family: Arial, sans-serif; color: #000; background: #fff; padding: 16px; border-radius: 8px;">
                    <div style="display: flex; justify-content: space-between; align-items: center; border-bottom: 2px solid #000; padding-bottom: 8px; margin-bottom: 12px;">
                        <div>
                            <h2 style="margin: 0; font-size: 16pt; font-weight: bold; text-transform: uppercase;">📋 JADE — RELATÓRIO DE MONTAGEM DE ASSENTOS DE CAMPO</h2>
                            <div style="font-size: 10pt; color: #444;">Solenidade: <strong>{event.get('nome','')}</strong> | Data: <strong>{event.get('data_evento','N/I')}</strong></div>
                        </div>
                        <div style="text-align: right; font-size: 9pt;">
                            <div><strong>COMSOC / CERIMONIAL</strong></div>
                            <div>Total: {len(allocated)} Assentos</div>
                        </div>
                    </div>

                    <table style="width: 100%; border-collapse: collapse; font-size: 10pt;">
                        <thead>
                            <tr style="background: #1f4e79; color: #fff; text-align: left;">
                                <th style="padding: 6px; border: 1px solid #333; width: 40px; text-align: center;">[  ]</th>
                                <th style="padding: 6px; border: 1px solid #333; width: 90px;">Assento</th>
                                <th style="padding: 6px; border: 1px solid #333;">Posto / Graduação e Nome da Autoridade</th>
                                <th style="padding: 6px; border: 1px solid #333; width: 140px;">Tipo de Placa</th>
                                <th style="padding: 6px; border: 1px solid #333; width: 110px;">Status Placa</th>
                            </tr>
                        </thead>
                        <tbody>
                            {rows_html}
                        </tbody>
                    </table>

                    <div style="margin-top: 20px; font-size: 9pt; display: flex; justify-content: space-between; border-top: 1px solid #999; padding-top: 8px;">
                        <div>Responsável pela Montagem: _____________________________________</div>
                        <div>Visto do Chefe do Cerimonial: _____________________________________</div>
                    </div>
                </div>
                ''')

        diag.open()

    def download_template():
        """Gera o modelo oficial JADE com colunas compatíveis com o importador inteligente por blocos."""
        try:
            output = io.BytesIO()
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Protection
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'LISTA DE CONVITE'

            # ── Cores de cabeçalho ──
            fill_header = PatternFill('solid', fgColor='1F4E79')
            fill_bloco  = PatternFill('solid', fgColor='D6E4F7')
            fill_conf   = PatternFill('solid', fgColor='00B050')   # Verde = Confirmado
            fill_rec    = PatternFill('solid', fgColor='C00000')   # Vermelho = Recusado
            fill_pend   = PatternFill('solid', fgColor='FFFFFF')   # Branco = Pendente
            font_header = Font(bold=True, color='FFFFFF', size=10)
            font_bloco  = Font(bold=True, color='1F4E79', size=10)
            font_data   = Font(size=10)
            align_c     = Alignment(horizontal='center', vertical='center', wrap_text=True)
            align_l     = Alignment(horizontal='left', vertical='center', wrap_text=True)
            thin        = Side(style='thin', color='BFBFBF')
            border      = Border(left=thin, right=thin, top=thin, bottom=thin)

            COLUNAS = [
                ('QTD / BLOCO', 8),
                ('POSTO / GRAD', 12),
                ('NOME', 35),
                ('CARGO / FUNÇÃO', 28),
                ('ENVIADO', 10),
                ('CONFIRMADO\n(verde=conf, vermelho=rec)', 14),
                ('CÔNJUGE\n(E = sim)', 10),
                ('TELEFONE', 18),
                ('E-MAIL', 30),
                ('Nº ANTIGUIDADE', 12),
            ]

            # Linha 1: Título geral
            ws.merge_cells('A1:J1')
            titulo = ws['A1']
            titulo.value = 'CONVIDADOS — SOLENIDADE [NOME DO EVENTO]'
            titulo.font = Font(bold=True, color='FFFFFF', size=12)
            titulo.fill = PatternFill('solid', fgColor='0D2B45')
            titulo.alignment = align_c
            ws.row_dimensions[1].height = 22

            # Linha 2: cabeçalho das colunas
            for col_idx, (col_name, col_width) in enumerate(COLUNAS, start=1):
                cell = ws.cell(row=2, column=col_idx, value=col_name)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_c
                cell.border = border
                ws.column_dimensions[get_column_letter(col_idx)].width = col_width
            ws.row_dimensions[2].height = 30

            # Linha 3: nome do bloco/categoria (lida automaticamente pelo importador)
            ws.merge_cells('A3:J3')
            bloco = ws['A3']
            bloco.value = 'EX-COMANDANTES DA MARINHA'
            bloco.font = font_bloco
            bloco.fill = fill_bloco
            bloco.alignment = align_c
            ws.row_dimensions[3].height = 18

            # Dados de exemplo
            exemplos = [
                (1, 'AE (Ref)', 'MAURO RODRIGUES PEREIRA',    'EX-MINISTRO DA MARINHA',    'E', 1,  'E', '(21) 99964-5831', 'exemplo@terra.com.br', 1, 'conf'),
                (2, 'AE (Ref)', 'ROBERTO DE GUIMARÃES CARVALHO', 'EX-COMANDANTE DA MARINHA', 'E', '', 'E', '(21) 99872-2458', 'exemplo@gmail.com',    2, 'rec'),
                (3, 'VA',       'CARLOS SOUZA NEVES',          'VICE-CHEFE DO ESTADO-MAIOR', '',  '', '',  '',                '',                     3, 'pend'),
            ]

            for row_num, (qtd, posto, nome, cargo, env, conf, conj, tel, email, ant, status) in enumerate(exemplos, start=4):
                dados = [qtd, posto, nome, cargo, env, conf, conj, tel, email, ant]
                for col_idx, val in enumerate(dados, start=1):
                    cell = ws.cell(row=row_num, column=col_idx, value=val)
                    cell.font = font_data
                    cell.alignment = align_l if col_idx in (3, 4, 8, 9) else align_c
                    cell.border = border
                # Cor da célula de confirmação (coluna F = índice 6)
                conf_cell = ws.cell(row=row_num, column=6)
                if status == 'conf':
                    conf_cell.fill = fill_conf
                    conf_cell.font = Font(bold=True, color='FFFFFF', size=10)
                elif status == 'rec':
                    conf_cell.fill = fill_rec
                    conf_cell.font = Font(bold=True, color='FFFFFF', size=10)
                else:
                    conf_cell.fill = fill_pend
                ws.row_dimensions[row_num].height = 16

            # Linha de instrução final
            inst_row = len(exemplos) + 4 + 1
            ws.merge_cells(f'A{inst_row}:J{inst_row}')
            inst = ws.cell(row=inst_row, column=1,
                value='💡 INSTRUÇÕES: Coluna A = número ou nome do bloco/categoria. Coluna F = confirmação (célula VERDE = confirmado, VERMELHA = recusado, BRANCA/VAZIA = pendente). '
                      'Coluna G = cônjuge/acompanhante (E = sim). Coluna J = número de antiguidade na lista. '
                      'Não altere a estrutura das colunas. O sistema lê blocos por categorias automaticamente.')
            inst.font = Font(italic=True, size=9, color='595959')
            inst.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            ws.row_dimensions[inst_row].height = 40

            # Congelar cabeçalho
            ws.freeze_panes = 'A3'

            wb.save(output)
            output.seek(0)
            ui.download(output.read(), 'modelo_oficial_jade_convite.xlsx')
            ui.notify('📥 Modelo oficial JADE baixado com sucesso!', color='positive')
        except Exception as ex:
            ui.notify(f'Erro ao gerar modelo: {ex}', color='negative')
            print(f'[TEMPLATE ERR] {ex}')

    async def handle_import_list(e, event_id):
        try:
            import inspect, io
            file_obj = getattr(e, 'file', None)
            if not file_obj and hasattr(e, 'files') and e.files:
                file_obj = e.files[0]
            
            if not file_obj:
                ui.notify('❌ Nenhum arquivo de planilha detectado.', color='negative')
                return

            file_bytes = file_obj.read() if hasattr(file_obj, 'read') else getattr(file_obj, 'content', None)
            if inspect.isawaitable(file_bytes):
                file_bytes = await file_bytes
            elif hasattr(file_bytes, 'read'):
                file_bytes = file_bytes.read()

            if not file_bytes:
                ui.notify('❌ Arquivo de planilha vazio.', color='negative')
                return

            file_name = getattr(file_obj, 'name', 'planilha.xlsx').lower()

            db = get_service_db_connection() or get_db_connection()
            if not db:
                ui.notify('❌ Banco de dados indisponível.', color='negative')
                return
            
            res_exist = db.table('jade_convidados').select('*').eq('evento_id', event_id).execute()
            existing_list = res_exist.data if res_exist.data else []
            
            existing_map = {}
            for item in existing_list:
                if not item.get('convidado_principal_id'):
                    key = f"{(item.get('nome') or '').strip().upper()}|{(item.get('posto_graduacao') or '').strip().upper()}"
                    existing_map[key] = item

            count_inserted = 0
            count_updated = 0

            wb = None
            if file_name.endswith('.xlsx') or file_name.endswith('.xlsm'):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
                except Exception:
                    wb = None

            if wb:
                sheet = wb.active
                current_category = "Geral"
                for r in range(1, sheet.max_row + 1):
                    row_cells = [sheet.cell(row=r, column=c) for c in range(1, sheet.max_column + 1)]
                    row_vals = [c.value for c in row_cells]

                    if not any(v is not None for v in row_vals):
                        continue

                    c0 = str(row_vals[0]).strip() if len(row_vals) > 0 and row_vals[0] is not None else ""
                    c1 = str(row_vals[1]).strip() if len(row_vals) > 1 and row_vals[1] is not None else ""
                    c2 = str(row_vals[2]).strip() if len(row_vals) > 2 and row_vals[2] is not None else ""

                    if c0.upper() in ('QTD', 'QUANTIDADE', 'QTD / BLOCO') or c1.upper() in ('POSTO', 'GRAU', 'POSTO / GRAD'):
                        continue

                    if str(row_vals[0]).startswith('CONVIDADOS') or str(row_vals[0]).startswith('💡 INSTRUÇÕES'):
                        continue

                    if c0 and not c0.isdigit() and c0.upper() not in ('QTD', 'QUANTIDADE', 'QTD / BLOCO'):
                        current_category = c0
                        continue

                    nome = c2
                    if not nome or nome.upper() in ('NOME', 'CONVITE', 'NONE', 'NAN'):
                        continue

                    posto = c1
                    cargo = str(row_vals[3]).strip() if len(row_vals) > 3 and row_vals[3] is not None else ""
                    conf_val = row_vals[5] if len(row_vals) > 5 else None
                    conjuge_val = row_vals[6] if len(row_vals) > 6 else None
                    ant_val = row_vals[9] if len(row_vals) > 9 else None
                    try:
                        numero_antiguidade = int(float(str(ant_val))) if ant_val not in (None, '', 'None') else None
                    except (ValueError, TypeError):
                        numero_antiguidade = None

                    conf_fill = ""
                    if len(row_cells) > 5 and row_cells[5].fill and row_cells[5].fill.start_color:
                        conf_fill = str(row_cells[5].fill.start_color.rgb or "")

                    status_conf = "pendente"
                    status_placa = "nao_necessaria"

                    if any(g in conf_fill for g in ('00B050', '92D050', '00FF00')):
                        status_conf = "confirmado"
                        status_placa = "pendente"
                    elif any(rc in conf_fill for rc in ('FF0000', 'FA1717', 'C00000', 'FF5555')):
                        status_conf = "recusado"
                        status_placa = "nao_necessaria"
                    else:
                        if conf_val in (1, 2, '1', '2', 'SIM', 'Sim', 'sim', 'CONFIRMADO', 'conf', 'CONF'):
                            status_conf = "confirmado"
                            status_placa = "pendente"

                    max_acomp = 0
                    if isinstance(conf_val, int) and conf_val > 1:
                        max_acomp = conf_val - 1
                    elif str(conf_val).strip().isdigit() and int(str(conf_val).strip()) > 1:
                        max_acomp = int(str(conf_val).strip()) - 1
                    elif conjuge_val in ('E', '1', 1) or (isinstance(conjuge_val, str) and len(conjuge_val) > 2):
                        max_acomp = 1

                    key = f"{nome.strip().upper()}|{posto.strip().upper()}"
                    
                    conv_data = {
                        'evento_id': event_id,
                        'nome': nome,
                        'posto_graduacao': posto,
                        'cargo_funcao': cargo,
                        'categoria': current_category,
                        'status_confirmacao': status_conf,
                        'max_acompanhantes': max_acomp,
                    }
                    if numero_antiguidade is not None:
                        conv_data['numero_antiguidade'] = numero_antiguidade
                    
                    conv_data_with_placa = dict(conv_data)
                    conv_data_with_placa['status_placa'] = status_placa

                    if key in existing_map:
                        existing_id = existing_map[key]['id']
                        try:
                            db.table('jade_convidados').update(conv_data_with_placa).eq('id', existing_id).execute()
                        except Exception:
                            safe = {k: v for k, v in conv_data.items() if k != 'numero_antiguidade'}
                            db.table('jade_convidados').update(safe).eq('id', existing_id).execute()
                        
                        sync_companions(existing_id, nome, max_acomp, event_id, current_category)
                        count_updated += 1
                    else:
                        res_ins = None
                        try:
                            res_ins = db.table('jade_convidados').insert(conv_data_with_placa).execute()
                        except Exception:
                            safe = {k: v for k, v in conv_data.items() if k != 'numero_antiguidade'}
                            safe['status_placa'] = status_placa
                            res_ins = db.table('jade_convidados').insert(safe).execute()
                        
                        if res_ins and res_ins.data:
                            new_id = res_ins.data[0]['id']
                            sync_companions(new_id, nome, max_acomp, event_id, current_category)
                        count_inserted += 1

            else:
                # Leitor Pandas Universal (.csv, .tsv, .txt, .ods, .xls)
                bio_p = io.BytesIO(file_bytes)
                df_p = None
                if file_name.endswith('.csv') or file_name.endswith('.tsv') or file_name.endswith('.txt'):
                    for sep in [';', ',', '\t', '|']:
                        try:
                            bio_p.seek(0)
                            df_test = pd.read_csv(bio_p, sep=sep, encoding='utf-8-sig', dtype=str)
                            if len(df_test.columns) > 1:
                                df_p = df_test
                                break
                        except Exception:
                            pass
                    if df_p is None:
                        bio_p.seek(0)
                        df_p = pd.read_csv(bio_p, encoding='utf-8-sig', dtype=str)
                elif file_name.endswith('.xls'):
                    try:
                        df_p = pd.read_excel(bio_p, engine='xlrd', dtype=str)
                    except Exception:
                        df_p = pd.read_excel(bio_p, dtype=str)
                elif file_name.endswith('.ods'):
                    try:
                        df_p = pd.read_excel(bio_p, engine='odf', dtype=str)
                    except Exception:
                        df_p = pd.read_excel(bio_p, dtype=str)
                else:
                    df_p = pd.read_excel(bio_p, dtype=str)

                if df_p is not None:
                    if 'Nome' not in df_p.columns and 'NOME' not in [str(c).upper() for c in df_p.columns]:
                        for r_idx in range(min(5, len(df_p))):
                            row_str = [str(val).upper() for val in df_p.iloc[r_idx].values]
                            if any(k in row_str for k in ('NOME', 'AUTORIDADE', 'CONVIDADO')):
                                df_p.columns = df_p.iloc[r_idx]
                                df_p = df_p.iloc[r_idx + 1:]
                                break

                    col_map = {}
                    for col in df_p.columns:
                        c_clean = str(col).strip().upper()
                        if c_clean in ('NOME', 'NOME COMPLETO', 'AUTORIDADE', 'NOME DA AUTORIDADE', 'NOME DO CONVIDADO', 'CONVIDADO'):
                            col_map[col] = 'Nome'
                        elif c_clean in ('POSTO', 'POSTO/GRADUAÇÃO', 'POSTO/GRADUACAO', 'POSTO / GRADUAÇÃO', 'POSTO / GRAD', 'GRADUAÇÃO', 'GRADUACAO', 'POSTO_GRADUACAO'):
                            col_map[col] = 'Posto/Graduacao'
                        elif c_clean in ('CARGO', 'CARGO/FUNÇÃO', 'CARGO/FUNCAO', 'CARGO / FUNÇÃO', 'FUNÇÃO', 'FUNCAO', 'TÍTULO', 'TITULO', 'CARGO_FUNCAO'):
                            col_map[col] = 'Cargo/Função'
                        elif c_clean in ('CATEGORIA', 'SETOR', 'BLOCO', 'GRUPO', 'QTD / BLOCO'):
                            col_map[col] = 'Categoria'
                        elif c_clean in ('MAX ACOMPANHANTES', 'ACOMPANHANTES', 'ACOMP', 'CÔNJUGE', 'CÔNJUGE\n(E = sim)', 'N_ACOMPANHANTES', 'QTD ACOMPANHANTES', 'QTD_ACOMP'):
                            col_map[col] = 'Max Acompanhantes'
                        elif c_clean in ('ANTIGUIDADE', 'Nº ANTIGUIDADE', 'NUMERO_ANTIGUIDADE'):
                            col_map[col] = 'Antiguidade'
                    
                    df_p = df_p.rename(columns=col_map)
                    
                    if 'Nome' not in df_p.columns and len(df_p.columns) > 0:
                        df_p = df_p.rename(columns={df_p.columns[0]: 'Nome'})

                    for _, row in df_p.iterrows():
                        nome_p = str(row.get('Nome', '')).strip().upper()
                        if not nome_p or nome_p in ('NAN', 'NONE', 'NOME', 'NOME COMPLETO'):
                            continue
                        posto_p = str(row.get('Posto/Graduacao', '')).strip() if pd.notna(row.get('Posto/Graduacao')) and str(row.get('Posto/Graduacao')) != 'nan' else ''
                        cargo_p = str(row.get('Cargo/Função', '')).strip() if pd.notna(row.get('Cargo/Função')) and str(row.get('Cargo/Função')) != 'nan' else ''
                        cat_p = str(row.get('Categoria', 'Geral')).strip() if pd.notna(row.get('Categoria')) and str(row.get('Categoria')) != 'nan' else 'Geral'
                        try:
                            acomp_p = int(row.get('Max Acompanhantes', 0)) if pd.notna(row.get('Max Acompanhantes')) else 0
                        except Exception:
                            acomp_p = 0

                        ant_p = None
                        try:
                            if pd.notna(row.get('Antiguidade')):
                                ant_p = int(float(str(row.get('Antiguidade'))))
                        except Exception:
                            ant_p = None

                        key_p = f"{nome_p}|{posto_p}"
                        guest_data = {
                            'evento_id': event_id,
                            'nome': nome_p,
                            'posto_graduacao': posto_p,
                            'cargo_funcao': cargo_p,
                            'categoria': cat_p,
                            'status_confirmacao': 'confirmado',
                            'status_placa': 'pendente',
                            'max_acompanhantes': acomp_p
                        }
                        if ant_p is not None:
                            guest_data['numero_antiguidade'] = ant_p

                        if key_p in existing_map:
                            e_id = existing_map[key_p]['id']
                            try:
                                db.table('jade_convidados').update(guest_data).eq('id', e_id).execute()
                            except Exception:
                                safe = {k: v for k, v in guest_data.items() if k != 'numero_antiguidade'}
                                db.table('jade_convidados').update(safe).eq('id', e_id).execute()
                            sync_companions(e_id, nome_p, acomp_p, event_id, cat_p)
                            count_updated += 1
                        else:
                            res_p = None
                            try:
                                res_p = db.table('jade_convidados').insert(guest_data).execute()
                            except Exception:
                                safe = {k: v for k, v in guest_data.items() if k != 'numero_antiguidade'}
                                res_p = db.table('jade_convidados').insert(safe).execute()
                            if res_p and res_p.data:
                                new_id = res_p.data[0]['id']
                                sync_companions(new_id, nome_p, acomp_p, event_id, cat_p)
                            count_inserted += 1

            ui.notify(f"✅ Importados {count_inserted} novos convidados ({count_updated} atualizados) com sucesso!", color='positive')
            render_content.refresh()
        except Exception as ex:
            print(f"[HANDLE IMPORT ERR] {ex}")
            ui.notify(f"Erro ao importar planilha: {ex}", color='red')

    render_content()

